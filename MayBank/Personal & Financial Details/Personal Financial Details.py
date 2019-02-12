from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)

path = "D:\\Automation Testing\\MayBank\\Personal & Financial Details\\Personal&Financial_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
Excel_Dictionary2 = {}
for i in range(1,110):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value
    Excel_Dictionary2[sheet.cell(i,1).value] = sheet.cell(i,3).value

driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_id("AddLoanApplication").click()
driver.find_element_by_id("homeBuyerSearchInputId").send_keys("A K S")
driver.find_element_by_id("homeBuyerSearchButtonId").click()
driver.find_element_by_id("selectedClientId").click()
driver.find_element_by_xpath('//*[@id="homeBuyerDetailsModelDiv"]/div/div/div[2]/div[3]/div[2]/button').click()
driver.find_element_by_id("singleApplicantID").click()
# driver.find_element_by_id("privateFlatID").click()
# driver.find_element_by_id("newHomeLoanImage").click()
# driver.find_element_by_xpath("(.//*[normalize-space(text()) and normalize-space(.)='Cancel'])[2]/following::button[2]").click()

if Excel_Dictionary1['application type 1'] != None:
    if Excel_Dictionary1['application type 1'] == 'Private':
        driver.find_element_by_id("privateFlatID").click()
    elif Excel_Dictionary1['application type 1'] == 'Executive Condo':
        driver.find_element_by_id("executiveCondoID").click()
    elif Excel_Dictionary1['application type 1'] == 'HDB':
        driver.find_element_by_id('hdbFlatID').click()        
    else:
        driver.find_element_by_id("privateFlatID").click()
else:
    driver.find_element_by_id("privateFlatID").click() 

if Excel_Dictionary1['application type 2'] != None:
    if Excel_Dictionary1['application type 2'] == 'New':
        driver.find_element_by_id("newHomeLoanImage").click()
    elif Excel_Dictionary1['application type 2'] == 'Refinance':
        driver.find_element_by_id("refinanceID").click()    
    else:
        driver.find_element_by_id("newHomeLoanImage").click()
else:
    driver.find_element_by_id("newHomeLoanImage").click()

driver.find_element_by_xpath("(.//*[normalize-space(text()) and normalize-space(.)='Cancel'])[2]/following::button[2]").click()

if Excel_Dictionary1['Salutation'] != None:
    driver.find_element_by_xpath('//*[@id="salutation"]').send_keys(Excel_Dictionary1['Salutation'])
if Excel_Dictionary1['Full Name as in NRIC/PASSPORT'] != None:     
    driver.find_element_by_xpath('//*[@id="fullName"]').clear()
    driver.find_element_by_xpath('//*[@id="fullName"]').send_keys(Excel_Dictionary1['Full Name as in NRIC/PASSPORT'])
if Excel_Dictionary1['Date of Birth'] != None:
    driver.find_element_by_name('M0009').clear()
    driver.find_element_by_name('M0009').send_keys(Excel_Dictionary1['Date of Birth'])
if Excel_Dictionary1['Country Of Birth'] != None:     
    driver.find_element_by_name('M0413').send_keys(Excel_Dictionary1['Country Of Birth'])
if Excel_Dictionary1['Race'] != None:
    driver.find_element_by_name('M0007').send_keys(Excel_Dictionary1['Race'])
    if Excel_Dictionary1['Race'] == 'Others':
        if Excel_Dictionary1['Other Race'] != None:
            driver.find_element_by_id('otherRace').send_keys(Excel_Dictionary1['Other Race'])
if Excel_Dictionary1['Marital Status'] != None:
    driver.find_element_by_name('M0005').send_keys(Excel_Dictionary1['Marital Status'])
    if Excel_Dictionary1['Marital Status'] == 'Others':
        if Excel_Dictionary1['Other Marital Status'] != None:
            driver.find_element_by_id('SOtherMaritalStatus').send_keys(Excel_Dictionary1['Other Marital Status'])





if Excel_Dictionary1['Nationality'] != None:
    driver.find_element_by_name('M0015').send_keys(Excel_Dictionary1['Nationality'])
if Excel_Dictionary1['Additional Nationality 1'] != None:
    driver.find_element_by_xpath('//*[@id="addothernationalityID"]/span').click()
    driver.find_element_by_id('additionalNationality').send_keys(Excel_Dictionary1['Additional Nationality 1'])
if Excel_Dictionary1['Additional Nationality 2'] != None:
    driver.find_element_by_xpath('//*[@id="addothernationalityID"]/span').click()
    driver.find_element_by_id('additionalNationality1').send_keys(Excel_Dictionary1['Additional Nationality 2'])
if Excel_Dictionary1['Singapore PR'] != None:  
    driver.find_element_by_xpath('//*[@id="option3"]').click()
if Excel_Dictionary1['Passport No.'] != None:
    driver.find_element_by_name('M0010').send_keys(Excel_Dictionary1['Passport No.'])
if Excel_Dictionary1['Issue Date'] != None:     
    driver.find_element_by_name('M0011').send_keys(Excel_Dictionary1['Issue Date'])
if Excel_Dictionary1['Expiry Date'] != None:     
    driver.find_element_by_name('M0012').send_keys(Excel_Dictionary1['Expiry Date'])
if Excel_Dictionary1['Country Of Issue'] != None:    
    driver.find_element_by_name('M0013').send_keys(Excel_Dictionary1['Country Of Issue'])
if Excel_Dictionary1['Previous Passport No.'] != None:
    driver.find_element_by_name('M0014').send_keys(Excel_Dictionary1['Previous Passport No.'])




if Excel_Dictionary1['Mother Maiden Name'] != None:    
    driver.find_element_by_name('M0018').send_keys(Excel_Dictionary1['Mother Maiden Name'])
if Excel_Dictionary1['Highest Education'] != None:
    driver.find_element_by_name('M0016').send_keys(Excel_Dictionary1['Highest Education'])
if Excel_Dictionary1['No. Of Dependents'] != None:    
    driver.find_element_by_name('M0019').send_keys(Excel_Dictionary1['No. Of Dependents'])
if Excel_Dictionary1['Age Of Dependent'] != None:
    driver.find_element_by_id('ageOfDependent1').send_keys(Excel_Dictionary1['Age Of Dependent'])
driver.find_element_by_xpath('//*[@id="singlePersoanlDetails"]/div[1]/div[2]/div[9]').click()
if Excel_Dictionary1['Email Address'] != None:   
    driver.find_element_by_name('M0031').send_keys(Excel_Dictionary1['Email Address'])
if Excel_Dictionary1['MCountry Code'] != None:
    Select(driver.find_element_by_id('ShandPhoneCountryCode')).select_by_visible_text(Excel_Dictionary1['MCountry Code'])
if Excel_Dictionary1['Mobile Number'] != None:    
    driver.find_element_by_name('M0030').send_keys(Excel_Dictionary1['Mobile Number'])
if Excel_Dictionary1['HCountry Code'] != None:
    Select(driver.find_element_by_id('ShomePhoneCountryCode')).select_by_visible_text(Excel_Dictionary1['HCountry Code'])
if Excel_Dictionary1['Home Phone'] != None:
    driver.find_element_by_name('M0028').send_keys(Excel_Dictionary1['Home Phone'])
if Excel_Dictionary1['OCountry Code'] != None:
    Select(driver.find_element_by_id('SofficePhoneCountryCode')).select_by_visible_text(Excel_Dictionary1['OCountry Code'])
if Excel_Dictionary1['Office Phone'] != None:    
    driver.find_element_by_name('M0029').send_keys(Excel_Dictionary1['Office Phone'])



if Excel_Dictionary1['Country Of Residence'] != None:    
    driver.find_element_by_name('M0021').send_keys(Excel_Dictionary1['Country Of Residence'])
    if Excel_Dictionary1['Country Of Residence'] == 'Singapore':
        if Excel_Dictionary1['Zip / Postal Code'] != None:
            driver.find_element_by_id('SHomePostalCode').send_keys(Excel_Dictionary1['Zip / Postal Code'])
        if Excel_Dictionary1['Residential Address'] != None:
            driver.find_element_by_id('homeAddress').send_keys(Excel_Dictionary1['Residential Address'])
        if Excel_Dictionary1['Unit No'] != None:
            driver.find_element_by_id('selfResidentialUnitNo').send_keys(Excel_Dictionary1['Unit No'])  
    else:
        if Excel_Dictionary1['Residential Address Line 1'] != None:    
            driver.find_element_by_id('homeAddressCopy').send_keys(Excel_Dictionary1['Residential Address Line 1'])
        if Excel_Dictionary1['Residential Address Line 2'] != None:
            driver.find_element_by_id('homeAddressLine2').send_keys(Excel_Dictionary1['Residential Address Line 2'])
        if Excel_Dictionary1['City'] != None:    
            driver.find_element_by_name('M0485').send_keys(Excel_Dictionary1['City'])
        if Excel_Dictionary1['State'] != None:    
            driver.find_element_by_name('M0486').send_keys(Excel_Dictionary1['State'])
        if Excel_Dictionary1['Zip / Postal Code'] != None:    
            driver.find_element_by_id('SHomePostalCodeCopy').send_keys(Excel_Dictionary1['Zip / Postal Code'])

if Excel_Dictionary1['Country Of Office'] != None:
    driver.find_element_by_name('M0454').send_keys(Excel_Dictionary1['Country Of Office'])
    if Excel_Dictionary1['Country Of Office'] == 'Singapore':
        if Excel_Dictionary1['OZip / Postal Code'] != None:    
            driver.find_element_by_id('SOfficePostalCode').send_keys(Excel_Dictionary1['OZip / Postal Code'])
        if Excel_Dictionary1['Office Address'] != None:
            driver.find_element_by_id('officeAddress').send_keys(Excel_Dictionary1['Office Address'])    
        if Excel_Dictionary1['OUnit No'] != None:
            driver.find_element_by_id('selfOfficeUnitNo').send_keys(Excel_Dictionary1['OUnit No'])    
    else:
        if Excel_Dictionary1['Office Address Line 1'] != None:    
            driver.find_element_by_id('officeAddressCopy').send_keys(Excel_Dictionary1['Office Address Line 1'])
        if Excel_Dictionary1['Office Address Line 2'] != None:    
            driver.find_element_by_id('selfOfficeAddressLine2').send_keys(Excel_Dictionary1['Office Address Line 2'])
        if Excel_Dictionary1['Ocity'] != None:    
            driver.find_element_by_name('M0489').send_keys(Excel_Dictionary1['Ocity'])
        if Excel_Dictionary1['Ostate'] != None:    
            driver.find_element_by_name('M0490').send_keys(Excel_Dictionary1['Ostate'])
        if Excel_Dictionary1['OZip / Postal Code'] != None:    
            driver.find_element_by_id('SOfficePostalCodeCopy').send_keys(Excel_Dictionary1['OZip / Postal Code'])



time.sleep(2)    
if Excel_Dictionary1['Mailing Preference'] != None:
    if Excel_Dictionary1['Mailing Preference'] == 'Office':
        driver.find_element_by_id('office').click()
    elif Excel_Dictionary1['Mailing Preference'] == 'Residence':
        driver.find_element_by_id('home').click()   

if Excel_Dictionary1['Residential Address same as NRIC '] != None:
    if Excel_Dictionary1['Residential Address same as NRIC '] == 'Yes':
        driver.find_element_by_id('residentialAddressSameAsNRICYes').click()    
    elif Excel_Dictionary1['Residential Address same as NRIC '] == 'No':     
        driver.find_element_by_id('residentialAddressSameAsNRICNo').click()

if Excel_Dictionary1['Residential Type'] != None:    
    driver.find_element_by_name('M0035').send_keys(Excel_Dictionary1['Residential Type'])
    if Excel_Dictionary1['Residential Type'] == 'Others':
        if Excel_Dictionary1['Other Residential Type'] != None:
            driver.find_element_by_id('residentailTypeOthers').send_keys(Excel_Dictionary1['Other Residential Type'])

if Excel_Dictionary1['Residential Status'] != None:    
    driver.find_element_by_name('M0032').send_keys(Excel_Dictionary1['Residential Status'])
    if Excel_Dictionary1['Residential Status'] == 'Others':
        if Excel_Dictionary1['Other Residential Status'] != None:
            driver.find_element_by_id('otherResidentailStatus').send_keys(Excel_Dictionary1['Other Residential Status'])
    if Excel_Dictionary1['Residential Status'] == 'Rented':
        if Excel_Dictionary1['Rental Amount in S$ Per Month'] != None:
            driver.find_element_by_name('M0034').send_keys(Excel_Dictionary1['Rental Amount in S$ Per Month'])

if Excel_Dictionary1['Length of Residency_Years'] != None:    
    driver.find_element_by_name('M0038').send_keys(Excel_Dictionary1['Length of Residency_Years'])
if Excel_Dictionary1['Length of Residency_Months'] != None:    
    driver.find_element_by_name('M0037').send_keys(Excel_Dictionary1['Length of Residency_Months'])


if Excel_Dictionary1['Country of Tax Residency 1'] != None:    
    driver.find_element_by_id('taxResidencyCountry').send_keys(Excel_Dictionary1['Country of Tax Residency 1'])
    if Excel_Dictionary1['Tax No 1'] != None:
        driver.find_element_by_id('USTaxNo').send_keys(Excel_Dictionary1['Tax No 1'])
if Excel_Dictionary1['Country of Tax Residency 2'] != None:
    driver.find_element_by_xpath('//*[@id="addothertaxID"]/span').click()
    driver.find_element_by_id('taxResidencyCountry1').send_keys(Excel_Dictionary1['Country of Tax Residency 2'])
    if Excel_Dictionary1['Tax No 2'] != None:
        driver.find_element_by_id('USTaxNo1').send_keys(Excel_Dictionary1['Tax No 2'])
if Excel_Dictionary1['Country of Tax Residency 3'] != None:
    driver.find_element_by_xpath('//*[@id="addothertaxID"]/span').click()
    driver.find_element_by_id('taxResidencyCountry2').send_keys(Excel_Dictionary1['Country of Tax Residency 3'])        
    if Excel_Dictionary1['Tax No 3'] != None:
        driver.find_element_by_id('USTaxNo2').send_keys(Excel_Dictionary1['Tax No 3'])    



if (Excel_Dictionary1['Employment Status'] != None) and (Excel_Dictionary1['Employment Status'] != 'Homemaker' or Excel_Dictionary1['Employment Status'] != 'Retired' or Excel_Dictionary1['Employment Status'] != 'Student' or Excel_Dictionary1['Employment Status'] != 'Unemployed'):    
    if Excel_Dictionary1['Employment Status'] != None:    
        driver.find_element_by_name('M0095').send_keys(Excel_Dictionary1['Employment Status'])
    if Excel_Dictionary1['Name of Company'] != None:
        driver.find_element_by_name('M0096').send_keys(Excel_Dictionary1['Name of Company'])
    if Excel_Dictionary1['Job Title'] != None:
        driver.find_element_by_name('M0098').send_keys(Excel_Dictionary1['Job Title'])
    if Excel_Dictionary1['Level'] != None:
        driver.find_element_by_name('M0099').send_keys(Excel_Dictionary1['Level'])
    if (Excel_Dictionary1['Employment Status'] == 'Self Employed') or (Excel_Dictionary1['Level'] == 'Owner,Director or C-Level'):
        if Excel_Dictionary1['Percentage shares in Company %'] != None:
            driver.find_element_by_name('M0097').send_keys(Excel_Dictionary1['Percentage shares in Company %'])
    if Excel_Dictionary1['Occupation Type'] != None:
        driver.find_element_by_name('M0100').send_keys(Excel_Dictionary1['Occupation Type'])
        if Excel_Dictionary1['Occupation Type'] == 'Others':
            if Excel_Dictionary1['Other Occupation Type'] != None:
                driver.find_element_by_id('Self_Other_Occupation_Current_Company').send_keys(Excel_Dictionary1['Other Occupation Type'])
    if Excel_Dictionary1['Industry Type'] != None:
        driver.find_element_by_name('M0102').send_keys(Excel_Dictionary1['Industry Type'])
        if Excel_Dictionary1['Industry Type'] == 'Others':
            if Excel_Dictionary1['Other Industry Type'] != None:
                driver.find_element_by_id('Self_Other_Industry_Current_Company').send_keys(Excel_Dictionary1['Other Industry Type'])
    if Excel_Dictionary1['Length of Employment_Years'] != None:
        driver.find_element_by_name('M0104').send_keys(Excel_Dictionary1['Length of Employment_Years'])
    if Excel_Dictionary1['Length of Employment_Months'] != None:
        driver.find_element_by_name('M0105').send_keys(Excel_Dictionary1['Length of Employment_Months'])


if Excel_Dictionary1['Name of Previous Company'] != None:
    driver.find_element_by_name('M0106').send_keys(Excel_Dictionary1['Name of Previous Company'])
if Excel_Dictionary1['Prev_Job_Title'] != None:
    driver.find_element_by_name('M0107').send_keys(Excel_Dictionary1['Prev_Job_Title'])
if Excel_Dictionary1['Prev_Occupation'] != None:
    driver.find_element_by_name('M0108').send_keys(Excel_Dictionary1['Prev_Occupation'])
if Excel_Dictionary1['Prev_Industry Type'] != None:
    driver.find_element_by_name('M0110').send_keys(Excel_Dictionary1['Prev_Industry Type'])
if Excel_Dictionary1['Prev_Length of Employment_Years'] != None:
    driver.find_element_by_name('M0112').send_keys(Excel_Dictionary1['Prev_Length of Employment_Years'])
if Excel_Dictionary1['Prev_Length of Employment_Months'] != None:
    driver.find_element_by_name('M0113').send_keys(Excel_Dictionary1['Prev_Length of Employment_Months'])


if (Excel_Dictionary1['Key Relationship'] != None) and (Excel_Dictionary1['Key Relationship'] == 'Yes'):
    driver.find_element_by_id('selfKeyRelationshipYes').click()
    if Excel_Dictionary1['Who is the person for the Position?'] != None:
        if Excel_Dictionary1['Who is the person for the Position?'] == 'Myself':
            driver.find_element_by_id('selfPersonPositionYes').click()
        elif Excel_Dictionary1['Who is the person for the Position?'] == 'My Family Member':
            driver.find_element_by_id('selfPersonPositionNo').click()  
            if Excel_Dictionary1['Relationship to Applicant'] != None:
                driver.find_element_by_name('M0549').send_keys(Excel_Dictionary1['Relationship to Applicant'])
    if Excel_Dictionary1['What is the status of the position?'] != None:
        if Excel_Dictionary1['What is the status of the position?'] == 'Currently Hold':
            driver.find_element_by_id('selfRelationshipCurrentlyHold').click()
            if Excel_Dictionary1['From Year'] != None:
                driver.find_element_by_name('M0421').send_keys(Excel_Dictionary1['From Year'])
        elif Excel_Dictionary1['What is the status of the position?'] == 'Have Held':
            driver.find_element_by_id('selfRelationshipHaveHeld').click()
            if Excel_Dictionary1['From Year'] != None:
                driver.find_element_by_name('M0421').send_keys(Excel_Dictionary1['From Year'])
            if Excel_Dictionary1['To'] != None:
                driver.find_element_by_name('M0422').send_keys(Excel_Dictionary1['To'])      
        elif Excel_Dictionary1['What is the status of the position?'] == 'Actively Seeking':
            driver.find_element_by_id('selfRelationshipActivelySeeking').click()  
        elif Excel_Dictionary1['What is the status of the position?'] == 'Being considered':
            driver.find_element_by_id('selfRelationshipBeingConsidered').click() 
    if Excel_Dictionary1['Name'] != None:
        driver.find_element_by_name('M0419').send_keys(Excel_Dictionary1['Name'])      
    if Excel_Dictionary1['Position'] != None:
        driver.find_element_by_name('M0420').send_keys(Excel_Dictionary1['Position'])                  
    if Excel_Dictionary1['Country'] != None:
        driver.find_element_by_name('M0548').send_keys(Excel_Dictionary1['Country'])                                 
else:
    driver.find_element_by_id('selfKeyRelationshipNo').click()    



if Excel_Dictionary1['Monthly Fixed Income'] != None:
    driver.find_element_by_id('Self_Basic_Employment_Income').send_keys(Excel_Dictionary1['Monthly Fixed Income'])
if Excel_Dictionary1['Total(NOA)'] != None:
    driver.find_element_by_id('Self_Total_Annual_Income').send_keys(Excel_Dictionary1['Total(NOA)'])
if Excel_Dictionary1['Rental Income'] != None:
    driver.find_element_by_id('Self_Rental_Income').send_keys(Excel_Dictionary1['Rental Income'])
if Excel_Dictionary1['Enter other income item name'] != None:
    driver.find_element_by_xpath('//*[@id="OwnerBorrowerFieldsToggle"]/div[13]/div/div/div[1]/div/button/span').click()
    driver.find_element_by_id('otherIncomeItem').click() 
    driver.find_element_by_id('otherIncomeItem').send_keys(Excel_Dictionary1['Enter other income item name'])
    driver.find_element_by_id('addotherincomeID').click()
    if Excel_Dictionary1['Other Income Amount'] != None:
        driver.find_element_by_xpath('//*[@id="self_Other_Income4"]').send_keys(Excel_Dictionary1['Other Income Amount'])

if (Excel_Dictionary1['Inheritance / Gift'] != None) and (Excel_Dictionary1['Inheritance / Gift'] == 'Yes'):
    driver.find_element_by_id('M0202_0').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Directorship_or_Dividends'] != None) and (Excel_Dictionary1['Directorship_or_Dividends'] == 'Yes'):
    driver.find_element_by_id('M0202_1').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Employment Income'] != None) and (Excel_Dictionary1['Employment Income'] == 'Yes'):
    driver.find_element_by_id('M0202_2').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Own Business'] != None) and (Excel_Dictionary1['Own Business'] == 'Yes'):
    driver.find_element_by_id('M0202_3').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Rental income'] != None) and (Excel_Dictionary1['Rental income'] == 'Yes'):
    driver.find_element_by_id('M0202_4').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Investments'] != None) and (Excel_Dictionary1['Investments'] == 'Yes'):
    driver.find_element_by_id('M0202_5').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Savings'] != None) and (Excel_Dictionary1['Savings'] == 'Yes'):
    driver.find_element_by_id('M0202_6').send_keys(Keys.SPACE)
if (Excel_Dictionary1['Sale of Property'] != None) and (Excel_Dictionary1['Sale of Property'] == 'Yes'):
    driver.find_element_by_id('M0202_7').send_keys(Keys.SPACE)






if Excel_Dictionary1['Do you own other properties in Singapore/Overseas'] != None:
    if Excel_Dictionary1['Do you own other properties in Singapore/Overseas'] == 'Yes':
        driver.find_element_by_id('selfOtherPropertyYes').click()
        if Excel_Dictionary1['Other Properties Property Name'] != None:
            driver.find_element_by_name('M0517').send_keys(Excel_Dictionary1['Other Properties Property Name'])
        if (Excel_Dictionary1['Other Properties Country'] != None) and (Excel_Dictionary1['Other Properties Country'] == 'Singapore'):
            driver.find_element_by_name('M0518').send_keys(Excel_Dictionary1['Other Properties Country'])
        if Excel_Dictionary1['Other Properties Zip / Postal Code'] != None:
            driver.find_element_by_name('M0519').send_keys(Excel_Dictionary1['Other Properties Zip / Postal Code'])    
        if Excel_Dictionary1['Other Properties Property Address'] != None:
            driver.find_element_by_name('M0268').send_keys(Excel_Dictionary1['Other Properties Property Address'])    
        if Excel_Dictionary1['Other Properties Unit No'] != None:
            driver.find_element_by_name('M0520').send_keys(Excel_Dictionary1['Other Properties Unit No'])    
    elif Excel_Dictionary1['Do you own other properties in Singapore/Overseas'] == 'No':
        driver.find_element_by_id('selfOtherPropertyNo').click()


# if Excel_Dictionary1['Own Business'] != None:
#     driver.find_element_by_id(Excel_Dictionary1['Own Business']).send_keys(Keys.SPACE)
# if Excel_Dictionary1['Employment Income'] != None:
#     driver.find_element_by_id(Excel_Dictionary1['Employment Income']).send_keys(Keys.SPACE)


############################################  JOINT APPLICANT NO 2  ###################################





driver.find_element_by_xpath("(.//*[normalize-space(text()) and normalize-space(.)='Remove'])[3]/following::img[1]").click()
time.sleep(5)
driver.find_element_by_id('jointOwnerOnlyPlusBorrowerId').click()
if Excel_Dictionary2['Salutation'] != None:
    driver.find_element_by_xpath('//*[@id="Jsalutation"]').send_keys(Excel_Dictionary2['Salutation'])
if Excel_Dictionary2['Full Name as in NRIC/PASSPORT'] != None:
    driver.find_element_by_xpath('//*[@id="JfullName"]').clear()
    driver.find_element_by_xpath('//*[@id="JfullName"]').send_keys(Excel_Dictionary2['Full Name as in NRIC/PASSPORT'])
if Excel_Dictionary2['Date of Birth'] != None:
    driver.find_element_by_name('M0059').clear()
    driver.find_element_by_name('M0059').send_keys(Excel_Dictionary2['Date of Birth'])
if Excel_Dictionary2['Country Of Birth'] != None:
    driver.find_element_by_name('M0414').send_keys(Excel_Dictionary2['Country Of Birth'])
if Excel_Dictionary2['Race'] != None:
    driver.find_element_by_name('M0057').send_keys(Excel_Dictionary2['Race'])
    if Excel_Dictionary2['Race'] == 'Others':
        if Excel_Dictionary1['Other Race'] != None:
            driver.find_element_by_id('JotherRace').send_keys(Excel_Dictionary2['Other Race'])
if Excel_Dictionary2['Marital Status'] != None:
    driver.find_element_by_name('M0055').send_keys(Excel_Dictionary2['Marital Status'])
    if Excel_Dictionary2['Marital Status'] == 'Others':
        if Excel_Dictionary2['Other Marital Status'] != None:
            driver.find_element_by_id('JOtherMaritalStatus').send_keys(Excel_Dictionary2['Other Marital Status'])


if Excel_Dictionary2['Nationality'] != None:
    driver.find_element_by_name('M0065').send_keys(Excel_Dictionary2['Nationality'])
if Excel_Dictionary2['Additional Nationality 1'] != None:
    driver.find_element_by_xpath('//*[@id="JaddothernationalityID"]/span').click()
    driver.find_element_by_id('JadditionalNationality').send_keys(Excel_Dictionary2['Additional Nationality 1'])
if Excel_Dictionary2['Additional Nationality 2'] != None:
    driver.find_element_by_xpath('//*[@id="JaddothernationalityID"]/span').click()
    driver.find_element_by_id('JadditionalNationality1').send_keys(Excel_Dictionary2['Additional Nationality 2'])
if Excel_Dictionary2['Singapore PR'] != None:    
    driver.find_element_by_xpath('//*[@id="option3"]').click()
if Excel_Dictionary2['Passport No.'] != None:
    driver.find_element_by_name('M0060').send_keys(Excel_Dictionary2['Passport No.'])
if Excel_Dictionary2['Issue Date'] != None:
    driver.find_element_by_name('M0061').send_keys(Excel_Dictionary2['Issue Date'])
if Excel_Dictionary2['Expiry Date'] != None:
    driver.find_element_by_name('M0062').send_keys(Excel_Dictionary2['Expiry Date'])
if Excel_Dictionary2['Country Of Issue'] != None:
    driver.find_element_by_name('M0063').send_keys(Excel_Dictionary2['Country Of Issue'])
if Excel_Dictionary2['Previous Passport No.'] != None:
    driver.find_element_by_name('M0064').send_keys(Excel_Dictionary2['Previous Passport No.'])
if Excel_Dictionary2['Nric No.'] != None:
    driver.find_element_by_name('M0060').send_keys(Excel_Dictionary2['Nric No.'])
if Excel_Dictionary2['Mother Maiden Name'] != None:
    driver.find_element_by_name('M0068').send_keys(Excel_Dictionary2['Mother Maiden Name'])
if Excel_Dictionary2['Highest Education'] != None:
    driver.find_element_by_name('M0066').send_keys(Excel_Dictionary2['Highest Education'])
if Excel_Dictionary2['No. Of Dependents'] != None:
    driver.find_element_by_name('M0069').send_keys(Excel_Dictionary2['No. Of Dependents'])
if Excel_Dictionary2['Age Of Dependent'] != None:
    driver.find_element_by_id('JageOfDependent1').send_keys(Excel_Dictionary2['Age Of Dependent'])
if Excel_Dictionary2['Email Address'] != None:
    driver.find_element_by_name('M0081').send_keys(Excel_Dictionary2['Email Address'])
if Excel_Dictionary2['MCountry Code'] != None:
    Select(driver.find_element_by_id('JhandPhoneCountryCode')).select_by_visible_text(Excel_Dictionary2['MCountry Code'])
if Excel_Dictionary2['Mobile Number'] != None:
    driver.find_element_by_name('M0080').send_keys(Excel_Dictionary2['Mobile Number'])
if Excel_Dictionary2['HCountry Code'] != None:
    Select(driver.find_element_by_id('JhomePhoneCountryCode')).select_by_visible_text(Excel_Dictionary2['HCountry Code'])
if Excel_Dictionary2['Home Phone'] != None:
    driver.find_element_by_name('M0078').send_keys(Excel_Dictionary2['Home Phone'])
if Excel_Dictionary2['OCountry Code'] != None:
    Select(driver.find_element_by_id('JofficePhoneCountryCode')).select_by_visible_text(Excel_Dictionary2['OCountry Code'])
if Excel_Dictionary2['Office Phone'] != None:
    driver.find_element_by_name('M0079').send_keys(Excel_Dictionary2['Office Phone'])



if Excel_Dictionary2['Country Of Residence'] != None:    
    driver.find_element_by_name('M0071').send_keys(Excel_Dictionary2['Country Of Residence'])
    if Excel_Dictionary2['Country Of Residence'] == 'Singapore':
        if Excel_Dictionary2['Zip / Postal Code'] != None:
            driver.find_element_by_id('jointHomePostalCode').send_keys(Excel_Dictionary2['Zip / Postal Code'])
        if Excel_Dictionary2['Residential Address'] != None:
            driver.find_element_by_id('jointHomeAddress').send_keys(Excel_Dictionary2['Residential Address'])
        if Excel_Dictionary2['Unit No'] != None:
            driver.find_element_by_id('jointResidentialUnitNo').send_keys(Excel_Dictionary2['Unit No'])  
    else:
        if Excel_Dictionary2['Residential Address Line 1'] != None:    
            driver.find_element_by_id('jointHomeAddressLine1Copy').send_keys(Excel_Dictionary2['Residential Address Line 1'])
        if Excel_Dictionary2['Residential Address Line 2'] != None:
            driver.find_element_by_id('jointHomeAddressLine2').send_keys(Excel_Dictionary2['Residential Address Line 2'])
        if Excel_Dictionary2['City'] != None:    
            driver.find_element_by_name('M0497').send_keys(Excel_Dictionary2['City'])
        if Excel_Dictionary2['State'] != None:    
            driver.find_element_by_name('M0498').send_keys(Excel_Dictionary2['State'])
        if Excel_Dictionary2['Zip / Postal Code'] != None:    
            driver.find_element_by_id('jointHomePostalCodeCopy').send_keys(Excel_Dictionary2['Zip / Postal Code'])

if Excel_Dictionary2['Country Of Office'] != None:
    driver.find_element_by_name('M0455').send_keys(Excel_Dictionary2['Country Of Office'])
    if Excel_Dictionary2['Country Of Office'] == 'Singapore':
        if Excel_Dictionary2['OZip / Postal Code'] != None:    
            driver.find_element_by_id('jointOfficePostalCode').send_keys(Excel_Dictionary2['OZip / Postal Code'])
        if Excel_Dictionary2['Office Address'] != None:
            driver.find_element_by_id('jointOfficeAddress').send_keys(Excel_Dictionary2['Office Address'])    
        if Excel_Dictionary2['OUnit No'] != None:
            driver.find_element_by_id('jointOfficeUnitNo').send_keys(Excel_Dictionary2['OUnit No'])    
    else:
        if Excel_Dictionary2['Office Address Line 1'] != None:    
            driver.find_element_by_id('jointOfficeAddressCopy').send_keys(Excel_Dictionary2['Office Address Line 1'])
        if Excel_Dictionary2['Office Address Line 2'] != None:    
            driver.find_element_by_id('jointOfficeAddressLine2').send_keys(Excel_Dictionary2['Office Address Line 2'])
        if Excel_Dictionary2['Ocity'] != None:    
            driver.find_element_by_name('M0501').send_keys(Excel_Dictionary2['Ocity'])
        if Excel_Dictionary2['Ostate'] != None:    
            driver.find_element_by_name('M0502').send_keys(Excel_Dictionary2['Ostate'])
        if Excel_Dictionary2['OZip / Postal Code'] != None:    
            driver.find_element_by_id('jointOfficePostalCodeCopy').send_keys(Excel_Dictionary2['OZip / Postal Code'])


time.sleep(3)
if Excel_Dictionary2['Mailing Preference'] != None:
    if Excel_Dictionary2['Mailing Preference'] == 'Office':
        driver.find_element_by_id('Joffice').click()
    elif Excel_Dictionary2['Mailing Preference'] == 'Residence':
        driver.find_element_by_id('Jhome').click()    


if Excel_Dictionary2['Residential Address same as NRIC '] != None:
    if Excel_Dictionary2['Residential Address same as NRIC '] == 'Yes':
        driver.find_element_by_id('JresidentialAddressSameAsNRICYes').click()
    elif Excel_Dictionary2['Residential Address same as NRIC '] == 'No':    
        driver.find_element_by_id('JresidentialAddressSameAsNRICNo').click()

if Excel_Dictionary2['Residential Type'] != None:
    driver.find_element_by_name('M0085').send_keys(Excel_Dictionary2['Residential Type'])
    if Excel_Dictionary2['Residential Type'] == 'Others':
        if Excel_Dictionary2['Other Residential Type'] != None:
            driver.find_element_by_id('JresidentailTypeOthers').send_keys(Excel_Dictionary2['Other Residential Type'])

if Excel_Dictionary2['Residential Status'] != None:    
    driver.find_element_by_name('M0082').send_keys(Excel_Dictionary2['Residential Status'])
    if Excel_Dictionary2['Residential Status'] == 'Others':
        if Excel_Dictionary2['Other Residential Status'] != None:
            driver.find_element_by_id('JotherResidentailStatus').send_keys(Excel_Dictionary2['Other Residential Status'])
    if Excel_Dictionary2['Residential Status'] == 'Rented':
        if Excel_Dictionary2['Rental Amount in S$ Per Month'] != None:
            driver.find_element_by_name('M0084').send_keys(Excel_Dictionary2['Rental Amount in S$ Per Month'])



# driver.find_element_by_name('M0084').send_keys(Excel_Dictionary2['Rental Amount in S$ Per Month'])
if Excel_Dictionary2['Length of Residency_Years'] != None:
    driver.find_element_by_name('M0088').send_keys(Excel_Dictionary2['Length of Residency_Years'])
if Excel_Dictionary2['Length of Residency_Months'] != None:
    driver.find_element_by_name('M0087').send_keys(Excel_Dictionary2['Length of Residency_Months'])

if Excel_Dictionary2['Country of Tax Residency 1'] != None:    
    driver.find_element_by_id('JtaxResidencyCountry').send_keys(Excel_Dictionary2['Country of Tax Residency 1'])
    if Excel_Dictionary2['Tax No 1'] != None:
        driver.find_element_by_id('JUSTaxNo').send_keys(Excel_Dictionary2['Tax No 1'])
if Excel_Dictionary2['Country of Tax Residency 2'] != None:
    driver.find_element_by_xpath('//*[@id="JaddothertaxID"]/span').click()
    driver.find_element_by_id('JtaxResidencyCountry1').send_keys(Excel_Dictionary2['Country of Tax Residency 2'])
    if Excel_Dictionary2['Tax No 2'] != None:
        driver.find_element_by_id('JUSTaxNo1').send_keys(Excel_Dictionary2['Tax No 2'])
if Excel_Dictionary2['Country of Tax Residency 3'] != None:
    driver.find_element_by_xpath('//*[@id="JaddothertaxID"]/span').click()
    driver.find_element_by_id('JtaxResidencyCountry2').send_keys(Excel_Dictionary2['Country of Tax Residency 3'])        
    if Excel_Dictionary2['Tax No 3'] != None:
        driver.find_element_by_id('JUSTaxNo2').send_keys(Excel_Dictionary2['Tax No 3'])    


if (Excel_Dictionary2['Employment Status'] != None) and (Excel_Dictionary2['Employment Status'] != 'Homemaker' or Excel_Dictionary2['Employment Status'] != 'Retired' or Excel_Dictionary2['Employment Status'] != 'Student' or Excel_Dictionary2['Employment Status'] != 'Unemployed'):
    if Excel_Dictionary2['Employment Status'] != None:
        driver.find_element_by_name('M0158').send_keys(Excel_Dictionary2['Employment Status'])
    if Excel_Dictionary2['Name of Company'] != None:
        driver.find_element_by_name('M0127').send_keys(Excel_Dictionary2['Name of Company'])
    if Excel_Dictionary2['Job Title'] != None:
        driver.find_element_by_name('M0129').send_keys(Excel_Dictionary2['Job Title'])
    if Excel_Dictionary2['Level'] != None:
        driver.find_element_by_name('M0130').send_keys(Excel_Dictionary2['Level'])
    if (Excel_Dictionary2['Employment Status'] == 'Self Employed') or (Excel_Dictionary2['Level'] == 'Owner,Director or C-Level'):
        if Excel_Dictionary2['Percentage shares in Company %'] != None:
            driver.find_element_by_name('M0128').send_keys(Excel_Dictionary2['Percentage shares in Company %'])    
    if Excel_Dictionary2['Occupation Type'] != None:
        driver.find_element_by_name('M0131').send_keys(Excel_Dictionary2['Occupation Type'])
        if Excel_Dictionary2['Occupation Type'] == 'Others':
            if Excel_Dictionary2['Other Occupation Type'] != None:
                driver.find_element_by_id('Joint_Other_Occupation_Current_Company').send_keys(Excel_Dictionary2['Other Occupation Type'])
    if Excel_Dictionary2['Industry Type'] != None:
        driver.find_element_by_name('M0133').send_keys(Excel_Dictionary2['Industry Type'])
        if Excel_Dictionary2['Industry Type'] == 'Others':
            if Excel_Dictionary2['Other Industry Type'] != None:
                driver.find_element_by_id('Joint_Other_Industry_Current_Company').send_keys(Excel_Dictionary2['Other Industry Type'])
    if Excel_Dictionary2['Length of Employment_Years'] != None:
        driver.find_element_by_name('M0135').send_keys(Excel_Dictionary2['Length of Employment_Years'])
    if Excel_Dictionary2['Length of Employment_Months'] != None:
        driver.find_element_by_name('M0136').send_keys(Excel_Dictionary2['Length of Employment_Months'])

    
if Excel_Dictionary2['Name of Previous Company'] != None:
    driver.find_element_by_name('M0137').send_keys(Excel_Dictionary2['Name of Previous Company'])
if Excel_Dictionary2['Prev_Job_Title'] != None:
    driver.find_element_by_name('M0138').send_keys(Excel_Dictionary2['Prev_Job_Title'])
if Excel_Dictionary2['Prev_Occupation'] != None:
    driver.find_element_by_name('M0139').send_keys(Excel_Dictionary2['Prev_Occupation'])
if Excel_Dictionary2['Prev_Industry Type'] != None:
    driver.find_element_by_name('M0141').send_keys(Excel_Dictionary2['Prev_Industry Type'])
if Excel_Dictionary2['Prev_Length of Employment_Years'] != None:
    driver.find_element_by_name('M0143').send_keys(Excel_Dictionary2['Prev_Length of Employment_Years'])
if Excel_Dictionary2['Prev_Length of Employment_Months'] != None:
    driver.find_element_by_name('M0144').send_keys(Excel_Dictionary2['Prev_Length of Employment_Months'])

if (Excel_Dictionary2['Key Relationship'] != None) and (Excel_Dictionary2['Key Relationship'] == 'Yes'):
    driver.find_element_by_id('jointKeyRelationshipYes').click()
    if Excel_Dictionary2['Who is the person for the Position?'] != None:
        if Excel_Dictionary2['Who is the person for the Position?'] == 'Myself':
            driver.find_element_by_id('jointPersonPositionYes').click()
        elif Excel_Dictionary2['Who is the person for the Position?'] == 'My Family Member':
            driver.find_element_by_id('jointPersonPositionNo').click()  
            if Excel_Dictionary2['Relationship to Applicant'] != None:
                driver.find_element_by_name('M0574').send_keys(Excel_Dictionary2['Relationship to Applicant'])
    if Excel_Dictionary2['What is the status of the position?'] != None:
        if Excel_Dictionary2['What is the status of the position?'] == 'Currently Hold':
            driver.find_element_by_id('jointRelationshipCurrentlyHold').click()
            if Excel_Dictionary2['From Year'] != None:
                driver.find_element_by_name('M0426').send_keys(Excel_Dictionary2['From Year'])
        elif Excel_Dictionary2['What is the status of the position?'] == 'Have Held':
            driver.find_element_by_id('jointRelationshipHaveHeld').click()
            if Excel_Dictionary2['From Year'] != None:
                driver.find_element_by_name('M0426').send_keys(Excel_Dictionary2['From Year'])
            if Excel_Dictionary2['To'] != None:
                driver.find_element_by_name('M0427').send_keys(Excel_Dictionary2['To'])      
        elif Excel_Dictionary2['What is the status of the position?'] == 'Actively Seeking':
            driver.find_element_by_id('jointRelationshipActivelySeeking').click()  
        elif Excel_Dictionary2['What is the status of the position?'] == 'Being considered':
            driver.find_element_by_id('jointRelationshipBeingConsidered').click() 
    if Excel_Dictionary2['Name'] != None:
        driver.find_element_by_name('M0424').send_keys(Excel_Dictionary2['Name'])      
    if Excel_Dictionary2['Position'] != None:
        driver.find_element_by_name('M0425').send_keys(Excel_Dictionary2['Position'])                  
    if Excel_Dictionary2['Country'] != None:
        driver.find_element_by_name('M0573').send_keys(Excel_Dictionary2['Country'])                                 
else:
    driver.find_element_by_id('jointKeyRelationshipNo').click()    


if Excel_Dictionary2['Monthly Fixed Income'] != None:
    driver.find_element_by_id('Joint_Basic_Employment_Income').send_keys(Excel_Dictionary2['Monthly Fixed Income'])
if Excel_Dictionary2['Total(NOA)'] != None:
    driver.find_element_by_id('Joint_Total_Annual_Income').send_keys(Excel_Dictionary2['Total(NOA)'])
if Excel_Dictionary2['Rental Income'] != None:
    driver.find_element_by_id('Joint_Rental_Income').send_keys(Excel_Dictionary2['Rental Income'])
if Excel_Dictionary2['Enter other income item name'] != None:
    driver.find_element_by_xpath('//*[@id="jointOwnerBorrowerFieldsToggle"]/div[13]/div[1]/div/div/div/button/span').click()
    driver.find_element_by_id('jOtherIncomeItem').click() 
    driver.find_element_by_id('jOtherIncomeItem').send_keys(Excel_Dictionary2['Enter other income item name'])
    driver.find_element_by_id('jAddotherincomeID').click()
    if Excel_Dictionary2['Other Income Amount'] != None:
        driver.find_element_by_xpath('//*[@id="joint_Other_Income2"]').send_keys(Excel_Dictionary2['Other Income Amount'])
    

if (Excel_Dictionary2['Inheritance / Gift'] != None) and (Excel_Dictionary2['Inheritance / Gift'] == 'Yes'):
    driver.find_element_by_id('M0412_0').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Directorship_or_Dividends'] != None) and (Excel_Dictionary2['Directorship_or_Dividends'] == 'Yes'):
    driver.find_element_by_id('M0412_1').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Employment Income'] != None) and (Excel_Dictionary2['Employment Income'] == 'Yes'):
    driver.find_element_by_id('M0412_2').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Own Business'] != None) and (Excel_Dictionary2['Own Business'] == 'Yes'):
    driver.find_element_by_id('M0412_3').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Rental income'] != None) and (Excel_Dictionary2['Rental income'] == 'Yes'):
    driver.find_element_by_id('M0412_4').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Investments'] != None) and (Excel_Dictionary2['Investments'] == 'Yes'):
    driver.find_element_by_id('M0412_5').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Savings'] != None) and (Excel_Dictionary2['Savings'] == 'Yes'):
    driver.find_element_by_id('M0412_6').send_keys(Keys.SPACE)
if (Excel_Dictionary2['Sale of Property'] != None) and (Excel_Dictionary2['Sale of Property'] == 'Yes'):
    driver.find_element_by_id('M0412_7').send_keys(Keys.SPACE)



if Excel_Dictionary2['Do you own other properties in Singapore/Overseas'] != None:
    if Excel_Dictionary2['Do you own other properties in Singapore/Overseas'] == 'Yes':
        driver.find_element_by_id('jointOwnOtherPropertiesInSingaporeYes').click()
        if Excel_Dictionary2['Other Properties Property Name'] != None:
            driver.find_element_by_name('M0521').send_keys(Excel_Dictionary2['Other Properties Property Name'])
        if (Excel_Dictionary2['Other Properties Country'] != None) and (Excel_Dictionary2['Other Properties Country'] == 'Singapore'):
            driver.find_element_by_name('M0522').send_keys(Excel_Dictionary2['Other Properties Country'])
        if Excel_Dictionary2['Other Properties Zip / Postal Code'] != None:
            driver.find_element_by_name('M0523').send_keys(Excel_Dictionary2['Other Properties Zip / Postal Code'])    
        if Excel_Dictionary2['Other Properties Property Address'] != None:
            driver.find_element_by_name('M0399').send_keys(Excel_Dictionary2['Other Properties Property Address'])    
        if Excel_Dictionary2['Other Properties Unit No'] != None:
            driver.find_element_by_name('M0524').send_keys(Excel_Dictionary2['Other Properties Unit No'])    
    elif Excel_Dictionary2['Do you own other properties in Singapore/Overseas'] == 'No':
        driver.find_element_by_id('jointOwnOtherPropertiesInSingaporeNo').click()


driver.find_element_by_xpath('//*[@id="loan"]').click()
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
time.sleep(4)
driver.find_element_by_class_name('loan-submit-button').click()
time.sleep(3)
driver.find_element_by_xpath('//*[@id="checkbox-select"]').click()
driver.find_element_by_xpath('//*[@id="maybank-deselected"]').click()
driver.find_element_by_xpath('//*[@id="downloadForms"]').click()
time.sleep(70)
driver.quit()
print("Form Filling Done")
