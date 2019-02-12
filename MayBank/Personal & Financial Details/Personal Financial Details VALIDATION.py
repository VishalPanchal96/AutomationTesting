from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time


path = "D:\\Automation Testing\\MayBank\\Personal & Financial Details\\Personal&Financial_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Self_Dictionary = {}
Joint_Dictionary = {}
for i in range(1,80):
    Self_Dictionary[sheet.cell(i,1).value] = sheet.cell(i,2).value
    Joint_Dictionary[sheet.cell(i,1).value] = sheet.cell(i,3).value

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)
driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_xpath('//*[@id="loansTable"]/tbody/tr[1]/td[8]/div/button').click()
time.sleep(3)
element = driver.find_element_by_xpath('//*[@id="applicationIdField"]').text
driver.quit()

Updated_PDF = "D:\\vaibhav\\Vaibhav Clone\\mortgage-webapp\\"+element+"_May Bank_Updated1.pdf"
try:
    pdf = PyPDF2.PdfFileReader(Updated_PDF)
except:
    error = "PDF file Not Found."
    print(error)

pdfData = pdf.getFields()
outputdict={}

# Creating a dictionary from pdf data    
for data in pdfData:
    info = pdfData[data]
    if "/V" in info:
        # if info["/V"] == "/" or info["/V"] == "" :
        #     pass
        # else :
        tag=info["/T"]
        value=info["/V"]
        # tag=tag.replace("_"," ")
        if "/" in value:
            value=value.replace("/", "")
        outputdict[tag]=value

def addComma(inputValue):
    if 'S$' in str(inputValue):
        inputValue = inputValue.replace('S$',"")
    if ',' in str(inputValue):
        return str(inputValue)
    if '.' in str(inputValue):
        parts = str(inputValue).split(".")
        intPart = parts[0]
        floatPart = parts[1]
        intPart = "{:,}".format(int(float(intPart)))
        if floatPart != '0' and floatPart != '00' :
            retstr= str(intPart) + "." + str(floatPart) 
        else :
            retstr= str(intPart)
        return retstr
    elif inputValue :
        try:
            return "{:,}".format(int(inputValue))
        except:
            parts = inputValue.split(".")
            intPart = parts[0]
            floatPart = parts[1]
            intPart = "{:,}".format(int(float(intPart)))
            return intPart + "." + floatPart
    else:
        return inputValue

def remove_dot_zero(input):
    if(float(input) % 1 == 0):
        return int(input)
    else:
        return input

results = []

########################################################## Validation For Single #############################################################

#Self Salutation Validation
if (Self_Dictionary['Salutation'] != None) and (Self_Dictionary['Salutation'] == 'Mr.' or Self_Dictionary['Salutation'] == 'Mdm.' or Self_Dictionary['Salutation'] == 'Mrs.' or Self_Dictionary['Salutation'] == 'Ms.' or Self_Dictionary['Salutation'] == 'Dr.'):
    if(Self_Dictionary['Salutation'] == 'Mr.') and (outputdict['Self_Salutation_Mr'] == 'Yes'):
        results.append("Self Salutation 'Mr.' Passed")
    elif(Self_Dictionary['Salutation'] == 'Mdm.') and (outputdict['Self_Salutation_Mdm'] == 'Yes'):
        results.append("Self Salutation 'Mdm.' Passed")
    elif(Self_Dictionary['Salutation'] == 'Mrs.') and (outputdict['Self_Salutation_Mrs'] == 'Yes'):
        results.append("Self Salutation 'Mrs.' Passed")
    elif(Self_Dictionary['Salutation'] == 'Ms.') and (outputdict['Self_Salutation_Ms'] == 'Yes'):
        results.append("Self Salutation 'Ms.' Passed")
    elif(Self_Dictionary['Salutation'] == 'Dr.') and (outputdict['Self_Salutation_Dr'] == 'Yes'):
        results.append("Self Salutation 'Dr.' Passed")
    else:
        results.append("Self Salutation Failed")
else:
    results.append("Self Salutation Data Not Available in UI or Input File or the Data is Incorrect")

#Self Passport No. Validation
if Self_Dictionary['Passport No.'] != None:
    if (Self_Dictionary['Passport No.'].upper()) == outputdict['Self_NRIC_Passport_No']:
        results.append("Self Passport Data Passed")
    else:
        results.append("Self Passport Data Failed")   
    if (outputdict['Self_Passport_Check'] == 'Yes'):
        results.append("Self Passport Tick Passed")
    else:
        results.append("Self Passport Tick Failed")        
else:
    results.append("Self Passport Data Not Available in UI or Input File")

#Self NRIC No. Validation
if 'Self_NRIC_Passport_No' in outputdict and 'Self_NRIC_Check' in outputdict:
    if Self_Dictionary['Nric No.'] != None: 
        if (Self_Dictionary['Nric No.'].upper()) == outputdict['Self_NRIC_Passport_No']:
            results.append("Self NRIC No. Data Passed")
        else:
            results.append("Self NRIC No. Data Failed") 
        if outputdict['Self_NRIC_Check'] == 'Yes':
            results.append("Self NRIC No. Tick Passed")
        else:
            results.append("Self NRIC No. Tick Failed")           
    else:
        results.append("Self NRIC No. Data Not Available in UI or Input File")

#Self Marital Status Validation
if (Self_Dictionary['Marital Status'] != None) and (Self_Dictionary['Marital Status'] == 'Single' or Self_Dictionary['Marital Status'] == 'Married' or Self_Dictionary['Marital Status'] == 'Divorced' or Self_Dictionary['Marital Status'] == 'Widowed' or Self_Dictionary['Marital Status'] == 'Others'):
    if (Self_Dictionary['Marital Status'] == 'Single') and (outputdict['Self_Marital_Status_Single'] == 'Yes'):
        results.append("Self Marital Status 'Single' Passed")
    elif (Self_Dictionary['Marital Status'] == 'Married') and (outputdict['Self_Marital_Status_Married'] == 'Yes'):
        results.append("Self Marital Status 'Married' Passed")
    elif (Self_Dictionary['Marital Status'] == 'Divorced') and (outputdict['Self_Marital_Status_Divorced'] == 'Yes'):
        results.append("Self Marital Status 'Divorced' Passed")
    elif (Self_Dictionary['Marital Status'] == 'Widowed') and (outputdict['Self_Marital_Status_OthersCheck'] == 'Yes') and (outputdict['Self_Marital_Status_OthersText'] == 'Widowed'):
        results.append("Self Marital Status 'Widowed' Passed")
    elif (Self_Dictionary['Marital Status'] == 'Others') and (outputdict['Self_Marital_Status_OthersCheck'] == 'Yes') and (Self_Dictionary['Other Marital Status'] != None) (Self_Dictionary['Other Marital Status'] == outputdict['Self_Marital_Status_OthersText']):
        results.append("Self Marital Status 'Others' with data and tick Passed")
    elif (Self_Dictionary['Marital Status'] == 'Others') and (outputdict['Self_Marital_Status_OthersCheck']) and (Self_Dictionary['Marital Status'] == None) and (outputdict['Self_Marital_Status_OthersText'] == ''):
        results.append("Self Marital Status 'Others' without data with tick Passed")    
    else:
        results.append("Self Marital Status Failed")
else:
    results.append("Self Marital Status Data Not Available in UI or Input File or the Data is Incorrect")        

#Self Number of Dependents Validation
if (str(Self_Dictionary['No. Of Dependents']) == outputdict['Self_Number_of_Dependents']):
    results.append("Self No. Of Dependents Passed")
else:
    results.append("Self No. Of Dependents Failed: Actual result- "+outputdict['Self_Number_of_Dependents']+" Expected Result- "+str(Self_Dictionary['No. Of Dependents']))    

#Self Highest Education Set 1 Validation
if (Self_Dictionary['Highest Education'] != None) and (Self_Dictionary['Highest Education'] == 'No Formal Education' or Self_Dictionary['Highest Education'] == 'Primary' or Self_Dictionary['Highest Education'] == 'Secondary' or Self_Dictionary['Highest Education'] == 'O Levels' or Self_Dictionary['Highest Education'] == 'N Levels' or Self_Dictionary['Highest Education'] == 'A Levels'):
    if (Self_Dictionary['Highest Education'] == "No Formal Education") and (outputdict['Self_Highest_Education_Primary'] == 'Yes'):
        results.append("Self Highest Education 'No Formal Education' Passed")
    elif (Self_Dictionary['Highest Education'] == "Primary") and (outputdict['Self_Highest_Education_Primary'] == 'Yes'):
        results.append("Self Highest Education 'Primary' Passed")
    elif (Self_Dictionary['Highest Education'] == "Secondary") and (outputdict['Self_Highest_Education_Secondary'] == 'Yes'):
        results.append("Self Highest Education 'Secondary' Passed")    
    elif (Self_Dictionary['Highest Education'] == "O Levels") and (outputdict['Self_Highest_Education_Secondary'] == 'Yes'):
        results.append("Self Highest Education 'O Levels' Passed")    
    elif (Self_Dictionary['Highest Education'] == "N Levels") and (outputdict['Self_Highest_Education_Secondary'] == 'Yes'):
        results.append("Self Highest Education 'N Levels' Passed")
    elif (Self_Dictionary['Highest Education'] == "A Levels") and (outputdict['Self_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Self Highest Education 'A Levels' Passed")          
    else:
        results.append("Self Highest Education Failed")
else:
    results.append("Self Highest Education Data Not Available in UI or Input File or the Data is Incorrect")

#Self Highest Education Set 2 Validation
if (Self_Dictionary['Highest Education'] != None) and (Self_Dictionary['Highest Education'] == 'NITEC/Higher NITEC' or Self_Dictionary['Highest Education'] == 'Pre-University' or Self_Dictionary['Highest Education'] == 'Diploma' or Self_Dictionary['Highest Education'] == 'Degree' or Self_Dictionary['Highest Education'] == 'Post-Graduate' or Self_Dictionary['Highest Education'] == 'Technical Certification'):
    if (Self_Dictionary['Highest Education'] == "NITEC/Higher NITEC") and (outputdict['Self_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Self Highest Education 'NITEC/Higher NITEC' Passed")
    elif (Self_Dictionary['Highest Education'] == "Pre-University") and (outputdict['Self_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Self Highest Education 'Pre-University' Passed")
    elif (Self_Dictionary['Highest Education'] == "Diploma") and (outputdict['Self_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Self Highest Education 'Diploma' Passed")    
    elif (Self_Dictionary['Highest Education'] == "Degree") and (outputdict['Self_Highest_Education_Degree'] == 'Yes'):
        results.append("Self Highest Education 'Degree' Passed")    
    elif (Self_Dictionary['Highest Education'] == "Post-Graduate") and (outputdict['Self_Highest_Education_Post_Graduate'] == 'Yes'):
        results.append("Self Highest Education 'Post-Graduate' Passed")
    elif (Self_Dictionary['Highest Education'] == "Technical Certification") and (outputdict['Self_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Self Highest Education 'Technical Certification' Passed")          
    else:
        results.append("Self Highest Education Failed")
else:
    results.append("Self Highest Education Data Not Available in UI or Input File or the Data is Incorrect")    

#Self Residential Address Same as NRIC Validation
if (Self_Dictionary['Residential Address same as NRIC '] != None) and (Self_Dictionary['Residential Address same as NRIC '] == 'No' or Self_Dictionary['Residential Address same as NRIC '] == 'Yes'):
    if (Self_Dictionary['Residential Address same as NRIC '] == 'No') and (outputdict['Self_Residential_Address_same_as_NRIC_No'] == 'Yes'):
        results.append("Self Residential Address same as NRIC 'No' Passed")
    elif (Self_Dictionary['Residential Address same as NRIC '] == 'Yes') and (outputdict['Self_Residential_Address_same_as_NRIC_Yes'] == 'Yes'):
        results.append("Self Residential Address same as NRIC 'Yes' Passed")
    else:
        results.append("Self Residential Address same as NRIC Failed")
else:
    results.append("Self Residential Address same as NRIC Data not Available in UI or Input file or Data is Incorrect")            

#Self Residential Address Validation
if Self_Dictionary['Country Of Residence'] != 'Singapore':
    Self_Residential_Address = (Self_Dictionary['Residential Address Line 1']+", "+Self_Dictionary['Residential Address Line 2']+", "+Self_Dictionary['City']+", "+Self_Dictionary['State']+", "+Self_Dictionary['Country Of Residence']).lower()
    actual_residential_output = outputdict['Self_Residential_Address'].replace("_","") + outputdict['Self_Residential_Address_Dummy']

    if Self_Residential_Address == actual_residential_output:
        results.append("Self Residential Address Passed")
    else:
        results.append("Self Residential Address Failed")   
elif Self_Dictionary['Country Of Residence'] == 'Singapore':
    Self_Residential_Address_S = (Self_Dictionary['Residential Address']+", "+Self_Dictionary['Unit No']+", "+Self_Dictionary['Country Of Residence']).lower()
    actual_residential_output_S = outputdict['Self_Residential_Address'].replace("_","") + outputdict['Self_Residential_Address_Dummy']            

    if Self_Residential_Address_S == actual_residential_output_S:
        results.append("Self Residential Address Passed")
    else:
        results.append("Self Residential Address Failed") 

#Self Residential Address Postal Code Validation
if str(Self_Dictionary['Zip / Postal Code']) == outputdict['Self_Home_Postal_Code']:
    results.append("Self Residential Postal Code Passed")
else:
    results.append("Self Residential Postal Code Failed")      

#Self Mailing Address Validation
if Self_Dictionary['Country Of Office'] != 'Singapore'
    Self_Mailing_Address = (Self_Dictionary['Office Address Line 1']+", "+Self_Dictionary['Office Address Line 2']+", "+Self_Dictionary['Ocity']+", "+Self_Dictionary['Ostate']+", "+Self_Dictionary['Country Of Office']).lower()
    actual_mailing_output = outputdict['Self_Mailing_Address'].replace("_","") + outputdict['Self_Mailing_Address_Dummy']

    if Self_Mailing_Address == actual_mailing_output:
        results.append("Self Mailing Address Passed")
    else:
        results.append("Self Mailing Address Failed")    
elif Self_Dictionary['Country Of Office'] == 'Singapore':
    Self_Mailing_Address_S = (Self_Dictionary['Office Address']+", "+Self_Dictionary['OUnit No']+", "+Self_Dictionary['Country Of Office']).lower()
    actual_mailing_output_S = outputdict['Self_Mailing_Address'].replace("_","") + outputdict['Self_Mailing_Address_Dummy']

    if Self_Mailing_Address_S == actual_mailing_output_S:
        results.append("Self Mailing Address Passed")
    else:
        results.append("Self Mailing Address Failed")    

#Self Mailing Address Postal Code Validation
if str(Self_Dictionary['OZip / Postal Code']) == outputdict['Self_Mailing_Postal_Code']:
    results.append("Self Mailing Postal Code Passed")
else:
    results.append("Self Mailing Postal Code Failed")

#Self Mobile,home,office Phone Number Validation 
self_mobile_number = "+"+Self_Dictionary['MCountry Code'].split("+")[1]+"-"+str(Self_Dictionary['Mobile Number'])
self_home_phone = "+"+Self_Dictionary['HCountry Code'].split("+")[1]+"-"+str(Self_Dictionary['Home Phone'])
self_office_phone = "+"+Self_Dictionary['OCountry Code'].split("+")[1]+"-"+str(Self_Dictionary['Office Phone'])

if self_mobile_number == outputdict['Self_Hand_phone']:
    results.append("Self Mobile No. Passed: Actual output- "+outputdict['Self_Hand_phone']+" Expected output- "+self_mobile_number)
else:
    results.append("Self Mobile No. Failed: Actual output- "+outputdict['Self_Hand_phone']+" Expected output- "+self_mobile_number)

if self_home_phone == outputdict['Self_Home_phone']:
    results.append("Self Home Phone Passed: Actual output- "+outputdict['Self_Home_phone']+" Expected output- "+self_home_phone)
else:
    results.append("Self Home Phone Failed: Actual output- "+outputdict['Self_Home_phone']+" Expected output- "+self_home_phone)

if self_office_phone == outputdict['Self_Office_phone']:
    results.append("Self Office Phone Passed: Actual output- "+outputdict['Self_Office_phone']+" Expected output- "+self_office_phone)
else:
    results.append("Self Office Phone Failed: Actual output- "+outputdict['Self_Office_phone']+" Expected output- "+self_office_phone)

#Self Email Address Validation
if Self_Dictionary['Email Address'] == outputdict['Self_Email_address']:
    results.append("Self Email Address Passed")
else:
    results.append("Self Email Address Failed")  

#Self Residential Status Set 1 Validation
if (Self_Dictionary['Residential Status'] != None) and (Self_Dictionary['Residential Status'] == 'Fully Paid' or Self_Dictionary['Residential Status'] == 'Oustanding Mortgage' or Self_Dictionary['Residential Status'] == 'Live with family' or Self_Dictionary['Residential Status'] == 'Parents'):
    if (Self_Dictionary['Residential Status'] == 'Fully Paid') and (outputdict['Self_Residential_Status_Fully_Owned'] == 'Yes'):
        results.append("Self Residential Status 'Fully Paid' Passed")
    elif (Self_Dictionary['Residential Status'] == 'Oustanding Mortgage') and (outputdict['Self_Residential_Status_Mortgaged'] == 'Yes'):
        results.append("Self Residential Status 'Oustanding Mortgage' Passed")
    elif (Self_Dictionary['Residential Status'] == 'Live with family') and (outputdict['Self_Residential_Status_Others'] == 'Yes') and (Self_Dictionary['Residential Status'] == outputdict['Self_Residential_Status_OthersText']):
        results.append("Self Residential Status 'Live with family'   Passed")
    elif (Self_Dictionary['Residential Status'] == 'Parents') and (outputdict['Self_Residential_Status_Parents'] == 'Yes'):
        results.append("Self Residential Status 'Parents' Passed")
    else:
        results.append("Self Residential Status Failed")     
else:
    results.append("Self Residential Status Data not Available in UI or Input file or Data is Incorrect")

#Self Residential Status Set 2 Validation
if (Self_Dictionary['Residential Status'] != None) and (Self_Dictionary['Residential Status'] == 'Relative' or Self_Dictionary['Residential Status'] == 'Rented' or Self_Dictionary['Residential Status'] == "Employer's" or Self_Dictionary['Residential Status'] == 'Others'):
    if (Self_Dictionary['Residential Status'] == 'Relative') and (outputdict['Self_Residential_Status_Others'] == 'Yes') and (Self_Dictionary['Residential Status'] == outputdict['Self_Residential_Status_OthersText']):
        results.append("Self Residential Status 'Relative' Passed")
    elif (Self_Dictionary['Residential Status'] == 'Rented') and (outputdict['Self_Residential_Status_Rented'] == 'Yes'):
        results.append("Self Residential Status 'Rented' Passed")
    elif (Self_Dictionary['Residential Status'] == "Employer's") and (outputdict['Self_Residential_Status_Employers'] == 'Yes'):
        results.append("Self Residential Status 'Employer's' Passed")
    elif (Self_Dictionary['Residential Status'] == 'Others') and (outputdict['Self_Residential_Status_Others'] == 'Yes') and (Self_Dictionary['Other Residential Status'] != None) and (Self_Dictionary['Other Residential Status'] == outputdict['Self_Residential_Status_OthersText']):
        results.append("Self Residential Status 'Others' with Tick and Data Passed")
    elif (Self_Dictionary['Residential Status'] == 'Others') and (outputdict['Self_Residential_Status_Others'] == 'Yes') and (Self_Dictionary['Other Residential Status'] == None) and (outputdict['Self_Residential_Status_OthersText'] == ''):
        results.append("Self Residential Status 'Others' with Tick without Data Passed")
    else:
        results.append("Self Residential Status Failed")     
else:
    results.append("Self Residential Status Data not Available in UI or Input file or Data is Incorrect")

#Self Residential Type Set 1 Validation
if (Self_Dictionary['Residential Type'] != None) and (Self_Dictionary['Residential Type'] == 'HDB' or Self_Dictionary['Residential Type'] == 'HUDC' or Self_Dictionary['Residential Type'] == 'Executive Condominium' or Self_Dictionary['Residential Type'] == 'Private Condominium' or Self_Dictionary['Residential Type'] == 'Apartment'):
    if (Self_Dictionary['Residential Type'] == 'HDB') and (outputdict['Self_Residential_Type_HDB'] == 'Yes'):
        results.append("Self Residential Type 'HDB' Passed")
    elif (Self_Dictionary['Residential Type'] == 'HUDC') and (outputdict['Self_Residential_Type_HUDC'] == 'Yes'):
        results.append("Self Residential Type 'HUDC' Passed")
    elif (Self_Dictionary['Residential Type'] == 'Executive Condominium') and (outputdict['Self_Residential_Type_ExecCondo'] == 'Yes'):
        results.append("Self Residential Type 'Executive Condominium' Passed")        
    elif (Self_Dictionary['Residential Type'] == 'Private Condominium') and (outputdict['Self_Residential_Type_Private'] == 'Yes'):
        results.append("Self Residential Type 'Private Condominium' Passed")
    elif (Self_Dictionary['Residential Type'] == 'Apartment') and (outputdict['Self_Residential_Type_Private'] == 'Yes'):
        results.append("Self Residential Type 'Apartment' Passed")
    else:
        results.append("Self Residential Type Failed")
else:
    results.append("Self Residential Type Data Not Available in UI or Input file or Data is Incorrect")                    

#Self Residential Type Set 2 Validation
if (Self_Dictionary['Residential Type'] != None) and (Self_Dictionary['Residential Type'] == 'Maisonette/Town house' or Self_Dictionary['Residential Type'] == "Terrace's" or Self_Dictionary['Residential Type'] == 'Semi-D' or Self_Dictionary['Residential Type'] == 'Bungalow' or Self_Dictionary['Residential Type'] == 'Others'):
    if (Self_Dictionary['Residential Type'] == 'Maisonette/Town house') and (outputdict['Self_Residential_Type_Landed'] == 'Yes'):
        results.append("Self Residential Type 'Maisonette/Town house' Passed")
    elif (Self_Dictionary['Residential Type'] == "Terrace's") and (outputdict['Self_Residential_Type_Landed'] == 'Yes'):
        results.append("Self Residential Type 'Terrace's' Passed")
    elif (Self_Dictionary['Residential Type'] == 'Semi-D') and (outputdict['Self_Residential_Type_Landed'] == 'Yes'):
        results.append("Self Residential Type 'Semi-D' Passed")        
    elif (Self_Dictionary['Residential Type'] == 'Bungalow') and (outputdict['Self_Residential_Type_Landed'] == 'Yes'):
        results.append("Self Residential Type 'Bungalow' Passed")
    elif (Self_Dictionary['Residential Type'] == 'Others') and (outputdict['Self_Residential_Type_Others'] == 'Yes') and (Self_Dictionary['Other Residential Type'] != None) and (Self_Dictionary['Other Residential Type'] == outputdict['Self_Residential_Type_OthersText']): 
        results.append("Self Residential Type 'Others' with Tick and Data Passed")
    elif (Self_Dictionary['Residential Type'] == 'Others') and (outputdict['Self_Residential_Type_Others'] == 'Yes') and (Self_Dictionary['Other Residential Type'] == None) and (outputdict['Self_Residential_Type_OthersText'] == ''): 
        results.append("Self Residential Type 'Others' with Tick and without Data Passed")
    else:
        results.append("Self Residential Type Failed")
else:
    results.append("Self Residential Type Data Not Available in UI or Input file or Data is Incorrect")                    

#Self Length of Residency Years and Months Validation
if str(Self_Dictionary['Length of Residency_Years']) == outputdict['Self_Length_of_Stay_Current_Residence_Years']:
    results.append("Self Length of Residency 'Years' Passed: Actual output- "+outputdict['Self_Length_of_Stay_Current_Residence_Years']+" Expected output- "+str(Self_Dictionary['Length of Residency_Years']))
else:
    results.append("Self Length of Residency 'Years' Failed: Actual output- "+outputdict['Self_Length_of_Stay_Current_Residence_Years']+" Expected output- "+str(Self_Dictionary['Length of Residency_Years']))

if str(Self_Dictionary['Length of Residency_Months']) == outputdict['Self_Length_of_Stay_Current_Residence_Months']:
    results.append("Self Length of Residency 'Months' Passed: Actual output- "+outputdict['Self_Length_of_Stay_Current_Residence_Months']+" Expected output- "+str(Self_Dictionary['Length of Residency_Months']))
else:
    results.append("Self Length of Residency 'Months' Failed: Actual output- "+outputdict['Self_Length_of_Stay_Current_Residence_Months']+" Expected output- "+str(Self_Dictionary['Length of Residency_Months']))

#Self Employment Status Set 1 Validation
if (Self_Dictionary['Employment Status'] != None) and (Self_Dictionary['Employment Status'] == 'Employee' or Self_Dictionary['Employment Status'] == 'Self Employed' or Self_Dictionary['Employment Status'] == 'Sales/Commission-based' or Self_Dictionary['Employment Status'] == 'Homemaker'):
    if (Self_Dictionary['Employment Status'] == 'Employee') and (outputdict['Self_Employment_Status_Employee'] == 'Yes'):
        results.append("Self Employement Status 'Employee' Passed")
    elif (Self_Dictionary['Employment Status'] == 'Self Employed') and (outputdict['Self_Employment_Status_SelfEmployed'] == 'Yes'):
        results.append("Self Employement Status 'Self Employed' Passed")
    elif (Self_Dictionary['Employment Status'] == 'Sales/Commission-based') and (outputdict['Self_Employment_Status_Employee'] == 'Yes'):
        results.append("Self Employement Status 'Sales/Commission-based' Passed")
    elif (Self_Dictionary['Employment Status'] == 'Homemaker') and (outputdict['Self_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Self Employement Status 'Homemaker' Passed")
    else:
        results.append("Self Employement Status Failed")
else:
    results.append("Self Employement Status Data not Available in UI or Input file or Data is Incorrect")            

#Self Employment Status Set 2 Validation
if (Self_Dictionary['Employment Status'] != None) and (Self_Dictionary['Employment Status'] == 'Retired' or Self_Dictionary['Employment Status'] == 'Student' or Self_Dictionary['Employment Status'] == 'Contract' or Self_Dictionary['Employment Status'] == 'Unemployed'):
    if (Self_Dictionary['Employment Status'] == 'Retired') and (outputdict['Self_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Self Employement Status 'Retired' Passed")
    elif (Self_Dictionary['Employment Status'] == 'Student') and (outputdict['Self_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Self Employement Status 'Student' Passed")
    elif (Self_Dictionary['Employment Status'] == 'Contract') and (outputdict['Self_Employment_Status_Employee'] == 'Yes'):
        results.append("Self Employement Status 'Contract' Passed")
    elif (Self_Dictionary['Employment Status'] == 'Unemployed') and (outputdict['Self_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Self Employement Status 'Unemployed' Passed")
    else:
        results.append("Self Employement Status Failed")
else:
    results.append("Self Employement Status Data not Available in UI or Input file or Data is Incorrect")             

#Self Current Name of Company Validation
if Self_Dictionary['Name of Company'] == outputdict['Self_Name_of_Current_Company']:
    results.append("Self Name of Company Passed")
else:
    results.append("Self Name of Company Failed")

#Self Current Company Address Validation
self_current_company_address = (Self_Dictionary['Office Address Line 1']+", "+Self_Dictionary['Office Address Line 2']+", "+Self_Dictionary['Ocity']+", "+Self_Dictionary['Ostate']+", "+Self_Dictionary['Country Of Office']).lower()
actual_company_address_output = outputdict['Self_Employment_Office_Address'].replace("_","") + outputdict['Self_Employment_Office_Address_Dummy']         

if self_current_company_address == actual_company_address_output:
    results.append("Self Current Company Address Passed")
else:
    results.append("Self Current Company Address Failed")    

#Self Current Company Postal Code Validation
if str(Self_Dictionary['OZip / Postal Code']) == outputdict['Self_Employment_Office_Postal_Code']:
    results.append("Self Current Company Postal Code Passed")
else:
    results.append("Self Current Company Postal Code Failed")    

#Self Current Company Job title Validation 
if Self_Dictionary['Job Title'] == outputdict['Self_Job_Title_Current_Company']:
    results.append("Self Current Company Job Title Passed")
else:
    results.append("Self Current Company Job Title Failed")    

#Self Current Company Industry Type Set 1 Validation
if (Self_Dictionary['Industry Type'] != None) and (Self_Dictionary['Industry Type'] == 'Banking/Finance' or Self_Dictionary['Industry Type'] == 'Building/Construction' or Self_Dictionary['Industry Type'] == 'F&B' or Self_Dictionary['Industry Type'] == 'Government/Stat Board' or Self_Dictionary['Industry Type'] == 'Healthcare' or Self_Dictionary['Industry Type'] == 'Insurance Company'):
    if (Self_Dictionary['Industry Type'] == 'Banking/Finance') and (outputdict['Self_Industry_Current_Company_BankingFinance'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Banking/Finance' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Building/Construction') and (outputdict['Self_Industry_Current_Company_BuildingConstruction'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Building/Construction' Passed")
    elif (Self_Dictionary['Industry Type'] == 'F&B') and (outputdict['Self_Industry_Current_Company_RetailFB'] == 'Yes'):
        results.append("Self Current Company Industry Type 'F&B' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Government/Stat Board') and (outputdict['Self_Industry_Current_Company_GovernmentStatBoard'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Government/Stat Board' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Healthcare') and (outputdict['Self_Industry_Current_Company_Others'] == 'Yes') and (Self_Dictionary['Industry Type'] == outputdict['Self_Industry_Current_Company_OthersText']):
        results.append("Self Current Company Industry Type 'Healthcare' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Insurance Company') and (outputdict['Self_Industry_Current_Company_BankingFinance'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Insurance Company' Passed")
    else:
        results.append("Self Current Company Industry Type Failed")
else:
    results.append("Self Current Company Industry Type Data Not Available in UI or Input File or Data is Incorrect")        

#Self Current Company Industry Type Set 2 Validation
if (Self_Dictionary['Industry Type'] != None) and (Self_Dictionary['Industry Type'] == 'IT/Communications' or Self_Dictionary['Industry Type'] == 'Manufacturing' or Self_Dictionary['Industry Type'] == 'Professional Firm' or Self_Dictionary['Industry Type'] == 'Retail' or Self_Dictionary['Industry Type'] == 'Travel/Hospitality' or Self_Dictionary['Industry Type'] == 'Others'):
    if (Self_Dictionary['Industry Type'] == 'IT/Communications') and (outputdict['Self_Industry_Current_Company_ITCommunications'] == 'Yes'):
        results.append("Self Current Company Industry Type 'IT/Communications' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Manufacturing') and (outputdict['Self_Industry_Current_Company_Manufacturing'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Manufacturing' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Professional Firm') and (outputdict['Self_Industry_Current_Company_Others'] == 'Yes') and (Self_Dictionary['Industry Type'] == outputdict['Self_Industry_Current_Company_OthersText']):
        results.append("Self Current Company Industry Type 'Professional Firm' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Retail') and (outputdict['Self_Industry_Current_Company_RetailFB'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Retail' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Travel/Hospitality') and (outputdict['Self_Industry_Current_Company_TravelHospitality'] == 'Yes'):
        results.append("Self Current Company Industry Type 'Travel/Hospitality' Passed")
    elif (Self_Dictionary['Industry Type'] == 'Others') and (outputdict['Self_Industry_Current_Company_Others'] == 'Yes') and (Self_Dictionary['Other Industry Type'] != None) and (Self_Dictionary['Other Industry Type'] == outputdict['Self_Industry_Current_Company_OthersText']):
        results.append("Self Current Company Industry Type 'Others' with Tick and Data Passed")
    elif (Self_Dictionary['Industry Type'] == 'Others') and (outputdict['Self_Industry_Current_Company_Others'] == 'Yes') and (Self_Dictionary['Other Industry Type'] == None) and (outputdict['Self_Industry_Current_Company_OthersText'] == ''):
        results.append("Self Current Company Industry Type 'Others' with Tick Without Data Passed")
    else:
        results.append("Self Current Company Industry Type Failed")
else:
    results.append("Self Current Company Industry Type Data Not Available in UI or Input File or Data is Incorrect")        

#Self Current Company Level Set 1 Validation
if (Self_Dictionary['Level'] != None) and (Self_Dictionary['Level'] == 'Owner,Director or C-Level' or Self_Dictionary['Level'] == 'Senior Management' or Self_Dictionary['Level'] == 'Middle Management'):
    if (Self_Dictionary['Level'] == 'Owner,Director or C-Level') and (outputdict['Self_Level_Current_Company_SeniorManagement'] == 'Yes'):
        results.append("Self Current Company Level/Position 'Owner,Director or C-Level' Passed")
    elif (Self_Dictionary['Level'] == 'Senior Management') and (outputdict['Self_Level_Current_Company_SeniorManagement'] == 'Yes'):
        results.append("Self Current Company Level/Position 'Senior Management' Passed")
    elif (Self_Dictionary['Level'] == 'Middle Management') and (outputdict['Self_Level_Current_Company_MiddleManagement'] == 'Yes'):
        results.append("Self Current Company Level/Position 'Middle Management' Passed")
    else:
        results.append("Self Current Company Level/Position Failed")
else:
    results.append("Self Current Company Level/Position Data not Available in UI or Input File or Data is Incorrect")                 

#Self Current Company Level Set 2 Validation
if (Self_Dictionary['Level'] != None) and (Self_Dictionary['Level'] == 'Manager or Supervisor' or Self_Dictionary['Level'] == 'Executive' or Self_Dictionary['Level'] == 'Admin or Clerical'):
    if (Self_Dictionary['Level'] == 'Manager or Supervisor') and (outputdict['Self_Level_Current_Company_Professional'] == 'Yes'):
        results.append("Self Current Company Level/Position 'Manager or Supervisor' Passed")
    elif (Self_Dictionary['Level'] == 'Executive') and (outputdict['Self_Level_Current_Company_Executive'] == 'Yes'):
        results.append("Self Current Company Level/Position 'Executive' Passed")
    elif (Self_Dictionary['Level'] == 'Admin or Clerical') and (outputdict['Self_Level_Current_Company_SkilledTradesClerical'] == 'Yes'):
        results.append("Self Current Company Level/Position 'Admin or Clerical' Passed")
    else:
        results.append("Self Current Company Level/Position Failed")
else:
    results.append("Self Current Company Level/Position Data not Available in UI or Input File or Data is Incorrect")                 

#Self Current Company Length of Service Years and Months Validation
if str(Self_Dictionary['Length of Employment_Years']) == outputdict['Self_Service_in_Current_Company_Years']:
    results.append("Self Current Company Length 'Years' Passed")
else:
    results.append("Self Current Company Length 'Years' Failed")

if str(Self_Dictionary['Length of Employment_Months']) == outputdict['Self_Service_in_Current_Company_Months']:
    results.append("Self Current Company Length 'Months' Passed")
else:
    results.append("Self Current Company Length 'Months' Failed")

#Self Annual and Monthly Fixed Income Validation
annual_income = Self_Dictionary['Total(NOA)'] + Self_Dictionary['Rental Income']

if str(addComma(annual_income)) == outputdict['Self_Annual_Income']:
    results.append("Self Annual Income Passed")
else:
    results.append("Self Annual Income Failed")    

if str(addComma(Self_Dictionary['Monthly Fixed Income'])) == outputdict['Self_Basic_Employment_Income_Monthly']:
    results.append("Self Monthly Fixed Income Passed")
else:
    results.append("Self Monthly Fixed Income Failed")

#Self Other Income Validation
other_income = (annual_income/12) - Self_Dictionary['Monthly Fixed Income']

if str(addComma(other_income)) == outputdict['Self_Other_Income']:
    results.append("Self Other Income Passed: Actual output- "+outputdict['Self_Other_Income']+" Expected output- "+str(other_income))
else:
    results.append("Self Other Income Failed: Actual output- "+outputdict['Self_Other_Income']+" Expected output- "+str(other_income))

#Self Name, Length of Service Years and Months of Previous Company Validation
if Self_Dictionary['Name of Previous Company'] == outputdict['Self_Name_of_Previous_Company']:
    results.append("Self Name of Previous Company Passed")
else:
    results.append("Self Name of Previous Company Failed")    

if str(Self_Dictionary['Prev_Length of Employment_Years']) == outputdict['Self_Service_in_Previous_Company_Years']:
    results.append("Self Length of Previous Company 'Years' Passed")
else:
    results.append("Self Length of Previous Company 'Years' Failed")    

if str(Self_Dictionary['Prev_Length of Employment_Months']) == outputdict['Self_Service_in_Previous_Company_Months']:
    results.append("Self Length of Previous Company 'Months' Passed")
else:
    results.append("Self Length of Previous Company 'Months' Failed")    

#Self Other Properties in Singapore/Overseas
if (Self_Dictionary['Do you own other properties in Singapore/Overseas'] != None) and (Self_Dictionary['Do you own other properties in Singapore/Overseas'] == 'No' or Self_Dictionary['Do you own other properties in Singapore/Overseas'] == 'Yes'):
    if (Self_Dictionary['Do you own other properties in Singapore/Overseas'] == 'No') and (outputdict['Self_Own_other_properties_in_Singapore_No'] == 'Yes'):
        results.append("Self Other Properties 'No' Tick Passed")
        if outputdict['Self_Address_of_other_property'] == '':
            results.append("Self Other Properties Blank Address with 'No' Tick Passed")
        else:
            results.append("Self Other Properties Blank Address with 'No' Tick Failed") 
    elif (Self_Dictionary['Do you own other properties in Singapore/Overseas'] == 'Yes') and (outputdict['Self_Own_other_properties_in_Singapore_Yes'] == 'Yes'):
        results.append("Self Other Properties 'Yes' Tick Passed") 
        self_other_properties_address = (Self_Dictionary['Other Properties Property Name']+", "+Self_Dictionary['Other Properties Property Address']+", "+Self_Dictionary['Other Properties Unit No']+", "+Self_Dictionary['Other Properties Country']+", "+Self_Dictionary['Other Properties Zip / Postal Code']).lower() 
        self_other_properties_actual_output = outputdict['Self_Address_of_other_property'].replace("_","")
        if self_other_properties_address == self_other_properties_actual_output:
            results.append("Self Other Properties Address with 'Yes' Passed")
        else:
            results.append("Self Other Properties Address with 'Yes' Failed")    
    else:
        results.append("Self Other Properties Failed")   
else:
    results.append("Self Other Properties Data Not Available in UI or Input File or Data is Incorrect")          
   

############################################################## JOINT APPLICANT ##############################################################

#Joint Salutation Validation
if (Joint_Dictionary['Salutation'] != None) and (Joint_Dictionary['Salutation'] == 'Mr.' or Joint_Dictionary['Salutation'] == 'Mdm.' or Joint_Dictionary['Salutation'] == 'Mrs.' or Joint_Dictionary['Salutation'] == 'Ms.' or Joint_Dictionary['Salutation'] == 'Dr.'):
    if(Joint_Dictionary['Salutation'] == 'Mr.') and (outputdict['Joint_Salutation_Mr'] == 'Yes'):
        results.append("Joint Salutation 'Mr.' Passed")
    elif(Joint_Dictionary['Salutation'] == 'Mdm.') and (outputdict['Joint_Salutation_Mdm'] == 'Yes'):
        results.append("Joint Salutation 'Mdm.' Passed")
    elif(Joint_Dictionary['Salutation'] == 'Mrs.') and (outputdict['Joint_Salutation_Mrs'] == 'Yes'):
        results.append("Joint Salutation 'Mrs.' Passed")
    elif(Joint_Dictionary['Salutation'] == 'Ms.') and (outputdict['Joint_Salutation_Ms'] == 'Yes'):
        results.append("Joint Salutation 'Ms.' Passed")
    elif(Joint_Dictionary['Salutation'] == 'Dr.') and (outputdict['Joint_Salutation_Dr'] == 'Yes'):
        results.append("Joint Salutation 'Dr.' Passed")
    else:
        results.append("Joint Salutation Failed")
else:
    results.append("Joint Salutation Data Not Available in UI or Input File or the Data is Incorrect")

#Joint Passport No. Validation
if Joint_Dictionary['Passport No.'] != None:
    if (Joint_Dictionary['Passport No.'].upper()) == outputdict['Joint_NRICPassport_No']:
        results.append("Joint Passport Data Passed")
    else:
        results.append("Joint Passport Data Failed")   
    if (outputdict['Joint_Passport_Check'] == 'Yes'):
        results.append("Joint Passport Tick Passed")
    else:
        results.append("Joint Passport Tick Failed")        
else:
    results.append("Joint Passport Data Not Available in UI or Input File")

#Joint NRIC No. Validation
if Joint_Dictionary['Nric No.'] != None: 
    if (Joint_Dictionary['Nric No.'].upper()) == outputdict['Joint_NRICPassport_No']:
        results.append("Joint NRIC No. Data Passed")
    else:
        results.append("Joint NRIC No. Data Failed") 
    if outputdict['Joint_NRIC_Check'] == 'Yes':
        results.append("Joint NRIC No. Tick Passed")
    else:
        results.append("Joint NRIC No. Tick Failed")           
else:
    results.append("Joint NRIC No. Data Not Available in UI or Input File")

#Joint Marital Status
if (Joint_Dictionary['Marital Status'] != None) and (Joint_Dictionary['Marital Status'] == 'Single' or Joint_Dictionary['Marital Status'] == 'Married' or Joint_Dictionary['Marital Status'] == 'Divorced' or Joint_Dictionary['Marital Status'] == 'Widowed' or Joint_Dictionary['Marital Status'] == 'Others'):
    if (Joint_Dictionary['Marital Status'] == 'Single') and (outputdict['Joint_Marital_Status_Single'] == 'Yes'):
        results.append("Joint Marital Status 'Single' Passed")
    elif (Joint_Dictionary['Marital Status'] == 'Married') and (outputdict['Joint_Marital_Status_Married'] == 'Yes'):
        results.append("Joint Marital Status 'Married' Passed")
    elif (Joint_Dictionary['Marital Status'] == 'Divorced') and (outputdict['Joint_Marital_Status_Divorced'] == 'Yes'):
        results.append("Joint Marital Status 'Divorced' Passed")
    elif (Joint_Dictionary['Marital Status'] == 'Widowed') and (outputdict['Joint_Marital_Status_OthersCheck'] == 'Yes') and (outputdict['Joint_Marital_Status_OthersText'] == 'Widowed'):
        results.append("Joint Marital Status 'Widowed' Passed")
    elif (Joint_Dictionary['Marital Status'] == 'Others') and (outputdict['Joint_Marital_Status_OthersCheck'] == 'Yes') and (Joint_Dictionary['Other Marital Status'] != None) (Joint_Dictionary['Other Marital Status'] == outputdict['Joint_Marital_Status_OthersText']):
        results.append("Joint Marital Status 'Others' with data and tick Passed")
    elif (Joint_Dictionary['Marital Status'] == 'Others') and (outputdict['Joint_Marital_Status_OthersCheck']) and (Joint_Dictionary['Marital Status'] == None) and (outputdict['Joint_Marital_Status_OthersText'] == ''):
        results.append("Joint Marital Status 'Others' without data with tick Passed")    
    else:
        results.append("Joint Marital Status Failed")
else:
    results.append("Joint Marital Status Data Not Available in UI or Input File or the Data is Incorrect")        

#Joint No. of Dependents Validation
if (str(Joint_Dictionary['No. Of Dependents']) == outputdict['Joint_Number_of_Dependents']):
    results.append("Joint No. Of Dependents Passed")
else:
    results.append("Joint No. Of Dependents Failed: Actual result- "+outputdict['Joint_Number_of_Dependents']+" Expected Result- "+str(Joint_Dictionary['No. Of Dependents']))    

#Joint Highest Education Set 1 Validation
if (Joint_Dictionary['Highest Education'] != None) and (Joint_Dictionary['Highest Education'] == 'No Formal Education' or Joint_Dictionary['Highest Education'] == 'Primary' or Joint_Dictionary['Highest Education'] == 'Secondary' or Joint_Dictionary['Highest Education'] == 'O Levels' or Joint_Dictionary['Highest Education'] == 'N Levels' or Joint_Dictionary['Highest Education'] == 'A Levels'):
    if (Joint_Dictionary['Highest Education'] == "No Formal Education") and (outputdict['Joint_Highest_Education_Primary'] == 'Yes'):
        results.append("Joint Highest Education 'No Formal Education' Passed")
    elif (Joint_Dictionary['Highest Education'] == "Primary") and (outputdict['Joint_Highest_Education_Primary'] == 'Yes'):
        results.append("Joint Highest Education 'Primary' Passed")
    elif (Joint_Dictionary['Highest Education'] == "Secondary") and (outputdict['Joint_Highest_Education_Secondary'] == 'Yes'):
        results.append("Joint Highest Education 'Secondary' Passed")    
    elif (Joint_Dictionary['Highest Education'] == "O Levels") and (outputdict['Joint_Highest_Education_Secondary'] == 'Yes'):
        results.append("Joint Highest Education 'O Levels' Passed")    
    elif (Joint_Dictionary['Highest Education'] == "N Levels") and (outputdict['Joint_Highest_Education_Secondary'] == 'Yes'):
        results.append("Joint Highest Education 'N Levels' Passed")
    elif (Joint_Dictionary['Highest Education'] == "A Levels") and (outputdict['Joint_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Joint Highest Education 'A Levels' Passed")          
    else:
        results.append("Joint Highest Education Failed")
else:
    results.append("Joint Highest Education Data Not Available in UI or Input File or the Data is Incorrect")

#Joint Highest Education Set 2 Validation
if (Joint_Dictionary['Highest Education'] != None) and (Joint_Dictionary['Highest Education'] == 'NITEC/Higher NITEC' or Joint_Dictionary['Highest Education'] == 'Pre-University' or Joint_Dictionary['Highest Education'] == 'Diploma' or Joint_Dictionary['Highest Education'] == 'Degree' or Joint_Dictionary['Highest Education'] == 'Post-Graduate' or Joint_Dictionary['Highest Education'] == 'Technical Certification'):
    if (Joint_Dictionary['Highest Education'] == "NITEC/Higher NITEC") and (outputdict['Joint_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Joint Highest Education 'NITEC/Higher NITEC' Passed")
    elif (Joint_Dictionary['Highest Education'] == "Pre-University") and (outputdict['Joint_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Joint Highest Education 'Pre-University' Passed")
    elif (Joint_Dictionary['Highest Education'] == "Diploma") and (outputdict['Joint_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Joint Highest Education 'Diploma' Passed")    
    elif (Joint_Dictionary['Highest Education'] == "Degree") and (outputdict['Joint_Highest_Education_Degree'] == 'Yes'):
        results.append("Joint Highest Education 'Degree' Passed")    
    elif (Joint_Dictionary['Highest Education'] == "Post-Graduate") and (outputdict['Joint_Highest_Education_Post_Graduate'] == 'Yes'):
        results.append("Joint Highest Education 'Post-Graduate' Passed")
    elif (Joint_Dictionary['Highest Education'] == "Technical Certification") and (outputdict['Joint_Highest_Education_Diploma_Or_PreUniversity'] == 'Yes'):
        results.append("Joint Highest Education 'Technical Certification' Passed")          
    else:
        results.append("Joint Highest Education Failed")
else:
    results.append("Joint Highest Education Data Not Available in UI or Input File or the Data is Incorrect")    

#Joint Residential Address Same as NRIC Validation
if (Joint_Dictionary['Residential Address same as NRIC '] != None) and (Joint_Dictionary['Residential Address same as NRIC '] == 'No' or Joint_Dictionary['Residential Address same as NRIC '] == 'Yes'):
    if (Joint_Dictionary['Residential Address same as NRIC '] == 'No') and (outputdict['Joint_Residential_Address_same_as_NRIC_No'] == 'Yes'):
        results.append("Joint Residential Address same as NRIC 'No' Passed")
    elif (Joint_Dictionary['Residential Address same as NRIC '] == 'Yes') and (outputdict['Joint_Residential_Address_same_as_NRIC_Yes'] == 'Yes'):
        results.append("Joint Residential Address same as NRIC 'Yes' Passed")
    else:
        results.append("Joint Residential Address same as NRIC Failed")
else:
    results.append("Joint Residential Address same as NRIC Data not Available in UI or Input file or Data is Incorrect")            

#Joint Residential Address Validation
if Joint_Dictionary['Country Of Residence'] != 'Singapore':
    Joint_Residential_Address = (Joint_Dictionary['Residential Address Line 1']+", "+Joint_Dictionary['Residential Address Line 2']+", "+Joint_Dictionary['City']+", "+Joint_Dictionary['State']+", "+Joint_Dictionary['Country Of Residence']).lower()
    Joint_actual_residential_output = outputdict['Joint_Residential_Address'].replace("_","") + outputdict['Joint_Residential_Address_Dummy']

    if Joint_Residential_Address == Joint_actual_residential_output:
        results.append("Joint Residential Address Passed")
    else:
        results.append("Joint Residential Address Failed")   
elif Joint_Dictionary['Country Of Residence'] == 'Singapore':
    Joint_Residential_Address_S = (Joint_Dictionary['Residential Address']+", "+Joint_Dictionary['Unit No']+", "+Joint_Dictionary['Country Of Residence']).lower()
    Joint_actual_residential_output_S = outputdict['Joint_Residential_Address'].replace("_","") + outputdict['Joint_Residential_Address_Dummy']

    if Joint_Residential_Address_S == Joint_actual_residential_output_S:
        results.append("Joint Residential Address Passed")
    else:
        results.append("Joint Residential Address Failed")          

#Joint Residential Address Postal Code Validation
if str(Joint_Dictionary['Zip / Postal Code']) == outputdict['Joint_Home_Postal_Code']:
    results.append("Joint Residential Postal Code Passed")
else:
    results.append("Joint Residential Postal Code Failed")      

#Joint Mailing Address Validation
if Joint_Dictionary['Country Of Office'] != 'Singapore':
    Joint_Mailing_Address = (Joint_Dictionary['Office Address Line 1']+", "+Joint_Dictionary['Office Address Line 2']+", "+Joint_Dictionary['Ocity']+", "+Joint_Dictionary['Ostate']+", "+Joint_Dictionary['Country Of Office']).lower()
    Joint_actual_mailing_output = outputdict['Joint_Mailing_Address'].replace("_","") + outputdict['Joint_Mailing_Address_Dummy']

    if Joint_Mailing_Address == Joint_actual_mailing_output:
        results.append("Joint Mailing Address Passed")
    else:
        results.append("Joint Mailing Address Failed")    
elif Joint_Dictionary['Country Of Office'] == 'Singapore':
    Joint_Mailing_Address_S = (Joint_Dictionary['Office Address']+", "+Joint_Dictionary['OUnit No']+", "+Joint_Dictionary['Country Of Office']).lower()
    Joint_actual_mailing_output_S = outputdict['Joint_Mailing_Address'].replace("_","") + outputdict['Joint_Mailing_Address_Dummy']

    if Joint_Mailing_Address_S == Joint_actual_mailing_output_S:
        results.append("Joint Mailing Address Passed")
    else:
        results.append("Joint Mailing Address Failed")

#Joint Mailing Address Postal Code Validation
if str(Joint_Dictionary['OZip / Postal Code']) == outputdict['Joint_Mailing_Postal_Code']:
    results.append("Joint Mailing Postal Code Passed")
else:
    results.append("Joint Mailing Postal Code Failed")
 
#Joint Mobile, Home, Office Phone Number Validation 
Joint_mobile_number = "+"+Joint_Dictionary['MCountry Code'].split("+")[1]+"-"+str(Joint_Dictionary['Mobile Number'])
Joint_home_phone = "+"+Joint_Dictionary['HCountry Code'].split("+")[1]+"-"+str(Joint_Dictionary['Home Phone'])
Joint_office_phone = "+"+Joint_Dictionary['OCountry Code'].split("+")[1]+"-"+str(Joint_Dictionary['Office Phone'])

if Joint_mobile_number == outputdict['Joint_Hand_phone']:
    results.append("Joint Mobile No. Passed: Actual output- "+outputdict['Joint_Hand_phone']+" Expected output- "+Joint_mobile_number)
else:
    results.append("Joint Mobile No. Failed: Actual output- "+outputdict['Joint_Hand_phone']+" Expected output- "+Joint_mobile_number)

if Joint_home_phone == outputdict['Joint_Home_phone']:
    results.append("Joint Home Phone Passed: Actual output- "+outputdict['Joint_Home_phone']+" Expected output- "+Joint_home_phone)
else:
    results.append("Joint Home Phone Failed: Actual output- "+outputdict['Joint_Home_phone']+" Expected output- "+Joint_home_phone)

if Joint_office_phone == outputdict['Joint_Office_phone']:
    results.append("Joint Office Phone Passed: Actual output- "+outputdict['Joint_Office_phone']+" Expected output- "+Joint_office_phone)
else:
    results.append("Joint Office Phone Failed: Actual output- "+outputdict['Joint_Office_phone']+" Expected output- "+Joint_office_phone)

#Joint Email Address Validation
if Joint_Dictionary['Email Address'] == outputdict['Joint_Email_address']:
    results.append("Joint Email Address Passed")
else:
    results.append("Joint Email Address Failed")  

#Joint Residential Status Set 1 Validation
if (Joint_Dictionary['Residential Status'] != None) and (Joint_Dictionary['Residential Status'] == 'Fully Paid' or Joint_Dictionary['Residential Status'] == 'Oustanding Mortgage' or Joint_Dictionary['Residential Status'] == 'Live with family' or Joint_Dictionary['Residential Status'] == 'Parents'):
    if (Joint_Dictionary['Residential Status'] == 'Fully Paid') and (outputdict['Joint_Residential_Status_Fully_Owned'] == 'Yes'):
        results.append("Joint Residential Status 'Fully Paid' Passed")
    elif (Joint_Dictionary['Residential Status'] == 'Oustanding Mortgage') and (outputdict['Joint_Residential_Status_Mortgaged'] == 'Yes'):
        results.append("Joint Residential Status 'Oustanding Mortgage' Passed")
    elif (Joint_Dictionary['Residential Status'] == 'Live with family') and (outputdict['Joint_Residential_Status_Others'] == 'Yes') and (Joint_Dictionary['Residential Status'] == outputdict['Joint_Residential_Status_OthersText']):
        results.append("Joint Residential Status 'Live with family'   Passed")
    elif (Joint_Dictionary['Residential Status'] == 'Parents') and (outputdict['Joint_Residential_Status_Parents'] == 'Yes'):
        results.append("Joint Residential Status 'Parents' Passed")
    else:
        results.append("Joint Residential Status Failed")     
else:
    results.append("Joint Residential Status Data not Available in UI or Input file or Data is Incorrect")

#Joint Residential Status Set 2 Validation
if (Joint_Dictionary['Residential Status'] != None) and (Joint_Dictionary['Residential Status'] == 'Relative' or Joint_Dictionary['Residential Status'] == 'Rented' or Joint_Dictionary['Residential Status'] == "Employer's" or Joint_Dictionary['Residential Status'] == 'Others'):
    if (Joint_Dictionary['Residential Status'] == 'Relative') and (outputdict['Joint_Residential_Status_Others'] == 'Yes') and (Joint_Dictionary['Residential Status'] == outputdict['Joint_Residential_Status_OthersText']):
        results.append("Joint Residential Status 'Relative' Passed")
    elif (Joint_Dictionary['Residential Status'] == 'Rented') and (outputdict['Joint_Residential_Status_Rented'] == 'Yes'):
        results.append("Joint Residential Status 'Rented' Passed")
    elif (Joint_Dictionary['Residential Status'] == "Employer's") and (outputdict['Joint_Residential_Status_Employers'] == 'Yes'):
        results.append("Joint Residential Status 'Employer's' Passed")
    elif (Joint_Dictionary['Residential Status'] == 'Others') and (outputdict['Joint_Residential_Status_Others'] == 'Yes') and (Joint_Dictionary['Other Residential Status'] != None) and (Joint_Dictionary['Other Residential Status'] == outputdict['Joint_Residential_Status_OthersText']):
        results.append("Joint Residential Status 'Others' with Tick and Data Passed")
    elif (Joint_Dictionary['Residential Status'] == 'Others') and (outputdict['Joint_Residential_Status_Others'] == 'Yes') and (Joint_Dictionary['Other Residential Status'] == None) and (outputdict['Joint_Residential_Status_OthersText'] == ''):
        results.append("Joint Residential Status 'Others' with Tick without Data Passed")
    else:
        results.append("Joint Residential Status Failed")     
else:
    results.append("Joint Residential Status Data not Available in UI or Input file or Data is Incorrect")

#Joint Residential Type Set 1 Validation
if (Joint_Dictionary['Residential Type'] != None) and (Joint_Dictionary['Residential Type'] == 'HDB' or Joint_Dictionary['Residential Type'] == 'HUDC' or Joint_Dictionary['Residential Type'] == 'Executive Condominium' or Joint_Dictionary['Residential Type'] == 'Private Condominium' or Joint_Dictionary['Residential Type'] == 'Apartment'):
    if (Joint_Dictionary['Residential Type'] == 'HDB') and (outputdict['Joint_Residential_Type_HDB'] == 'Yes'):
        results.append("Joint Residential Type 'HDB' Passed")
    elif (Joint_Dictionary['Residential Type'] == 'HUDC') and (outputdict['Joint_Residential_Type_HUDC'] == 'Yes'):
        results.append("Joint Residential Type 'HUDC' Passed")
    elif (Joint_Dictionary['Residential Type'] == 'Executive Condominium') and (outputdict['Joint_Residential_Type_ExecCondo'] == 'Yes'):
        results.append("Joint Residential Type 'Executive Condominium' Passed")        
    elif (Joint_Dictionary['Residential Type'] == 'Private Condominium') and (outputdict['Joint_Residential_Type_Private'] == 'Yes'):
        results.append("Joint Residential Type 'Private Condominium' Passed")
    elif (Joint_Dictionary['Residential Type'] == 'Apartment') and (outputdict['Joint_Residential_Type_Private'] == 'Yes'):
        results.append("Joint Residential Type 'Apartment' Passed")
    else:
        results.append("Joint Residential Type Failed")
else:
    results.append("Joint Residential Type Data Not Available in UI or Input file or Data is Incorrect")                    

#Joint Residential Type Set 2 Validation
if (Joint_Dictionary['Residential Type'] != None) and (Joint_Dictionary['Residential Type'] == 'Maisonette/Town house' or Joint_Dictionary['Residential Type'] == "Terrace's" or Joint_Dictionary['Residential Type'] == 'Semi-D' or Joint_Dictionary['Residential Type'] == 'Bungalow' or Joint_Dictionary['Residential Type'] == 'Others'):
    if (Joint_Dictionary['Residential Type'] == 'Maisonette/Town house') and (outputdict['Joint_Residential_Type_Landed'] == 'Yes'):
        results.append("Joint Residential Type 'Maisonette/Town house' Passed")
    elif (Joint_Dictionary['Residential Type'] == "Terrace's") and (outputdict['Joint_Residential_Type_Landed'] == 'Yes'):
        results.append("Joint Residential Type 'Terrace's' Passed")
    elif (Joint_Dictionary['Residential Type'] == 'Semi-D') and (outputdict['Joint_Residential_Type_Landed'] == 'Yes'):
        results.append("Joint Residential Type 'Semi-D' Passed")        
    elif (Joint_Dictionary['Residential Type'] == 'Bungalow') and (outputdict['Joint_Residential_Type_Landed'] == 'Yes'):
        results.append("Joint Residential Type 'Bungalow' Passed")
    elif (Joint_Dictionary['Residential Type'] == 'Others') and (outputdict['Joint_Residential_Type_Others'] == 'Yes') and (Joint_Dictionary['Other Residential Type'] != None) and (Joint_Dictionary['Other Residential Type'] == outputdict['Joint_Residential_Type_OthersText']): 
        results.append("Joint Residential Type 'Others' with Tick and Data Passed")
    elif (Joint_Dictionary['Residential Type'] == 'Others') and (outputdict['Joint_Residential_Type_Others'] == 'Yes') and (Joint_Dictionary['Other Residential Type'] == None) and (outputdict['Joint_Residential_Type_OthersText'] == ''): 
        results.append("Joint Residential Type 'Others' with Tick and without Data Passed")
    else:
        results.append("Joint Residential Type Failed")
else:
    results.append("Joint Residential Type Data Not Available in UI or Input file or Data is Incorrect")                    

#Joint Length of Residency Years and Months Validation
if str(Joint_Dictionary['Length of Residency_Years']) == outputdict['Joint_Length_of_Stay_Current_Residence_Years']:
    results.append("Joint Length of Residency 'Years' Passed: Actual output- "+outputdict['Joint_Length_of_Stay_Current_Residence_Years']+" Expected output- "+str(Joint_Dictionary['Length of Residency_Years']))
else:
    results.append("Joint Length of Residency 'Years' Failed: Actual output- "+outputdict['Joint_Length_of_Stay_Current_Residence_Years']+" Expected output- "+str(Joint_Dictionary['Length of Residency_Years']))

if str(Joint_Dictionary['Length of Residency_Months']) == outputdict['Joint_Length_of_Stay_Current_Residence_Months']:
    results.append("Joint Length of Residency 'Months' Passed: Actual output- "+outputdict['Joint_Length_of_Stay_Current_Residence_Months']+" Expected output- "+str(Joint_Dictionary['Length of Residency_Months']))
else:
    results.append("Joint Length of Residency 'Months' Failed: Actual output- "+outputdict['Joint_Length_of_Stay_Current_Residence_Months']+" Expected output- "+str(Joint_Dictionary['Length of Residency_Months']))

#Joint Employment Status Set 1 Validation
if (Joint_Dictionary['Employment Status'] != None) and (Joint_Dictionary['Employment Status'] == 'Employee' or Joint_Dictionary['Employment Status'] == 'Self Employed' or Joint_Dictionary['Employment Status'] == 'Sales/Commission-based' or Joint_Dictionary['Employment Status'] == 'Homemaker'):
    if (Joint_Dictionary['Employment Status'] == 'Employee') and (outputdict['Joint_Employment_Status_Employee'] == 'Yes'):
        results.append("Joint Employement Status 'Employee' Passed")
    elif (Joint_Dictionary['Employment Status'] == 'Self Employed') and (outputdict['Joint_Employment_Status_SelfEmployed'] == 'Yes'):
        results.append("Joint Employement Status 'Self Employed' Passed")
    elif (Joint_Dictionary['Employment Status'] == 'Sales/Commission-based') and (outputdict['Joint_Employment_Status_Employee'] == 'Yes'):
        results.append("Joint Employement Status 'Sales/Commission-based' Passed")
    elif (Joint_Dictionary['Employment Status'] == 'Homemaker') and (outputdict['Joint_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Joint Employement Status 'Homemaker' Passed")
    else:
        results.append("Joint Employement Status Failed")
else:
    results.append("Joint Employement Status Data not Available in UI or Input file or Data is Incorrect")            

#Joint Employment Status Set 2 Validation
if (Joint_Dictionary['Employment Status'] != None) and (Joint_Dictionary['Employment Status'] == 'Retired' or Joint_Dictionary['Employment Status'] == 'Student' or Joint_Dictionary['Employment Status'] == 'Contract' or Joint_Dictionary['Employment Status'] == 'Unemployed'):
    if (Joint_Dictionary['Employment Status'] == 'Retired') and (outputdict['Joint_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Joint Employement Status 'Retired' Passed")
    elif (Joint_Dictionary['Employment Status'] == 'Student') and (outputdict['Joint_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Joint Employement Status 'Student' Passed")
    elif (Joint_Dictionary['Employment Status'] == 'Contract') and (outputdict['Joint_Employment_Status_Employee'] == 'Yes'):
        results.append("Joint Employement Status 'Contract' Passed")
    elif (Joint_Dictionary['Employment Status'] == 'Unemployed') and (outputdict['Joint_Employment_Status_Unemployed'] == 'Yes'):
        results.append("Joint Employement Status 'Unemployed' Passed")
    else:
        results.append("Joint Employement Status Failed")
else:
    results.append("Joint Employement Status Data not Available in UI or Input file or Data is Incorrect")             

#Joint Name of Current Company Validation
if Joint_Dictionary['Name of Company'] == outputdict['Joint_Name_of_Current_Company']:
    results.append("Joint Name of Company Passed")
else:
    results.append("Joint Name of Company Failed")

#Joint Address of Current Company Validation
Joint_current_company_address = (Joint_Dictionary['Office Address Line 1']+", "+Joint_Dictionary['Office Address Line 2']+", "+Joint_Dictionary['Ocity']+", "+Joint_Dictionary['Ostate']+", "+Joint_Dictionary['Country Of Office']).lower()
Joint_actual_company_address_output = outputdict['Joint_Employment_Office_Address'].replace("_","") + outputdict['Joint_Employment_Office_Address_Dummy']         

if Joint_current_company_address == Joint_actual_company_address_output:
    results.append("Joint Current Company Address Passed")
else:
    results.append("Joint Current Company Address Failed")    

#Joint Current Company Address Postal Code Validation
if str(Joint_Dictionary['OZip / Postal Code']) == outputdict['Joint_Employment_Office_Postal_Code']:
    results.append("Joint Current Company Postal Code Passed")
else:
    results.append("Joint Current Company Postal Code Failed")    

#Joint Current Company Job Title Validation
if Joint_Dictionary['Job Title'] == outputdict['Joint_Job_Title_Current_Company']:
    results.append("Joint Current Company Job Title Passed")
else:
    results.append("Joint Current Company Job Title Failed")    

#Joint Current Company Industry Type Set 1 Validation
if (Joint_Dictionary['Industry Type'] != None) and (Joint_Dictionary['Industry Type'] == 'Banking/Finance' or Joint_Dictionary['Industry Type'] == 'Building/Construction' or Joint_Dictionary['Industry Type'] == 'F&B' or Joint_Dictionary['Industry Type'] == 'Government/Stat Board' or Joint_Dictionary['Industry Type'] == 'Healthcare' or Joint_Dictionary['Industry Type'] == 'Insurance Company'):
    if (Joint_Dictionary['Industry Type'] == 'Banking/Finance') and (outputdict['Joint_Industry_Current_Company_BankingFinance'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Banking/Finance' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Building/Construction') and (outputdict['Joint_Industry_Current_Company_BuildingConstruction'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Building/Construction' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'F&B') and (outputdict['Joint_Industry_Current_Company_RetailFB'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'F&B' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Government/Stat Board') and (outputdict['Joint_Industry_Current_Company_GovernmentStatBoard'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Government/Stat Board' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Healthcare') and (outputdict['Joint_Industry_Current_Company_Others'] == 'Yes') and (Joint_Dictionary['Industry Type'] == outputdict['Joint_Industry_Current_Company_OthersText']):
        results.append("Joint Current Company Industry Type 'Healthcare' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Insurance Company') and (outputdict['Joint_Industry_Current_Company_BankingFinance'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Insurance Company' Passed")
    else:
        results.append("Joint Current Company Industry Type Failed")
else:
    results.append("Joint Current Company Industry Type Data Not Available in UI or Input File or Data is Incorrect")        

#Joint Current Company Industry Type Set 2 Validation
if (Joint_Dictionary['Industry Type'] != None) and (Joint_Dictionary['Industry Type'] == 'IT/Communications' or Joint_Dictionary['Industry Type'] == 'Manufacturing' or Joint_Dictionary['Industry Type'] == 'Professional Firm' or Joint_Dictionary['Industry Type'] == 'Retail' or Joint_Dictionary['Industry Type'] == 'Travel/Hospitality' or Joint_Dictionary['Industry Type'] == 'Others'):
    if (Joint_Dictionary['Industry Type'] == 'IT/Communications') and (outputdict['Joint_Industry_Current_Company_ITCommunications'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'IT/Communications' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Manufacturing') and (outputdict['Joint_Industry_Current_Company_Manufacturing'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Manufacturing' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Professional Firm') and (outputdict['Joint_Industry_Current_Company_Others'] == 'Yes') and (Joint_Dictionary['Industry Type'] == outputdict['Joint_Industry_Current_Company_OthersText']):
        results.append("Joint Current Company Industry Type 'Professional Firm' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Retail') and (outputdict['Joint_Industry_Current_Company_RetailFB'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Retail' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Travel/Hospitality') and (outputdict['Joint_Industry_Current_Company_TravelHospitality'] == 'Yes'):
        results.append("Joint Current Company Industry Type 'Travel/Hospitality' Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Others') and (outputdict['Joint_Industry_Current_Company_Others'] == 'Yes') and (Joint_Dictionary['Other Industry Type'] != None) and (Joint_Dictionary['Other Industry Type'] == outputdict['Joint_Industry_Current_Company_OthersText']):
        results.append("Joint Current Company Industry Type 'Others' with Tick and Data Passed")
    elif (Joint_Dictionary['Industry Type'] == 'Others') and (outputdict['Joint_Industry_Current_Company_Others'] == 'Yes') and (Joint_Dictionary['Other Industry Type'] == None) and (outputdict['Joint_Industry_Current_Company_OthersText'] == ''):
        results.append("Joint Current Company Industry Type 'Others' with Tick Without Data Passed")
    else:
        results.append("Joint Current Company Industry Type Failed")
else:
    results.append("Joint Current Company Industry Type Data Not Available in UI or Input File or Data is Incorrect")        

#Joint Current Company Level Set 1 Validation
if (Joint_Dictionary['Level'] != None) and (Joint_Dictionary['Level'] == 'Owner,Director or C-Level' or Joint_Dictionary['Level'] == 'Senior Management' or Joint_Dictionary['Level'] == 'Middle Management'):
    if (Joint_Dictionary['Level'] == 'Owner,Director or C-Level') and (outputdict['Joint_Level_Current_Company_SeniorManagement'] == 'Yes'):
        results.append("Joint Current Company Level/Position 'Owner,Director or C-Level' Passed")
    elif (Joint_Dictionary['Level'] == 'Senior Management') and (outputdict['Joint_Level_Current_Company_SeniorManagement'] == 'Yes'):
        results.append("Joint Current Company Level/Position 'Senior Management' Passed")
    elif (Joint_Dictionary['Level'] == 'Middle Management') and (outputdict['Joint_Level_Current_Company_MiddleManagement'] == 'Yes'):
        results.append("Joint Current Company Level/Position 'Middle Management' Passed")
    else:
        results.append("Joint Current Company Level/Position Failed")
else:
    results.append("Joint Current Company Level/Position Data not Available in UI or Input File or Data is Incorrect")                 

#Joint Current Company Level Set 2 Validation
if (Joint_Dictionary['Level'] != None) and (Joint_Dictionary['Level'] == 'Manager or Supervisor' or Joint_Dictionary['Level'] == 'Executive' or Joint_Dictionary['Level'] == 'Admin or Clerical'):
    if (Joint_Dictionary['Level'] == 'Manager or Supervisor') and (outputdict['Joint_Level_Current_Company_Professional'] == 'Yes'):
        results.append("Joint Current Company Level/Position 'Manager or Supervisor' Passed")
    elif (Joint_Dictionary['Level'] == 'Executive') and (outputdict['Joint_Level_Current_Company_Executive'] == 'Yes'):
        results.append("Joint Current Company Level/Position 'Executive' Passed")
    elif (Joint_Dictionary['Level'] == 'Admin or Clerical') and (outputdict['Joint_Level_Current_Company_SkilledTradesClerical'] == 'Yes'):
        results.append("Joint Current Company Level/Position 'Admin or Clerical' Passed")
    else:
        results.append("Joint Current Company Level/Position Failed")
else:
    results.append("Joint Current Company Level/Position Data not Available in UI or Input File or Data is Incorrect")                 

#Joint Current Company Length of Service Years and Months Validation
if str(Joint_Dictionary['Length of Employment_Years']) == outputdict['Joint_Service_in_Current_Company_Years']:
    results.append("Joint Current Company Length 'Years' Passed")
else:
    results.append("Joint Current Company Length 'Years' Failed")

if str(Joint_Dictionary['Length of Employment_Months']) == outputdict['Joint_Service_in_Current_Company_Months']:
    results.append("Joint Current Company Length 'Months' Passed")
else:
    results.append("Joint Current Company Length 'Months' Failed")

#Joint Annual, Monthly fixed Income Validation
joint_annual_income = Joint_Dictionary['Total(NOA)'] + Joint_Dictionary['Rental Income']

if str(addComma(joint_annual_income)) == outputdict['Joint_Annual_Income']:
    results.append("Joint Annual Income Passed")
else:
    results.append("Joint Annual Income Failed")    

if str(addComma(Joint_Dictionary['Monthly Fixed Income'])) == outputdict['Joint_Basic_Employment_Income_Monthly']:
    results.append("Joint Monthly Fixed Income Passed")
else:
    results.append("Joint Monthly Fixed Income Failed")

#Joint Other Income Validation
joint_other_income = (joint_annual_income/12) - Joint_Dictionary['Monthly Fixed Income']

if str(addComma(format(float(joint_other_income),'.2f'))) == outputdict['Joint_Other_Income']:
    results.append("Joint Other Income Passed: Actual output- "+outputdict['Joint_Other_Income']+" Expected output- "+str(addComma(format(float(joint_other_income),'.2f'))))
else:
    results.append("Joint Other Income Failed: Actual output- "+outputdict['Joint_Other_Income']+" Expected output- "+str(addComma(format(float(joint_other_income),'.2f'))))

#Joint Name, Length of Service Years and Months of Previous Company
if Joint_Dictionary['Name of Previous Company'] == outputdict['Joint_Name_of_Previous_Company']:
    results.append("Joint Name of Previous Company Passed")
else:
    results.append("Joint Name of Previous Company Failed")    

if str(Joint_Dictionary['Prev_Length of Employment_Years']) == outputdict['Joint_Service_in_Previous_Company_Years']:
    results.append("Joint Length of Previous Company 'Years' Passed")
else:
    results.append("Joint Length of Previous Company 'Years' Failed")    

if str(Joint_Dictionary['Prev_Length of Employment_Months']) == outputdict['Joint_Service_in_Previous_Company_Months']:
    results.append("Joint Length of Previous Company 'Months' Passed")
else:
    results.append("Joint Length of Previous Company 'Months' Failed")    

#Joint Other Properties in Singapore/Overseas
if (Joint_Dictionary['Do you own other properties in Singapore/Overseas'] != None) and (Joint_Dictionary['Do you own other properties in Singapore/Overseas'] == 'No' or Joint_Dictionary['Do you own other properties in Singapore/Overseas'] == 'Yes'):
    if (Joint_Dictionary['Do you own other properties in Singapore/Overseas'] == 'No') and (outputdict['Joint_Own_other_properties_in_Singapore_No'] == 'Yes'):
        results.append("Joint Other Properties 'No' Tick Passed")
        if outputdict['Joint_Address_of_other_property'] == '':
            results.append("Joint Other Properties Blank Address with 'No' Tick Passed")
        else:
            results.append("Joint Other Properties Blank Address with 'No' Tick Failed") 
    elif (Joint_Dictionary['Do you own other properties in Singapore/Overseas'] == 'Yes') and (outputdict['Joint_Own_other_properties_in_Singapore_Yes'] == 'Yes'):
        results.append("Joint Other Properties 'Yes' Tick Passed") 
        Joint_other_properties_address = (Joint_Dictionary['Other Properties Property Name']+", "+Joint_Dictionary['Other Properties Property Address']+", "+Joint_Dictionary['Other Properties Unit No']+", "+Joint_Dictionary['Other Properties Country']+", "+Joint_Dictionary['Other Properties Zip / Postal Code']).lower() 
        Joint_other_properties_actual_output = outputdict['Joint_Address_of_other_property'].replace("_","")
        if Joint_other_properties_address == Joint_other_properties_actual_output:
            results.append("Joint Other Properties Address with 'Yes' Passed")
        else:
            results.append("Joint Other Properties Address with 'Yes' Failed")    
    else:
        results.append("Joint Other Properties Failed")   
else:
    results.append("Joint Other Properties Data Not Available in UI or Input File or Data is Incorrect") 


#Preparing Results File 
writesheet = openpyxl.Workbook()
Wsheet = writesheet.active
Wsheet.column_dimensions['A'].width = float(100)
if len(results) > 0:
    j = 1
    for i in range(len(results)):
        Wsheet.cell(j,1).value = results[i]
        j+=1

writesheet.save("D:\\Automation Testing\\MayBank\\Personal & Financial Details\\Results\\results.xlsx")   