from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time
import datetime
from datetime import date


path = "D:\\Automation Testing\\MayBank\\Loan\\loan_refinance_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
for i in range(1,50):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)
driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_xpath('//*[@id="loansTable"]/tbody/tr[1]/td[8]/div/button').click()
time.sleep(3)
element = driver.find_element_by_xpath('//*[@id="applicationIdField"]').text
driver.quit()

Updated_PDF = "D:\\vaibhav\\Vaibhav Clone\\mortgage-webapp\\"+element+"_May Bank_Updated1.pdf"
try:
    pdf = PyPDF2.PdfFileReader(Updated_PDF)
except:
    error = "PDF file Not Found."
    print(error)

pdfData = pdf.getFields()
outputdict={}

# Creating a dictionary from pdf data    
for data in pdfData:
    info = pdfData[data]
    if "/V" in info:
        # if info["/V"] == "/" or info["/V"] == "" :
        #     pass
        # else :
        tag=info["/T"]
        value=info["/V"]
        # tag=tag.replace("_"," ")
        if "/" in value:
            value=value.replace("/", "")
        outputdict[tag]=value

def addComma(inputValue):
    if 'S$' in str(inputValue):
        inputValue = inputValue.replace('S$',"")
    if ',' in str(inputValue):
        return str(inputValue)
    if '.' in str(inputValue):
        parts = str(inputValue).split(".")
        intPart = parts[0]
        floatPart = parts[1]
        intPart = "{:,}".format(int(float(intPart)))
        if floatPart != '0' and floatPart != '00' :
            retstr= str(intPart) + "." + str(floatPart) 
        else :
            retstr= str(intPart)
        return retstr
    elif inputValue :
        try:
            return "{:,}".format(int(inputValue))
        except:
            parts = inputValue.split(".")
            intPart = parts[0]
            floatPart = parts[1]
            intPart = "{:,}".format(int(float(intPart)))
            return intPart + "." + floatPart
    else:
        return inputValue

def remove_dot_zero(input):
    if(float(input) % 1 == 0):
        return int(input)
    else:
        return input

results = []

########################################################## Validation Starting #############################################################

if Excel_Dictionary1['Name of Financier'] != None:
    if Excel_Dictionary1['Name of Financier'] == outputdict['Name_of_Existing_Financial_Institution']:
        results.append("Name of Financier in Refinance Passed")
    else:
        results.append("Name of Financier in Refinance Failed")    
else:
    results.append("Name of Financier Data not Available in UI or Input File")


if Excel_Dictionary1['Outstanding Housing Loan Amount'] != None:
    if str(addComma(Excel_Dictionary1['Outstanding Housing Loan Amount'])) == outputdict['Housing_Loan_Amount_Outstanding']:
        results.append("Outstanding Housing Loan Amount Passed")
    else:
        results.append("Outstanding Housing Loan Amount Failed")    
else:
    results.append("Outstanding Housing Loan Amount Data not Available in UI or Input File")  


housing_loan_remaining_years = 0
if (Excel_Dictionary1['Original Loan Tenure'] != None) and (Excel_Dictionary1['Loan Disbursement Date'] != None):
    loan_disbursement_year = 0
    if len(str(Excel_Dictionary1['Loan Disbursement Date'])) == 7:
        loan_disbursement_year = str(Excel_Dictionary1['Loan Disbursement Date'])[3:]
    elif len(str(Excel_Dictionary1['Loan Disbursement Date'])) == 8:
        loan_disbursement_year = str(Excel_Dictionary1['Loan Disbursement Date'])[4:]    
    now = datetime.datetime.now()
    currentYear = now.year   
    housing_loan_remaining_years = int(Excel_Dictionary1['Original Loan Tenure']) - (int(currentYear) - int(loan_disbursement_year))
if str(housing_loan_remaining_years) == outputdict['Existing_Housing_Loan_Remaining_Yrs']:
    results.append("Outstanding Housing Loan Remaining Loan Tenure Years Passed")
else:
    results.append("Outstanding Housing Loan Remaining Loan Tenure Years Failed")      


if (Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] != None) and (Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] == 'Yes'):
    if Excel_Dictionary1['Amount Undisbursed'] != None:
        if str(addComma(Excel_Dictionary1['Amount Undisbursed'])) == outputdict['Housing_Loan_Amount_Undrawn_or_Undisbursed']:
            results.append("Amount Undisbursed in Housing Loan Refinance Passed")
        else:
            results.append("Amount Undisbursed in Housing Loan Refinance Failed")     
    else:
        results.append("Amount Undisbursed Data not Available in UI or Input File")
else:
    results.append("Is there any Undisbursed Amount in Housing Loan? Data not Available in UI or Input File or Data is Incorrect")


undisbursed_loan_remaining_years = 0
if (Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] != None) and (Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] == 'Yes'):
    if Excel_Dictionary1['Amount Undisbursed'] != None:
        if (Excel_Dictionary1['Original Loan Tenure'] != None) and (Excel_Dictionary1['Loan Disbursement Date'] != None):
            loan_disbursement_year = 0
            if len(str(Excel_Dictionary1['Loan Disbursement Date'])) == 7:
                loan_disbursement_year = str(Excel_Dictionary1['Loan Disbursement Date'])[3:]
            elif len(str(Excel_Dictionary1['Loan Disbursement Date'])) == 8:
                loan_disbursement_year = str(Excel_Dictionary1['Loan Disbursement Date'])[4:]    
            now = datetime.datetime.now()
            currentYear = now.year   
            undisbursed_loan_remaining_years = int(Excel_Dictionary1['Original Loan Tenure']) - (int(currentYear) - int(loan_disbursement_year))        
if undisbursed_loan_remaining_years > 0:
    if str(undisbursed_loan_remaining_years) == outputdict['Self_Housing_Undispersed_Amount_Remaining_Yrs']:
        results.append("Undisbursed Loan Amount Remaining Loan Tenure Years with Data Passed")
    else:
        results.append("Undisbursed Loan Amount Remaining Loan Tenure Years with Data Failed")
elif undisbursed_loan_remaining_years <= 0:
    if outputdict['Self_Housing_Undispersed_Amount_Remaining_Yrs'] == '':
        results.append("Undisbursed Loan Amount Remaining Loan Tenure Years with Blank Passed")
    else:
        results.append("Undisbursed Loan Amount Remaining Loan Tenure Years with Blank Failed") 


if (Excel_Dictionary1['application type 1'] != 'HDB'):
    if (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] != None) and (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == 'Yes'):
        if Excel_Dictionary1['Outstanding Equity Loan Amount'] != None:
            if str(addComma(Excel_Dictionary1['Outstanding Equity Loan Amount'])) == outputdict['Existing_Equity_Loan']:
                results.append("Outstanding Equity Loan Amount Refinance Passed")
            else:
                results.append("Outstanding Equity Loan Amount Refinance Failed")
        else:
            results.append("Outstanding Equity Loan Amount Data not Available in UI or Input File")
    else:
        results.append("Do you have an Outstanding Equity Loan? Data not Available in UI or Input File or Data is Incorrect")

    equity_loan_remaining_years = 0
    if (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] != None) and (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == 'Yes'):
        if (Excel_Dictionary1['Outstanding Equity Loan Disbursement Date'] != None) and (Excel_Dictionary1['Outstanding Equity Original Loan Tenure'] != None):
            loan_disbursement_year = 0
            if len(str(Excel_Dictionary1['Outstanding Equity Loan Disbursement Date'])) == 7:
                loan_disbursement_year = str(Excel_Dictionary1['Outstanding Equity Loan Disbursement Date'])[3:]
            elif len(str(Excel_Dictionary1['Outstanding Equity Loan Disbursement Date'])) == 8:
                loan_disbursement_year = str(Excel_Dictionary1['Outstanding Equity Loan Disbursement Date'])[4:]    
            now = datetime.datetime.now()
            currentYear = now.year   
            equity_loan_remaining_years = int(Excel_Dictionary1['Outstanding Equity Original Loan Tenure']) - (int(currentYear) - int(loan_disbursement_year))   
    elif (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == None) and (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == 'No'):
        equity_loan_remaining_years = ''

    if str(equity_loan_remaining_years) == outputdict['Existing_Equity_Loan_Remaining_Yrs']:
        results.append("Outstanding Equity Loan Remaining Years Passed")
    else:
        results.append("Outstanding Equity Loan Remaining Years Failed")   
                       

total_with_existing_bank = 0
if Excel_Dictionary1['Outstanding Housing Loan Amount'] != None:
    total_with_existing_bank = total_with_existing_bank + Excel_Dictionary1['Outstanding Housing Loan Amount'] 
    if (Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] != None) and (Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] == 'Yes'):
        if Excel_Dictionary1['Amount you would like to Pre-pay down'] != None:
            total_with_existing_bank = total_with_existing_bank - Excel_Dictionary1['Amount you would like to Pre-pay down']
        else:
            results.append("Amount you would like to Pre-pay down Data not Available in UI or Input File")
    else:
        results.append("Pre-Payment on the Housing Loan before Refinancing? Data not Available in UI or Input File or Data is Incorrect")
else:
    results.append("Outstanding Housing Loan Amount Data not Available in UI or Input File")                    

if (Excel_Dictionary1['application type 1'] != 'HDB'):
    if (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] != None) and (Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == 'Yes'):
        if Excel_Dictionary1['Outstanding Equity Loan Amount'] != None:
            total_with_existing_bank = total_with_existing_bank + Excel_Dictionary1['Outstanding Equity Loan Amount']
            if (Excel_Dictionary1['Pre-Payment on the Equity Loan before Refinancing?'] != None) and (Excel_Dictionary1['Pre-Payment on the Equity Loan before Refinancing?'] == 'Yes'):
                if Excel_Dictionary1['Equity Amount you would like to Pre-pay down'] != None:
                    total_with_existing_bank = total_with_existing_bank - Excel_Dictionary1['Equity Amount you would like to Pre-pay down']
                else:
                    results.append("Equity Amount you would like to Pre-pay down Data not Available in UI or Input File")
            else:
                results.append("Pre-Payment on the Equity Loan before Refinancing? Data not Available in UI or Input File or Data is Incorrect")
        else:
            results.append("Outstanding Equity Loan Amount Data not Available in UI or Input File")
    else:
        results.append("Do you have an Outstanding Equity Loan? Data not Available in UI or Input File or Data is Incorrect")                             

if str(addComma(total_with_existing_bank)) == outputdict['Total_Existing_Loan']:
    results.append("Total of Outstanding Housing and Outstanding Equity Loan Passed")
else:
    results.append("Total of Outstanding Housing and Outstanding Equity Loan Failed")    


cpf_utilised_all_owners = 0
if Excel_Dictionary1['CPF Utilised- Main'] != None:
    cpf_utilised_all_owners = cpf_utilised_all_owners + Excel_Dictionary1['CPF Utilised- Main']
if (Excel_Dictionary1['Are there additional owners?'] != None) and (Excel_Dictionary1['Are there additional owners?'] == 'Yes'):
    if Excel_Dictionary1['CPF Utilised- Joint 1'] != None:
        cpf_utilised_all_owners = cpf_utilised_all_owners + Excel_Dictionary1['CPF Utilised- Joint 1']
    if Excel_Dictionary1['CPF Utilised- Joint 2'] != None:
        cpf_utilised_all_owners = cpf_utilised_all_owners + Excel_Dictionary1['CPF Utilised- Joint 2']
    if Excel_Dictionary1['CPF Utilised- Joint 3'] != None:
        cpf_utilised_all_owners = cpf_utilised_all_owners + Excel_Dictionary1['CPF Utilised- Joint 3']      
if str(addComma(cpf_utilised_all_owners)) == outputdict['Total_Overall_CPF_Withdrawn_to_date']:
    results.append("CPF Utilised for All Owner's Refinance Passed")
else:
    results.append("CPF Utilised for All Owner's Refinance Failed")          


mortgage_loan_refinance = 0
if Excel_Dictionary1['Outstanding Housing Loan Amount'] != None:
    mortgage_loan_refinance = mortgage_loan_refinance + Excel_Dictionary1['Outstanding Housing Loan Amount'] 
    if (Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] != None) and (Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] == 'Yes'):
        if Excel_Dictionary1['Amount you would like to Pre-pay down'] != None:
            mortgage_loan_refinance = mortgage_loan_refinance - Excel_Dictionary1['Amount you would like to Pre-pay down']
        else:
            results.append("Amount you would like to Pre-pay down Data not Available in UI or Input File")
    else:
        results.append("Pre-Payment on the Housing Loan before Refinancing? Data not Available in UI or Input File or Data is Incorrect")
else:
    results.append("Outstanding Housing Loan Amount Data not Available in UI or Input File")                    


if mortgage_loan_refinance > 0:
    if str(addComma(mortgage_loan_refinance)) == outputdict['Self_Mortgage_Land_Loan_Refinance']:
        results.append("Mortgage/Land Loan (Refinance) with Data Passed")
    else:
        results.append("Mortgage/Land Loan (Refinance) with Data Failed")    
elif mortgage_loan_refinance <= 0:
    if outputdict['Self_Mortgage_Land_Loan_Refinance'] == '':
        results.append("Mortgage/Land Loan (Refinance) without Data Passed")     
    else:
        results.append("Mortgage/Land Loan (Refinance) without Data Failed")       


if Excel_Dictionary1['New Loan Tenure'] != None:
    if str(Excel_Dictionary1['New Loan Tenure']) == outputdict['Self_Mortgage_Land_Loan_Refinance_Yrs']:
        results.append("Housing Loan New Loan Tenure Passed")
    else:
        results.append("Housing Loan New Loan Tenure Failed")
else:
    results.append("Housing Loan New Loan Tenure Data not Available in UI or Input File")            


if (Excel_Dictionary1['application type 1'] != 'HDB'):
    if (Excel_Dictionary1['New Equity Loan Required?'] != None) and (Excel_Dictionary1['New Equity Loan Required?'] == 'Yes'):
        if Excel_Dictionary1['New Equity Loan for Personal Use'] != None:
            if str(addComma(Excel_Dictionary1['New Equity Loan for Personal Use'])) == outputdict['Self_Additional_Equity_Loan']:
                results.append("New Equity Loan for Personal Use Passed")
            else:
                results.append("New Equity Loan for Personal Use Failed")
        if Excel_Dictionary1['New Equity Loan Tenure'] != None:
            if str(Excel_Dictionary1['New Equity Loan Tenure']) == outputdict['Self_Additional_Equity_Loan_Yrs']:
                results.append("New Equity Loan Tenure Passed")
            else:
                results.append("New Equity Loan Tenure Failed")               
    else:
        results.append("New Equity Loan Required? Data not Available in UI or Input File or Data is Incorrect")                


if Excel_Dictionary1['Outstanding Housing Loan Amount'] != None:
    if outputdict['Self_Additional_Equity_Loan_Purpose'] == 'Personal Use':
        results.append("Loan Purpose 'Personal Use' Passed")
    else:
        results.append("Loan Purpose 'Personal Use' Failed")    


if Excel_Dictionary1['CPF for Monthly Repayment?'] != None:
    if Excel_Dictionary1['CPF for Monthly Repayment?'] == 'Yes':
        if outputdict['Self_CPF_for_Monthly_Repayment_Check_Yes'] == 'Yes':
            results.append("CPF for Monthly Repayment 'Yes' tick Passed")
        else:
            results.append("CPF For Monthly Repayment 'Yes' tick Failed")    
        if Excel_Dictionary1['CPF For Monthly Repayment Amount'] != None:
            cpf_Monthly_Installment = str(addComma(Excel_Dictionary1['CPF For Monthly Repayment Amount']))
            if cpf_Monthly_Installment == outputdict['Main_Account_Holder_CPF_per_month']:
                results.append("CPF For Monthly Repayment Amount Passed") 
            else:
                results.append("CPF For Monthly Repayment Amount Failed")
        else:
            results.append("CPF For Monthly Repayment Amount Data not Available in UI or Input File")               
    elif Excel_Dictionary1['CPF for Monthly Repayment?'] == 'No':
        if outputdict['Self_CPF_for_Monthly_Repayment_Check_No'] == 'Yes':
            results.append("CPF for Monthly Repayment 'No' tick Passed")
        else:
            results.append("CPF for Monthly Repayment 'No' tick Failed")    
    else:
        results.append("CPF for Monthly Repayment? Data is Incorrect")        
else:
    results.append("CPF for Monthly Repayment? Data not Available in UI or Input File")        




########################################################## Validation Ending #############################################################

#Preparing Results File 
writesheet = openpyxl.Workbook()
Wsheet = writesheet.active
Wsheet.column_dimensions['A'].width = float(100)
if len(results) > 0:
    j = 1
    for i in range(len(results)):
        Wsheet.cell(j,1).value = results[i]
        j+=1

writesheet.save("D:\\Automation Testing\\MayBank\\Loan\\Results\\loan_refinance.xlsx")   