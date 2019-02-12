from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)

path = "D:\\Automation Testing\\MayBank\\Loan\\loan_new_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
for i in range(1,60):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value
    
driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_id("AddLoanApplication").click()
driver.find_element_by_id("homeBuyerSearchInputId").send_keys("A K S")
driver.find_element_by_id("homeBuyerSearchButtonId").click()
driver.find_element_by_id("selectedClientId").click()
driver.find_element_by_xpath('//*[@id="homeBuyerDetailsModelDiv"]/div/div/div[2]/div[3]/div[2]/button').click()
driver.find_element_by_id("singleApplicantID").click()

if Excel_Dictionary1['application type 1'] != None:
    if Excel_Dictionary1['application type 1'] == 'Private':
        driver.find_element_by_id("privateFlatID").click()
    elif Excel_Dictionary1['application type 1'] == 'Executive Condo':
        driver.find_element_by_id("executiveCondoID").click()
    elif Excel_Dictionary1['application type 1'] == 'HDB':
        driver.find_element_by_id('hdbFlatID').click()        
    else:
        driver.find_element_by_id("privateFlatID").click()
else:
    driver.find_element_by_id("privateFlatID").click() 

driver.find_element_by_id("newHomeLoanImage").click()
driver.find_element_by_xpath("(.//*[normalize-space(text()) and normalize-space(.)='Cancel'])[2]/following::button[2]").click()
time.sleep(4)

if Excel_Dictionary1['Salutation'] != None:
    driver.find_element_by_xpath('//*[@id="salutation"]').send_keys(Excel_Dictionary1['Salutation'])
if Excel_Dictionary1['Full Name as in NRIC/PASSPORT'] != None:     
    driver.find_element_by_xpath('//*[@id="fullName"]').clear()
    driver.find_element_by_xpath('//*[@id="fullName"]').send_keys(Excel_Dictionary1['Full Name as in NRIC/PASSPORT'])     
if Excel_Dictionary1['Email Address'] != None:   
    driver.find_element_by_name('M0031').send_keys(Excel_Dictionary1['Email Address']) 

driver.find_element_by_xpath('//*[@id="loan"]').click()
time.sleep(4)   

if Excel_Dictionary1['Valuation Firm'] != None:
    driver.find_element_by_name('M0275').send_keys(Excel_Dictionary1['Valuation Firm'])
if Excel_Dictionary1['Name of Valuer'] != None:
    driver.find_element_by_name('M0531').send_keys(Excel_Dictionary1['Name of Valuer'])
if Excel_Dictionary1['Indicative Valuation'] != None:
    driver.find_element_by_name('M0273').send_keys(Excel_Dictionary1['Indicative Valuation'])
if Excel_Dictionary1['Date of Valuation'] != None:
    driver.find_element_by_name('M0274').send_keys(Excel_Dictionary1['Date of Valuation']) 

if Excel_Dictionary1['CPF for Monthly Repayment?'] != None:
    if Excel_Dictionary1['CPF for Monthly Repayment?'] == 'Yes':
        driver.find_element_by_id('isCPFForMonthlyRepaymentYes').click()
        time.sleep(1)
        if Excel_Dictionary1['CPF For Monthly Repayment Amount'] != None:
            driver.find_element_by_id('CPFForMonthlyRepaymentValue').send_keys(Excel_Dictionary1['CPF For Monthly Repayment Amount'])
            time.sleep(1)
    elif Excel_Dictionary1['CPF for Monthly Repayment?'] == 'No':
        driver.find_element_by_id('isCPFForMonthlyRepaymentNo').click()
        time.sleep(1)

if Excel_Dictionary1['CPF for Legal Fees?'] != None:
    if Excel_Dictionary1['CPF for Legal Fees?'] == 'Yes':
        driver.find_element_by_id('isCPFForLegalFeesYes').click()
        time.sleep(1)
        if Excel_Dictionary1['CPF for Legal Fees Amount'] != None:
            driver.find_element_by_id('CPFForLegalFeesValue').send_keys(Excel_Dictionary1['CPF for Legal Fees Amount'])  
            time.sleep(1)
    elif Excel_Dictionary1['CPF for Legal Fees?'] == 'No':
        driver.find_element_by_id('isCPFForLegalFeesNo').click()
        time.sleep(1)

if Excel_Dictionary1['CPF for Stamp Duty?'] != None:
    if Excel_Dictionary1['CPF for Stamp Duty?'] == 'Yes':
        driver.find_element_by_id('isCPFForStampDutyYes').click()
        time.sleep(1)
        if Excel_Dictionary1['CPF for Stamp Duty Amount'] != None:
            driver.find_element_by_id('CPFForStampDutyValue').send_keys(Excel_Dictionary1['CPF for Stamp Duty Amount'])  
            time.sleep(1)
    elif Excel_Dictionary1['CPF for Stamp Duty?'] == 'No':
        driver.find_element_by_id('isCPFForStampDutyNo').click()
        time.sleep(1)

if Excel_Dictionary1['Transaction Type'] != None:
    driver.find_element_by_name('M0280').send_keys(Excel_Dictionary1['Transaction Type'])

if (Excel_Dictionary1['application type 1'] != None) and (Excel_Dictionary1['application type 1'] == 'Private' or Excel_Dictionary1['application type 1'] == 'Executive Condo'):
    if Excel_Dictionary1['Discount/Rebate Yes/No'] != None:
        if Excel_Dictionary1['Discount/Rebate Yes/No'] == 'Yes':
            driver.find_element_by_id('isDiscountRebatesReceivedYes').click()
            time.sleep(1)
            if Excel_Dictionary1['Discount / Rebate / Benefits Amount'] != None:
                driver.find_element_by_name('M0543').send_keys(Excel_Dictionary1['Discount / Rebate / Benefits Amount'])
                time.sleep(1)
            if Excel_Dictionary1['Vendor / Developer / Third-Party Name'] != None:
                driver.find_element_by_name('M0542').send_keys(Excel_Dictionary1['Vendor / Developer / Third-Party Name']) 
                time.sleep(1)
            if Excel_Dictionary1['Furniture & Electrical Rebates'] != None:
                driver.find_element_by_id(Excel_Dictionary1['Furniture & Electrical Rebates']).send_keys(Keys.SPACE) 
                time.sleep(1)
            if Excel_Dictionary1['Household appliances/Stamp Duty Rebates'] != None:
                driver.find_element_by_id(Excel_Dictionary1['Household appliances/Stamp Duty Rebates']).send_keys(Keys.SPACE)
                time.sleep(1) 
            if Excel_Dictionary1['Cash & Vouchers Rebates'] != None:
                driver.find_element_by_id(Excel_Dictionary1['Cash & Vouchers Rebates']).send_keys(Keys.SPACE)
                time.sleep(1)
            if Excel_Dictionary1['Lucky Draw'] != None:
                driver.find_element_by_id(Excel_Dictionary1['Lucky Draw']).send_keys(Keys.SPACE)
                time.sleep(1)
            if Excel_Dictionary1['Benefits from Deferred Payment Scheme'] != None:
                driver.find_element_by_id(Excel_Dictionary1['Benefits from Deferred Payment Scheme']).send_keys(Keys.SPACE) 
                time.sleep(1)
            if Excel_Dictionary1['Others(Please Specify)'] != None:
                driver.find_element_by_id(Excel_Dictionary1['Others(Please Specify)']).send_keys(Keys.SPACE)
                time.sleep(1)
            if Excel_Dictionary1['Others(Please Specify) Data'] != None:
                driver.find_element_by_id('othersText').send_keys(Excel_Dictionary1['Others(Please Specify) Data'])   
                time.sleep(1)                         
        elif Excel_Dictionary1['Discount/Rebate Yes/No'] == 'No':
            driver.find_element_by_id('isDiscountRebatesReceivedNo').click()  
            time.sleep(1)  

if Excel_Dictionary1['Purchase Price'] != None:
    driver.find_element_by_name('M0281').send_keys(Excel_Dictionary1['Purchase Price'])

time.sleep(1)
if (Excel_Dictionary1['Transaction Type'] != None) and (Excel_Dictionary1['Transaction Type'] == 'Resale Market'):
    if (Excel_Dictionary1['application type 1'] != None) and (Excel_Dictionary1['application type 1'] == 'HDB' or Excel_Dictionary1['application type 1'] == 'Executive Condo'):
        if Excel_Dictionary1['HDB/CPF Grant'] != None:
            driver.find_element_by_name('M0286').send_keys(Excel_Dictionary1['HDB/CPF Grant']) 

if Excel_Dictionary1['Cash down payment (minimum 5%)'] != None:
    driver.find_element_by_name('M0284').send_keys(Excel_Dictionary1['Cash down payment (minimum 5%)'])  

if Excel_Dictionary1['CPF Down payment'] != None:
    driver.find_element_by_name('M0285').send_keys(Excel_Dictionary1['CPF Down payment'])              

if Excel_Dictionary1['Loan Tenure'] != None:
    driver.find_element_by_name('M0288').send_keys(Excel_Dictionary1['Loan Tenure'])    

time.sleep(1)
if Excel_Dictionary1['Bridging Loan Required?'] != None:
    if Excel_Dictionary1['Bridging Loan Required?'] == 'Yes':
        driver.find_element_by_id('isBridgingLoanRequiredYes').click()
        time.sleep(1)
        if Excel_Dictionary1['Type of Property'] != None:
            driver.find_element_by_name('M0296').send_keys(Excel_Dictionary1['Type of Property'])
        if Excel_Dictionary1['Expected/Actual Sales Date'] != None:
            driver.find_element_by_name('M0301').send_keys(Excel_Dictionary1['Expected/Actual Sales Date'])
        if Excel_Dictionary1['Country'] != None:
            driver.find_element_by_name('M0514').send_keys(Excel_Dictionary1['Country'])  
            time.sleep(1)      
        if Excel_Dictionary1['Zip / Postal Code'] != None:
            driver.find_element_by_name('M0472').send_keys(Excel_Dictionary1['Zip / Postal Code'])
        if Excel_Dictionary1['Address of Property to be sold'] != None:
            driver.find_element_by_name('M0302').send_keys(Excel_Dictionary1['Address of Property to be sold'])
        if Excel_Dictionary1['Unit No'] != None:
            driver.find_element_by_name('M0510').send_keys(Excel_Dictionary1['Unit No'])  
            time.sleep(1)  
        if Excel_Dictionary1['Selling Price of Property'] != None:
            driver.find_element_by_name('M0290').send_keys(Excel_Dictionary1['Selling Price of Property']) 
        if Excel_Dictionary1['Outstanding Loan'] != None:
            driver.find_element_by_name('M0291').send_keys(Excel_Dictionary1['Outstanding Loan'])      
        if Excel_Dictionary1['HDB Levy (if any)'] != None:
            driver.find_element_by_name('M0294').send_keys(Excel_Dictionary1['HDB Levy (if any)']) 
            time.sleep(1)
        if Excel_Dictionary1['Main Owner Name'] != None:    
            driver.find_element_by_name('M0545').send_keys(Excel_Dictionary1['Main Owner Name'])
        if Excel_Dictionary1['CPF Utilised- Main'] != None:
            driver.find_element_by_name('M0293').send_keys(Excel_Dictionary1['CPF Utilised- Main'])   
            time.sleep(1) 
        if Excel_Dictionary1['Are there additional owners?'] != None:
            if Excel_Dictionary1['Are there additional owners?'] == 'Yes':
                driver.find_element_by_id('areThereAdditionalOwnersYes').click()
                time.sleep(1) 
                if Excel_Dictionary1['Joint Owner 1 Name'] != None:
                    driver.find_element_by_xpath('//*[@id="addotherownerID"]/span').click()
                    time.sleep(1)
                    driver.find_element_by_id('nameOfOwner1').send_keys(Excel_Dictionary1['Joint Owner 1 Name'])
                    if Excel_Dictionary1['CPF Utilised- Joint 1'] != None:
                        driver.find_element_by_id('cpfUtilisedJoint1').send_keys(Excel_Dictionary1['CPF Utilised- Joint 1'])
                if Excel_Dictionary1['Joint Owner 2 Name'] != None:
                    driver.find_element_by_xpath('//*[@id="addotherownerID"]/span').click()
                    time.sleep(1)
                    driver.find_element_by_id('nameOfOwner2').send_keys(Excel_Dictionary1['Joint Owner 2 Name'])
                    if Excel_Dictionary1['CPF Utilised- Joint 2'] != None:
                        driver.find_element_by_id('cpfUtilisedJoint2').send_keys(Excel_Dictionary1['CPF Utilised- Joint 2'])
                if Excel_Dictionary1['Joint Owner 3 Name'] != None:
                    driver.find_element_by_xpath('//*[@id="addotherownerID"]/span').click()  
                    time.sleep(1)
                    driver.find_element_by_id('nameOfOwner3').send_keys(Excel_Dictionary1['Joint Owner 3 Name'])    
                    if Excel_Dictionary1['CPF Utilised- Joint 3'] != None:
                        driver.find_element_by_id('cpfUtilisedJoint3').send_keys(Excel_Dictionary1['CPF Utilised- Joint 3'])                  
            elif Excel_Dictionary1['Are there additional owners?'] == 'No':
                driver.find_element_by_id('areThereAdditionalOwnersNo').click()  
                time.sleep(1) 
        if Excel_Dictionary1['Bridging Loan CPF'] != None:
            driver.find_element_by_name('M0297').send_keys(Excel_Dictionary1['Bridging Loan CPF'])        
        if Excel_Dictionary1['Bridging Loan Cash'] != None:
            driver.find_element_by_name('M0298').send_keys(Excel_Dictionary1['Bridging Loan Cash']) 
        if Excel_Dictionary1['Bridging Loan Tenure'] != None:
            driver.find_element_by_name('M0300').send_keys(Excel_Dictionary1['Bridging Loan Tenure'])
            time.sleep(1)                   
    elif Excel_Dictionary1['Bridging Loan Required?'] == 'No':
        driver.find_element_by_id('isBridgingLoanRequiredNo').click()
        time.sleep(1)        

driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
driver.find_element_by_class_name('loan-submit-button').click()
time.sleep(3)
driver.find_element_by_xpath('//*[@id="checkbox-select"]').click()
driver.find_element_by_xpath('//*[@id="maybank-deselected"]').click()
driver.find_element_by_xpath('//*[@id="downloadForms"]').click()
time.sleep(70)
driver.quit()
print("Form Filling Done")            