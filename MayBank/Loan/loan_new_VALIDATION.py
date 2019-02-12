from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time


path = "D:\\Automation Testing\\MayBank\\Loan\\loan_new_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
for i in range(1,50):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)
driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_xpath('//*[@id="loansTable"]/tbody/tr[1]/td[8]/div/button').click()
time.sleep(3)
element = driver.find_element_by_xpath('//*[@id="applicationIdField"]').text
driver.quit()

Updated_PDF = "D:\\vaibhav\\Vaibhav Clone\\mortgage-webapp\\"+element+"_May Bank_Updated1.pdf"
try:
    pdf = PyPDF2.PdfFileReader(Updated_PDF)
except:
    error = "PDF file Not Found."
    print(error)

pdfData = pdf.getFields()
outputdict={}

# Creating a dictionary from pdf data    
for data in pdfData:
    info = pdfData[data]
    if "/V" in info:
        # if info["/V"] == "/" or info["/V"] == "" :
        #     pass
        # else :
        tag=info["/T"]
        value=info["/V"]
        # tag=tag.replace("_"," ")
        if "/" in value:
            value=value.replace("/", "")
        outputdict[tag]=value

def addComma(inputValue):
    if 'S$' in str(inputValue):
        inputValue = inputValue.replace('S$',"")
    if ',' in str(inputValue):
        return str(inputValue)
    if '.' in str(inputValue):
        parts = str(inputValue).split(".")
        intPart = parts[0]
        floatPart = parts[1]
        intPart = "{:,}".format(int(float(intPart)))
        if floatPart != '0' and floatPart != '00' :
            retstr= str(intPart) + "." + str(floatPart) 
        else :
            retstr= str(intPart)
        return retstr
    elif inputValue :
        try:
            return "{:,}".format(int(inputValue))
        except:
            parts = inputValue.split(".")
            intPart = parts[0]
            floatPart = parts[1]
            intPart = "{:,}".format(int(float(intPart)))
            return intPart + "." + floatPart
    else:
        return inputValue

def remove_dot_zero(input):
    if(float(input) % 1 == 0):
        return int(input)
    else:
        return input

results = []

########################################################## Validation Starting #############################################################

if Excel_Dictionary1['application type 1'] != 'HDB':
    if Excel_Dictionary1['Discount/Rebate Yes/No'] != None:
        if Excel_Dictionary1['Discount/Rebate Yes/No'] == 'Yes':
            if Excel_Dictionary1['Discount / Rebate / Benefits Amount'] != None:
                if outputdict['Self_Benefits_from_Develop_Vendor_Yes'] == 'Yes':
                    results.append("Discount/Rebate/Benefits tick 'Yes' Passed")
                discount_rebate_amount = str(addComma(Excel_Dictionary1['Discount / Rebate / Benefits Amount'])) 
                if discount_rebate_amount == outputdict['Self_Benefits_Amount']:
                    results.append("Discount/Rebate/Benefits Amount Passed") 
        else:
            results.append("Discount/Rebates/Benefits Failed")              
    else:
        results.append("Discount/Rebate Yes/No Data not Available in UI or Input File")   


if Excel_Dictionary1['Purchase Price'] != None:
    purchase_price = str(addComma(Excel_Dictionary1['Purchase Price']))
    if purchase_price == outputdict['Purchase_Price']:
        results.append("Purchase Price Data Passed")
    else:
        results.append("Purchase Price Data Failed")        
else:
    results.append("Purchase Price Data not Available in UI or Input File")


if Excel_Dictionary1['Transaction Type'] != None:
    if (Excel_Dictionary1['Transaction Type'] == 'Direct from Developer') and (outputdict['Self_Transaction_Type_DevOrHDB'] == 'Yes'):
        results.append("Transaction Type 'Direct from Developer' 'tick' Passed")
    elif (Excel_Dictionary1['Transaction Type'] == 'Direct from HDB') and (outputdict['Self_Transaction_Type_DevOrHDB'] == 'Yes'):
        results.append("Transaction Type 'Direct from HDB' 'tick' Passed")    
    elif (Excel_Dictionary1['Transaction Type'] == 'Resale Market') and (outputdict['Self_Transaction_Type_SecondaryMarket'] == 'Yes'):
        results.append("Transaction Type 'Resale Market' 'tick' Passed") 
    else:
        results.append("Transaction Type Failed")       
else:
    results.append("Transaction Type Data not Available in UI or Input File")    

 
if Excel_Dictionary1['Purchase Price'] != None:
    if Excel_Dictionary1['Indicative Valuation'] != None:   
        cov = Excel_Dictionary1['Purchase Price'] - Excel_Dictionary1['Indicative Valuation']
        if cov > 0:
            if str(addComma(cov)) == outputdict['Self_Cash_Over_Valuation']:
                results.append("Cash Over Valuation Data Passed")
            else:
                results.append("Cash Over Valuation Data Failed")
        elif cov <= 0:
            if outputdict['Self_Cash_Over_Valuation'] == '':
                results.append("Cash Over Valuation Blank Passed")
            else:
                results.append("Cash Over Valuation Blank Failed") 
    else:
        results.append("Indicative Valuation Data not Available in UI or Input File")
else:
    results.append("Purchase Price Data not Available in UI or Input File")    


if Excel_Dictionary1['Cash down payment (minimum 5%)'] != None:
    cash_down_payment = str(addComma(Excel_Dictionary1['Cash down payment (minimum 5%)']))
    if cash_down_payment == outputdict['Cash_Deposit']:
        results.append("Cash down payment Passed")
    else:
        results.append("Cash down payment Failed")    
else:
    results.append("Cash down payment (minimum 5%) Data not Available in UI or Input File")    


if Excel_Dictionary1['CPF Down payment'] != None:
    cpf_down_payment = str(addComma(Excel_Dictionary1['CPF Down payment']))
    if cpf_down_payment == outputdict['CPF_Lump_Sum_Deposit']:
        results.append("CPF Down payment Passed")
    else:
        results.append("CPF Down payment Failed")    
else:
    results.append("CPF Down payment Data not Available in UI or Input File")    


if (Excel_Dictionary1['application type 1'] != None): 
    if (Excel_Dictionary1['application type 1'] == 'HDB' or Excel_Dictionary1['application type 1'] == 'Executive Condo'):
        if Excel_Dictionary1['HDB/CPF Grant'] != None:
            if str(addComma(Excel_Dictionary1['HDB/CPF Grant'])) == outputdict['Housing_Grant_for_HDB']:
                results.append("HDB/CPF Grant Passed")
            else:
                results.append("HDB/CPF Grant Failed")  
        else:
            results.append("HDB/CPF Grant Data not Available in UI or Input File")
    else:
        results.append("Application type Data is Incorrect") 
else:
    results.append("Application type Data is not Available in UI or Input File")    


housing_loan_required = 0
if Excel_Dictionary1['Purchase Price'] != None:
    housing_loan_required = Excel_Dictionary1['Purchase Price'] 
    if Excel_Dictionary1['Indicative Valuation'] != None:
        housing_loan_required = Excel_Dictionary1['Indicative Valuation']
    if Excel_Dictionary1['application type 1'] != 'Private':
        if Excel_Dictionary1['HDB/CPF Grant'] != None:
            housing_loan_required = housing_loan_required - Excel_Dictionary1['HDB/CPF Grant']
    if Excel_Dictionary1['Cash down payment (minimum 5%)'] != None:
        housing_loan_required = housing_loan_required - Excel_Dictionary1['Cash down payment (minimum 5%)']
    if Excel_Dictionary1['CPF Down payment'] != None:
        housing_loan_required = housing_loan_required - Excel_Dictionary1['CPF Down payment']
    if (Excel_Dictionary1['Bridging Loan Required?'] != None) and (Excel_Dictionary1['Bridging Loan Required?'] == 'Yes'):
        if Excel_Dictionary1['Bridging Loan CPF'] != None:
            housing_loan_required = housing_loan_required - Excel_Dictionary1['Bridging Loan CPF']
        if Excel_Dictionary1['Bridging Loan Cash'] != None:
            housing_loan_required = housing_loan_required - Excel_Dictionary1['Bridging Loan Cash']     

if housing_loan_required > 0:
    if str(addComma(housing_loan_required)) == outputdict['Self_Mortgage_Land_Loan_New']:
        results.append("Mortgage / Land Loan (New Purchase) Passed")
    else:
        results.append("Mortgage / Land Loan (New Purchase) Failed")
elif housing_loan_required <= 0:
    if outputdict['Self_Mortgage_Land_Loan_New'] == '':
        results.append("Mortgage / Land Loan (New Purchase) Passed")
    else:
        results.append("Mortgage / Land Loan (New Purchase) Failed")                  

if str(addComma(housing_loan_required)) == outputdict['Housing_Loan_Required']:
    results.append("Mortgage / Housing Loan Required Passed")
else:
    results.append("Mortgage / Housing Loan Required Failed")

bridging_loan_required = 0
if (Excel_Dictionary1['Bridging Loan Required?'] != None) and (Excel_Dictionary1['Bridging Loan Required?'] == 'Yes'):
    if Excel_Dictionary1['Bridging Loan CPF'] != None:
        bridging_loan_required = bridging_loan_required + Excel_Dictionary1['Bridging Loan CPF']
    if Excel_Dictionary1['Bridging Loan Cash'] != None:
        bridging_loan_required = bridging_loan_required + Excel_Dictionary1['Bridging Loan Cash']
else:
    results.append("Bridging Loan Required Data not Available in UI or Input File")

if str(addComma(bridging_loan_required)) == outputdict['Bridging_Loan_Required']:
    results.append("Bridging Loan Required Passed")
else:
    results.append("Bridging Loan Required Failed")    

total = housing_loan_required + bridging_loan_required
if str(addComma(total)) == outputdict['Self_Total_Financing_Required']:
    results.append("Total Loan Required Passed")
else:
    results.append("Total Loan Required Failed")        


bridging_loan_address = (Excel_Dictionary1['Address of Property to be sold'] +", "+ Excel_Dictionary1['Unit No'] +", "+ Excel_Dictionary1['Country']).lower()
output_bridging_loan_address = (outputdict['Address_of_Property_Sold_ToBeSold'] + outputdict['Address_of_Property_Sold_ToBeSold_Dummy']).replace("_","")

if bridging_loan_address == output_bridging_loan_address:
    results.append("Bridging Loan Address Passed")
else:
    results.append("Bridging Loan Address Failed")

if Excel_Dictionary1['Zip / Postal Code'] != None:
    if str(Excel_Dictionary1['Zip / Postal Code']) == outputdict['Property_Address_To_Be_Sold_Postal_Code']:
        results.append("Bridging Loan Address Zip/Postal Code Passed")
    else:
        results.append("Bridging Loan Address Zip/Postal Code Failed")
else:
    results.append("Zip / Postal Code Data not Available in UI or Input File")        

if (Excel_Dictionary1['Bridging Loan Required?'] != None) and (Excel_Dictionary1['Bridging Loan Required?'] == 'Yes'):
    if Excel_Dictionary1['Zip / Postal Code'] != None:
        if outputdict['Property_To_be_Sold_Check'] == 'Yes':
            results.append("Bridging Loan Property to be Sold tick Passed")
        else:
            results.append("Bridging Loan Property to be Sold tick Failed")    
    else:
        results.append("Zip / Postal Code Data not Available in UI or Input File")  
else:
    results.append("Bridging Loan Required? Data not Available in UI or Input File or Data is Incorrect")     
                     

if Excel_Dictionary1['Selling Price of Property'] != None:
    selling_price_property = str(addComma(Excel_Dictionary1['Selling Price of Property']))
    if selling_price_property == outputdict['Sale_Price_of_Existing_Property']:
        results.append("Bridging loan Selling Price Passed")
    else:
        results.append("Bridging loan Selling Price Failed")
else:
    results.append("Bridging loan Selling Price Data not Available in UI or Input File")


if Excel_Dictionary1['Outstanding Loan'] != None:
    outstanding_loan = str(addComma(Excel_Dictionary1['Outstanding Loan']))
    if outstanding_loan == outputdict['Oustanding_Loan']:
        results.append("Bridging loan Outstanding loan Passed")
    else:
        results.append("Bridging loan Outstanding loan Failed")
else:
    results.append("Bridging loan Outstanding loan Data not Available in UI or Input File") 


if Excel_Dictionary1['CPF Utilised- Main'] != None:
    if str(addComma(Excel_Dictionary1['CPF Utilised- Main'])) == outputdict['CPF_Utilised_Main']:
        results.append("CPF Utilised- Main Passed")
    else:
        results.append("CPF Utilised- Main Failed")    
else:
    results.append("CPF Utilised- Main Data not Available in UI or Input File")

cpf_utilised_joint = 0
if (Excel_Dictionary1['Are there additional owners?'] != None) and (Excel_Dictionary1['Are there additional owners?'] == 'Yes'):
    if Excel_Dictionary1['CPF Utilised- Joint 1'] != None:
        cpf_utilised_joint = cpf_utilised_joint + Excel_Dictionary1['CPF Utilised- Joint 1']
    if Excel_Dictionary1['CPF Utilised- Joint 2'] != None:
        cpf_utilised_joint = cpf_utilised_joint + Excel_Dictionary1['CPF Utilised- Joint 2']
    if Excel_Dictionary1['CPF Utilised- Joint 3'] != None:
        cpf_utilised_joint = cpf_utilised_joint + Excel_Dictionary1['CPF Utilised- Joint 3']    
else:        
    results.append("Are there additional Owners Data not Available in UI or Input File")

if str(addComma(cpf_utilised_joint)) == outputdict['CPF_Utilised_Joint']:
    results.append("CPF Utilised Joint Passed")
else:
    results.append("CPF Utilised Joint Failed")    


net_cash_proceeds = 0
if Excel_Dictionary1['Selling Price of Property'] != None:
    net_cash_proceeds = Excel_Dictionary1['Selling Price of Property']
    if Excel_Dictionary1['Outstanding Loan'] != None:
        net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['Outstanding Loan']
    if Excel_Dictionary1['CPF Utilised- Main'] != None:
        net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['CPF Utilised- Main']
    if (Excel_Dictionary1['Are there additional owners?'] != None) and (Excel_Dictionary1['Are there additional owners?'] == 'Yes'):
        if Excel_Dictionary1['CPF Utilised- Joint 1'] != None:
            net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['CPF Utilised- Joint 1']
        if Excel_Dictionary1['CPF Utilised- Joint 2'] != None:
            net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['CPF Utilised- Joint 2']
        if Excel_Dictionary1['CPF Utilised- Joint 3'] != None:
            net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['CPF Utilised- Joint 3']       
    if Excel_Dictionary1['HDB Levy (if any)'] != None:
        net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['HDB Levy (if any)']
        net_cash_proceeds_hdb_levy = addComma(net_cash_proceeds) + "Incl. HDB Levy: S$ "+addComma(Excel_Dictionary1['HDB Levy (if any)'])         
        if str(net_cash_proceeds_hdb_levy) == outputdict['Net_Cash_Proceeds']:
            results.append("Net Cash Proceeds With HDB Levy Data Passed")
        else: 
            results.append("Net Cash Proceeds With HDB Levy Data Failed")
    else:
        if str(addComma(net_cash_proceeds)) == outputdict['Net_Cash_Proceeds']:
            results.append("Net Cash Proceeds With Data Passed")
        else:
            results.append("Net Cash Proceeds With Data Failed")    
elif Excel_Dictionary1['Selling Price of Property'] == None:
    if Excel_Dictionary1['HDB Levy (if any)'] != None:
        # net_cash_proceeds = net_cash_proceeds - Excel_Dictionary1['HDB Levy (if any)']
        net_cash_proceeds_hdb_levy = net_cash_proceeds + "Incl. HDB Levy: S$ "+addComma(Excel_Dictionary1['HDB Levy (if any)'])         
        if str(net_cash_proceeds_hdb_levy) == outputdict['Net_Cash_Proceeds']:
            results.append("Net Cash Proceeds With HDB Levy Data Passed")
        else: 
            results.append("Net Cash Proceeds With HDB Levy Data Failed")
    else:
        if outputdict['Net_Cash_Proceeds'] == '0':
            results.append("Net Cash Proceeds with '0' Passed")
        else:
            results.append("Net Cash Proceeds with '0' Failed")    


if Excel_Dictionary1['Bridging Loan Cash'] != None:
    if str(addComma(Excel_Dictionary1['Bridging Loan Cash'])) == outputdict['Bridging_Loan_Cash']:
        results.append("Bridging Loan Cash Passed")
    else:
        results.append("Bridging Loan Cash Failed")
else:
    results.append("Bridging Loan Cash Data not Available in UI or Input File")             

if Excel_Dictionary1['Bridging Loan CPF'] != None:
    if str(addComma(Excel_Dictionary1['Bridging Loan CPF'])) == outputdict['Bridging_Loan_CPF']:
        results.append("Bridging Loan CPF Passed")
    else:
        results.append("Bridging Loan CPF Failed")        
else:
    results.append("Bridging Loan CPF Data not Available in UI or Input File") 

if Excel_Dictionary1['Bridging Loan Tenure'] != None:
    if str(Excel_Dictionary1['Bridging Loan Tenure']) == outputdict['Self_Bridging_Loan_Cash_Yrs']:
        results.append("Bridging Loan Tenure (Cash) Passed")
    else:
        results.append("Bridging Loan Tenure (Cash) Failed") 
    if str(Excel_Dictionary1['Bridging Loan Tenure']) == outputdict['Self_Bridging_Loan_CPF_Yrs']:
        results.append("Bridging Loan Tenure (CPF) Passed")
    else:
        results.append("Bridging Loan Tenure (CPF) Failed")
else:
    results.append("Bridging Loan Tenure Data not Available in UI or Input File")         
                              

if Excel_Dictionary1['Loan Tenure'] != None:
    if (str(Excel_Dictionary1['Loan Tenure']) == outputdict['Housing_Loan_Tenure_Yrs']):
        results.append("New Loan Details Loan Tenure Passed")
    else:
        results.append("New Loan Details Loan Tenure Failed")    
else:
    results.append("New Loan Details Loan Data not Available in UI or Input File")

             
if Excel_Dictionary1['CPF for Monthly Repayment?'] != None:
    if Excel_Dictionary1['CPF for Monthly Repayment?'] == 'Yes':
        if outputdict['Self_CPF_for_Monthly_Repayment_Check_Yes'] == 'Yes':
            results.append("CPF for Monthly Repayment 'Yes' tick Passed")
        else:
            results.append("CPF For Monthly Repayment 'Yes' tick Failed")    
        if Excel_Dictionary1['CPF For Monthly Repayment Amount'] != None:
            cpf_Monthly_Installment = str(addComma(Excel_Dictionary1['CPF For Monthly Repayment Amount']))
            if cpf_Monthly_Installment == outputdict['Main_Account_Holder_CPF_per_month']:
                results.append("CPF For Monthly Repayment Amount Passed") 
            else:
                results.append("CPF For Monthly Repayment Amount Failed")
        else:
            results.append("CPF For Monthly Repayment Amount Data not Available in UI or Input File")               
    elif Excel_Dictionary1['CPF for Monthly Repayment?'] == 'No':
        if outputdict['Self_CPF_for_Monthly_Repayment_Check_No'] == 'Yes':
            results.append("CPF for Monthly Repayment 'No' tick Passed")
        else:
            results.append("CPF for Monthly Repayment 'No' tick Failed")    
    else:
        results.append("CPF for Monthly Repayment? Data is Incorrect")        
else:
    results.append("CPF for Monthly Repayment? Data not Available in UI or Input File")        

              


########################################################## Validation Ending #############################################################

#Preparing Results File 
writesheet = openpyxl.Workbook()
Wsheet = writesheet.active
Wsheet.column_dimensions['A'].width = float(100)
if len(results) > 0:
    j = 1
    for i in range(len(results)):
        Wsheet.cell(j,1).value = results[i]
        j+=1

writesheet.save("D:\\Automation Testing\\MayBank\\Loan\\Results\\loan_new.xlsx")       