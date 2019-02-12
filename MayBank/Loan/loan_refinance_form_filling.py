from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)

path = "D:\\Automation Testing\\MayBank\\Loan\\loan_refinance_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
for i in range(1,80):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value
    
driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_id("AddLoanApplication").click()
driver.find_element_by_id("homeBuyerSearchInputId").send_keys("A K S")
driver.find_element_by_id("homeBuyerSearchButtonId").click()
driver.find_element_by_id("selectedClientId").click()
driver.find_element_by_xpath('//*[@id="homeBuyerDetailsModelDiv"]/div/div/div[2]/div[3]/div[2]/button').click()
driver.find_element_by_id("singleApplicantID").click()

if Excel_Dictionary1['application type 1'] != None:
    if Excel_Dictionary1['application type 1'] == 'Private':
        driver.find_element_by_id("privateFlatID").click()
    elif Excel_Dictionary1['application type 1'] == 'Executive Condo':
        driver.find_element_by_id("executiveCondoID").click()
    elif Excel_Dictionary1['application type 1'] == 'HDB':
        driver.find_element_by_id('hdbFlatID').click()        
    else:
        driver.find_element_by_id("privateFlatID").click()
else:
    driver.find_element_by_id("privateFlatID").click() 

driver.find_element_by_id("refinanceID").click()
driver.find_element_by_xpath("(.//*[normalize-space(text()) and normalize-space(.)='Cancel'])[2]/following::button[2]").click()
time.sleep(4)

if Excel_Dictionary1['Salutation'] != None:
    driver.find_element_by_xpath('//*[@id="salutation"]').send_keys(Excel_Dictionary1['Salutation'])
if Excel_Dictionary1['Full Name as in NRIC/PASSPORT'] != None:     
    driver.find_element_by_xpath('//*[@id="fullName"]').clear()
    driver.find_element_by_xpath('//*[@id="fullName"]').send_keys(Excel_Dictionary1['Full Name as in NRIC/PASSPORT'])     
if Excel_Dictionary1['Email Address'] != None:   
    driver.find_element_by_name('M0031').send_keys(Excel_Dictionary1['Email Address']) 

driver.find_element_by_xpath('//*[@id="loan"]').click()
time.sleep(4)   

if Excel_Dictionary1['Valuation Firm'] != None:
    driver.find_element_by_name('M0275').send_keys(Excel_Dictionary1['Valuation Firm'])
if Excel_Dictionary1['Name of Valuer'] != None:
    driver.find_element_by_name('M0531').send_keys(Excel_Dictionary1['Name of Valuer'])
if Excel_Dictionary1['Indicative Valuation'] != None:
    driver.find_element_by_name('M0273').send_keys(Excel_Dictionary1['Indicative Valuation'])
if Excel_Dictionary1['Date of Valuation'] != None:
    driver.find_element_by_name('M0274').send_keys(Excel_Dictionary1['Date of Valuation']) 

if Excel_Dictionary1['CPF for Monthly Repayment?'] != None:
    if Excel_Dictionary1['CPF for Monthly Repayment?'] == 'Yes':
        driver.find_element_by_id('isCPFForMonthlyRepaymentYes').click()
        time.sleep(1)
        if Excel_Dictionary1['CPF For Monthly Repayment Amount'] != None:
            driver.find_element_by_id('CPFForMonthlyRepaymentValue').send_keys(Excel_Dictionary1['CPF For Monthly Repayment Amount'])
            time.sleep(1)
    elif Excel_Dictionary1['CPF for Monthly Repayment?'] == 'No':
        driver.find_element_by_id('isCPFForMonthlyRepaymentNo').click()
        time.sleep(1)

if Excel_Dictionary1['CPF for Legal Fees?'] != None:
    if Excel_Dictionary1['CPF for Legal Fees?'] == 'Yes':
        driver.find_element_by_id('isCPFForLegalFeesYes').click()
        time.sleep(1)
        if Excel_Dictionary1['CPF for Legal Fees Amount'] != None:
            driver.find_element_by_id('CPFForLegalFeesValue').send_keys(Excel_Dictionary1['CPF for Legal Fees Amount'])  
            time.sleep(1)
    elif Excel_Dictionary1['CPF for Legal Fees?'] == 'No':
        driver.find_element_by_id('isCPFForLegalFeesNo').click()
        time.sleep(1)

if Excel_Dictionary1['Name of Financier'] != None:
    driver.find_element_by_name('M0303').send_keys(Excel_Dictionary1['Name of Financier'])           

if Excel_Dictionary1['Outstanding Housing Loan Amount'] != None:
    driver.find_element_by_name('M0304').send_keys(Excel_Dictionary1['Outstanding Housing Loan Amount'])   

if Excel_Dictionary1['Housing Current Interest Rate'] != None:
    driver.find_element_by_name('M0316').send_keys(Excel_Dictionary1['Housing Current Interest Rate'])   

time.sleep(1)
if Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] != None:
    if Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] == 'Yes':
        driver.find_element_by_id('option2').click()  
        time.sleep(1)
        if Excel_Dictionary1['Amount Undisbursed'] != None:
            driver.find_element_by_name('M0310').send_keys(Excel_Dictionary1['Amount Undisbursed'])
    elif Excel_Dictionary1['Is there any Undisbursed Amount in Housing Loan? '] == 'No':
        driver.find_element_by_id('option3').click()      
        time.sleep(1)   

if Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] != None:
    if Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] == 'Yes':
        driver.find_element_by_id('option2').click() 
        time.sleep(1)
        if Excel_Dictionary1['Amount you would like to Pre-pay down'] != None:
            driver.find_element_by_name('M0306').send_keys(Excel_Dictionary1['Amount you would like to Pre-pay down'])     
    elif Excel_Dictionary1['Pre-Payment on the Housing Loan before Refinancing? '] == 'No':
        driver.find_element_by_id('option3').click()      
        time.sleep(1)    

if Excel_Dictionary1['Original Loan Tenure'] != None:
    driver.find_element_by_name('M0311').send_keys(Excel_Dictionary1['Original Loan Tenure']) 

if Excel_Dictionary1['Loan Disbursement Date'] != None:
    driver.find_element_by_name('M0315').send_keys(Excel_Dictionary1['Loan Disbursement Date'])

if Excel_Dictionary1['New Loan Tenure'] != None:
    driver.find_element_by_name('M0480').send_keys(Excel_Dictionary1['New Loan Tenure']) 

time.sleep(1)
if (Excel_Dictionary1['application type 1'] != None) and (Excel_Dictionary1['application type 1'] != 'HDB'):
    if Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] != None:
        if Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == 'Yes':
            driver.find_element_by_id('option2').click()
            time.sleep(1)
            if Excel_Dictionary1['Outstanding Equity Loan Amount'] != None:
                driver.find_element_by_name('M0305').send_keys(Excel_Dictionary1['Outstanding Equity Loan Amount'])
            if Excel_Dictionary1['Outstanding Current Interest Rate'] != None:
                driver.find_element_by_name('M0481').send_keys(Excel_Dictionary1['Outstanding Current Interest Rate'])    
        elif Excel_Dictionary1['Do you have an Outstanding Equity Loan? '] == 'No':
            driver.find_element_by_id('option3').click()    
            time.sleep(1)
    if Excel_Dictionary1['Pre-Payment on the Equity Loan before Refinancing?'] != None:
        if Excel_Dictionary1['Pre-Payment on the Equity Loan before Refinancing?'] == 'Yes':
            driver.find_element_by_id('option2').click()
            time.sleep(1)
            if Excel_Dictionary1['Equity Amount you would like to Pre-pay down'] != None:
                driver.find_element_by_name('M0307').send_keys(Excel_Dictionary1['Equity Amount you would like to Pre-pay down'])        
        elif Excel_Dictionary1['Pre-Payment on the Equity Loan before Refinancing?'] == 'No':
            driver.find_element_by_id('option3').click()   
            time.sleep(1)   
    if Excel_Dictionary1['New Equity Loan Required?'] != None:
        if Excel_Dictionary1['New Equity Loan Required?'] == 'Yes':
            driver.find_element_by_id('isAdditionalEquityLoanRequired2Yes').click()
            time.sleep(1)
            if Excel_Dictionary1['New Equity Loan for Personal Use'] != None:
                driver.find_element_by_name('M0332').send_keys(Excel_Dictionary1['New Equity Loan for Personal Use'])    
            if Excel_Dictionary1['Equity Loan Tenure'] != None:
                driver.find_element_by_name('M0333').send_keys(Excel_Dictionary1['Equity Loan Tenure']) 
        elif Excel_Dictionary1['New Equity Loan Required?'] == 'No':
            driver.find_element_by_id('isAdditionalEquityLoanRequired2No').click()     
            time.sleep(1)          

time.sleep(1)
if Excel_Dictionary1['CPF Utilised- Main'] != None:
    driver.find_element_by_id('mainAccountHolderTotalCPFWithdrawnToDate').send_keys(Excel_Dictionary1['CPF Utilised- Main'])

if Excel_Dictionary1['CPF Utilised- Joint 1'] != None:
    driver.find_element_by_id('refinanceCPFUtilisedJoint1').send_keys(Excel_Dictionary1['CPF Utilised- Joint 1'])

time.sleep(1)
if Excel_Dictionary1['CPF Utilised- Joint 2'] != None:
    driver.find_element_by_id('refinanceCPFUtilisedJoint2').send_keys(Excel_Dictionary1['CPF Utilised- Joint 2'])

if Excel_Dictionary1['CPF Utilised- Joint 3'] != None:
    driver.find_element_by_id('refinanceCPFUtilisedJoint3').send_keys(Excel_Dictionary1['CPF Utilised- Joint 3'])

time.sleep(1)
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
driver.find_element_by_class_name('loan-submit-button').click()
time.sleep(3)
driver.find_element_by_xpath('//*[@id="checkbox-select"]').click()
driver.find_element_by_xpath('//*[@id="maybank-deselected"]').click()
driver.find_element_by_xpath('//*[@id="downloadForms"]').click()
time.sleep(70)
driver.quit()
print("Form Filling Done")            
