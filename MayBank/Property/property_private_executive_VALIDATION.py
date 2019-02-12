from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time


path = "D:\\Automation Testing\\MayBank\\Property\\property_private_executive_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
for i in range(1,36):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)
driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_xpath('//*[@id="loansTable"]/tbody/tr[1]/td[8]/div/button').click()
time.sleep(3)
element = driver.find_element_by_xpath('//*[@id="applicationIdField"]').text
driver.quit()

Updated_PDF = "D:\\vaibhav\\Vaibhav Clone\\mortgage-webapp\\"+element+"_May Bank_Updated1.pdf"
try:
    pdf = PyPDF2.PdfFileReader(Updated_PDF)
except:
    error = "PDF file Not Found."
    print(error)

pdfData = pdf.getFields()
outputdict={}

# Creating a dictionary from pdf data    
for data in pdfData:
    info = pdfData[data]
    if "/V" in info:
        # if info["/V"] == "/" or info["/V"] == "" :
        #     pass
        # else :
        tag=info["/T"]
        value=info["/V"]
        # tag=tag.replace("_"," ")
        if "/" in value:
            value=value.replace("/", "")
        outputdict[tag]=value

def addComma(inputValue):
    if 'S$' in str(inputValue):
        inputValue = inputValue.replace('S$',"")
    if ',' in str(inputValue):
        return str(inputValue)
    if '.' in str(inputValue):
        parts = str(inputValue).split(".")
        intPart = parts[0]
        floatPart = parts[1]
        intPart = "{:,}".format(int(float(intPart)))
        if floatPart != '0' and floatPart != '00' :
            retstr= str(intPart) + "." + str(floatPart) 
        else :
            retstr= str(intPart)
        return retstr
    elif inputValue :
        try:
            return "{:,}".format(int(inputValue))
        except:
            parts = inputValue.split(".")
            intPart = parts[0]
            floatPart = parts[1]
            intPart = "{:,}".format(int(float(intPart)))
            return intPart + "." + floatPart
    else:
        return inputValue

def remove_dot_zero(input):
    if(float(input) % 1 == 0):
        return int(input)
    else:
        return input

results = []

########################################################## Validation Starting #############################################################
if Excel_Dictionary1['Property Name'] != None:
    if (Excel_Dictionary1['Property Name'] == outputdict['Self_Property_Estate_Name']):
        results.append("Property Name in Details of Property to be Mortgaged Passed")
    else:
        results.append("Property Name in Details of Property to be Mortgaged Failed")
else:
    results.append("Property Name Data not Available in UI or Input file or Data is Incorrect") 

if Excel_Dictionary1['Zip / Postal Code'] != None:
    if (Excel_Dictionary1['Zip / Postal Code'] == outputdict['Self_Property_Postal_Code']):
        results.append("Property Postal Code in Details of Property to be Mortgaged Passed")  
    else:
        results.append("Property Postal Code in Details of Property to be Mortgaged Passed")
else:
    results.append("Property Postal Code Data not Available in UI or Input file or Data is Incorrect")                         

if Excel_Dictionary1['Unit No'] != None:
    if (str(Excel_Dictionary1['Unit No']) == outputdict['Self_Unit_No']):
        results.append("Property Unit No. in Details of Property to be Mortgaged Passed")
    else:
        results.append("Property Unit No. in Details of Property to be Mortgaged Failed")
else:
    results.append("Property Unit No. Data not Available in UI or Input file or Data is Incorrect")

property_Address = (Excel_Dictionary1['Property Address']).lower()
country = (Excel_Dictionary1['Country']).lower()
# if Excel_Dictionary1['Property Address'] != None:
#     property_Address = (Excel_Dictionary1['Property Address']).lower()
# if Excel_Dictionary1['Country'] != None:
#     country = (Excel_Dictionary1['Country']).lower()    

property_Address_Final = property_Address +", "+ country
if (property_Address_Final == outputdict['Self_Street_Name']):
    results.append("Property Name in Details of Property to be Mortgaged Passed")
else:
    results.append("Property Name in Details of Property to be Mortgaged Failed")    

if (Excel_Dictionary1['application type 2'] == 'New'):
    if Excel_Dictionary1['Have you purchase the property? '] != None:
        if (Excel_Dictionary1['Have you purchase the property? '] == 'No') and (outputdict['Self_OTP_Granted_No'] == 'Yes'):
            results.append("OTP Granted in Financing Requirements 'No' Passed")
        elif (Excel_Dictionary1['Have you purchase the property? '] == 'Yes'):
            if (Excel_Dictionary1['Date of Purchase'] == None) and (outputdict['Self_OTP_Granted_No'] == 'Yes'):
                results.append("OTP Granted in Financing Requirements For 'Yes' with No Date Passed")
            elif (Excel_Dictionary1['Date of Purchase'] != None) and (outputdict['Self_OTP_Granted_Yes'] == 'Yes'):
                outputDate = (outputdict['Self_Date_of_OTP']).replace("-","")
                if (str(Excel_Dictionary1['Date of Purchase']) == outputDate):
                    results.append("OTP Granted in Financing Requirements For 'Yes' with Date Passed")
        else:
            results.append("Have You Purchase the property/OTP Granted in Financing Requirements Failed")      
    else:
        results.append("Have You Purchase the property Data not Available in UI or Input File or Data is Incorrect")

if Excel_Dictionary1['application type 1'] == 'Private':
    if (Excel_Dictionary1['Type of Property'] != None) and (Excel_Dictionary1['Type of Property'] == 'Private Landed' or Excel_Dictionary1['Type of Property'] == 'Private Non-Landed'): 
        if (Excel_Dictionary1['Private Landed'] == 'Bungalow') and (outputdict['Self_Property_Bungalow'] == 'Yes'):
            results.append("Private Landed 'Bungalow' tick Passed")
        elif (Excel_Dictionary1['Private Landed'] == 'Semi-D') and (outputdict['Self_Property_Semi_D'] == 'Yes'):
            results.append("Private Landed 'Semi-D' tick Passed")    
        elif (Excel_Dictionary1['Private Landed'] == 'Corner Terrace') and (outputdict['Self_Property_Intermediate_Terrace'] == 'Yes'):
            results.append("Private Landed 'Corner Terrace' tick Passed")
            if "____" in outputdict['Self_Property_Intermediate_Cross_Line']:
                results.append("Private Landed  Strike for 'Corner Terrace' Passed") 
        elif (Excel_Dictionary1['Private Landed'] == 'Intermediate Terrace') and (outputdict['Self_Property_Intermediate_Terrace'] == 'Yes'):
            results.append("Private Landed 'Intermediate Terrace' tick Passed")
            if "____" in outputdict['Self_Property_CornerTerrace_Cross_Line']:
                results.append("Private Landed Strike for 'Intermediate Terrace' Passed")
        elif (Excel_Dictionary1['Private Landed'] == 'Others') and (outputdict['Self_Property_Others'] == 'Yes') and (Excel_Dictionary1['Private Landed Others'] != None) and (Excel_Dictionary1['Private Landed Others'] == outputdict['Self_Property_OthersText']):
            results.append("Private Landed 'Others' with tick and Data Passed")
        elif (Excel_Dictionary1['Private Landed'] == 'Others') and (outputdict['Self_Property_Others'] == 'Yes') and (Excel_Dictionary1['Private Landed Others'] == None) and (outputdict['Self_Property_OthersText'] == ''):
            results.append("Private Landed 'Others' with tick and without Data Passed") 

        elif (Excel_Dictionary1['Private Non-Landed'] == 'Cluster/Townhouse') and (outputdict['Self_Property_Others'] == 'Yes') and (outputdict['Self_Property_OthersText'] == 'Cluster/Townhouse'):
            results.append("Private Non-Landed 'Cluster/Townhouse' tick and text Passed")
        elif (Excel_Dictionary1['Private Non-Landed'] == 'HUDC') and (outputdict['Self_Property_HUDC_Privatised'] == 'Yes'):
            results.append("Private Non-Landed 'HUDC' tick Passed")  
        elif (Excel_Dictionary1['Private Non-Landed'] == 'Condominium') and (outputdict['Self_Property_Others'] == 'Yes') and (outputdict['Self_Property_OthersText'] == 'Condominium'):
            results.append("Private Non-Landed 'Condominium' tick and text Passed")  
        elif (Excel_Dictionary1['Private Non-Landed'] == 'Executive Condo') and (outputdict['Self_Property_Executive_Condominium'] == 'Yes'):
            results.append("Private Non-Landed 'Executive Condo' tick Passed")      
        elif (Excel_Dictionary1['Private Non-Landed'] == 'Apartment') and (outputdict['Self_Property_Apartment'] == 'Yes'):
            results.append("Private Non-Landed 'Apartment' tick Passed") 
        elif (Excel_Dictionary1['Private Non-Landed'] == 'Others') and (outputdict['Self_Property_Others'] == 'Yes') and (Excel_Dictionary1['Private Non-Landed Others'] != None) and (Excel_Dictionary1['Private Non-Landed Others'] == outputdict['Self_Property_OthersText']):
            results.append("Private Non-Landed 'Others' with tick and Data Passed")
        elif (Excel_Dictionary1['Private Non-Landed'] == 'Others') and (outputdict['Self_Property_Others'] == 'Yes') and (Excel_Dictionary1['Private Non-Landed Others'] == None) and (outputdict['Self_Property_OthersText'] == ''):
            results.append("Private Non-Landed 'Others' with tick and without Data Passed") 
        else:
            results.append("Type of Property Failed")
    else:
        results.append("Type of Property Data not Available in UI or Input File or Data is Incorrect")    

if (Excel_Dictionary1['Usage of Property'] != None):
    if (Excel_Dictionary1['Usage of Property'] == "Owner's Occupation") and (outputdict['Self_Pupose_Owner'] == 'Yes'):
        results.append("Usage of Property 'Owner's Occupation' tick Passed")
    elif (Excel_Dictionary1['Usage of Property'] == 'Investment') and (outputdict['Self_Pupose_Investment'] == 'Yes'):
        results.append("Usage of Property 'Investment' tick Passed")
        if (Excel_Dictionary1['Is Vacant or Occupied'] != None):
            if (Excel_Dictionary1['Is Vacant or Occupied'] == 'Vacant') and (outputdict['Self_Vacant'] == 'Yes'):
                results.append("Is Vacant or Occupied 'Vacant' tick Passed")
                if (Excel_Dictionary1['Expected Rental Amount'] != None):
                    if (Excel_Dictionary1['Expected Rental Monthly/Yearly'] != None):
                        if (Excel_Dictionary1['Expected Rental Monthly/Yearly'] == 'Annual'):
                            vacant_Monthly = str(addComma(format(float(Excel_Dictionary1['Expected Rental Amount']/12),'.2f')))
                            if vacant_Monthly == outputdict['Self_Expected_Rental']:
                                results.append("Vacant Monthly Calculation is Passed")
                            else:
                                results.append("Vacant Monthly Calculation is Failed")    
                        elif (Excel_Dictionary1['Expected Rental Monthly/Yearly'] == 'Monthly'):
                            vacant_Monthly = str(addComma(Excel_Dictionary1['Expected Rental Amount'])) 
                            if vacant_Monthly == outputdict['Self_Expected_Rental']:
                                results.append("Vacant Monthly Calculation is Passed")
                            else:
                                results.append("Vacant Monthly Calculation is Failed")     
                        else:
                            results.append("Expected Rental Monthly/Yearly Data is Incorrect")        
                    else:
                        results("Expected Rental Monthly/Yearly Data not Available in UI or Input File")             
                else:
                    results.append("Expected Rental Amount Data not Available in UI or Input File")                      
            elif (Excel_Dictionary1['Is Vacant or Occupied'] == 'Occupied') and (outputdict['Self_Occupied'] == 'Yes'):
                results.append("Is Vacant or Occupied 'Occupied' tick Passed")
                if (Excel_Dictionary1['Actual Rental Amount'] != None):
                    if (Excel_Dictionary1['Actual Rental Monthly/Yearly'] != None):
                        if (Excel_Dictionary1['Actual Rental Monthly/Yearly'] == 'Annual'):
                            occupied_Monthly = str(addComma(format(float(Excel_Dictionary1['Actual Rental Amount']/12),'.2f')))
                            if occupied_Monthly == outputdict['Self_Actual_Rental']:
                                results.append("Occupied Monthly Calculation is Passed")
                            else:
                                results.append("Occupied Monthly Calculation is Failed")    
                        elif (Excel_Dictionary1['Actual Rental Monthly/Yearly'] == 'Monthly'):
                            occupied_Monthly = str(addComma(Excel_Dictionary1['Actual Rental Amount']))
                            if occupied_Monthly == outputdict['Self_Actual_Rental']:
                                results.append("Occupied Monthly Calculation is Passed")
                            else:
                                results.append("Occupied Monthly Calculation is Failed")    
                        else:
                            results.append("Actual Rental Monthly/Yearly Data is Incorrect")
                    else:
                        results.append("Actual Rental Monthly/Yearly Data not Available in UI or Input File") 
                else:
                    results.append("Actual Rental Amount Data not Available in UI or Input File")                       
                if (Excel_Dictionary1['Rental Expires On'] != None):
                    rental_expires = (outputdict['Self_Rental_Expires_On']).replace("-","")
                    if (str(Excel_Dictionary1['Rental Expires On']) == rental_expires):
                        results.append("Occupied Rental Expires On Passed")     
            else:
                results.append("Is Vacant or Occupied Data is Incorrect")            
        else:
            results.append("Is Vacant or Occupied Data not Available in UI or Input File")   
    else:
        results.append("Usage of Property Data is Incorrect")                       
else:
    results.append("Usage of Property Data not Available in UI or Input File")

if (Excel_Dictionary1['Tenure of Property'] != None):
    if (Excel_Dictionary1['Tenure of Property'] == 'Freehold') and (outputdict['Self_Tenure_of_Property_Freehold_Check'] == 'Yes'):
        results.append("Tenure of Property 'Freehold' Passed")
    elif (Excel_Dictionary1['Tenure of Property'] == 999) and (outputdict['Self_Tenure_of_Property_LeaseHold_Check'] == 'Yes') and (str(Excel_Dictionary1['Tenure of Property']) == outputdict['Self_Tenure_of_Property_LeaseHold_Text']):
        results.append("Tenure of Property '999' Passed")    
    elif (Excel_Dictionary1['Tenure of Property'] == 199) and (outputdict['Self_Tenure_of_Property_LeaseHold_Check'] == 'Yes') and (str(Excel_Dictionary1['Tenure of Property']) == outputdict['Self_Tenure_of_Property_LeaseHold_Text']):
        results.append("Tenure of Property '199' Passed")
    elif (Excel_Dictionary1['Tenure of Property'] == 99) and (outputdict['Self_Tenure_of_Property_LeaseHold_Check'] == 'Yes') and (str(Excel_Dictionary1['Tenure of Property']) == outputdict['Self_Tenure_of_Property_LeaseHold_Text']):
        results.append("Tenure of Property '99' Passed")
    elif (Excel_Dictionary1['Tenure of Property'] == 60) and (outputdict['Self_Tenure_of_Property_LeaseHold_Check'] == 'Yes') and (str(Excel_Dictionary1['Tenure of Property']) == outputdict['Self_Tenure_of_Property_LeaseHold_Text']):
        results.append("Tenure of Property '60' Passed")
    elif (Excel_Dictionary1['Tenure of Property'] == 'Others'): 
        if (Excel_Dictionary1['Tenure of Property Others'] != None): 
            if (outputdict['Self_Tenure_of_Property_LeaseHold_Check'] == 'Yes') and (str(Excel_Dictionary1['Tenure of Property Others']) == outputdict['Self_Tenure_of_Property_LeaseHold_Text']):
                results.append("Tenure of Property 'Others' with tick and Data Passed") 
        elif (outputdict['Self_Tenure_of_Property_LeaseHold_Check'] == 'Yes') and (Excel_Dictionary1['Tenure of Property Others'] == None) and (outputdict['Self_Tenure_of_Property_LeaseHold_Text'] == ''):
            results.append("Tenure of Property 'Others' with tick without Data Passed")               

if (Excel_Dictionary1['Tenure of Property'] != None) and (Excel_Dictionary1['Tenure of Property'] != 'Freehold'):
    if (Excel_Dictionary1['Tenure w.e.f'] != None):
        if (str(Excel_Dictionary1['Tenure w.e.f']) == (outputdict['Self_Tenure_WEF']).replace("-","")):
            results.append("Tenure w.e.f Passed")
        else:
            results.append("Tenure w.e.f Failed")    
    else:
        results.append("Tenure w.e.f Data not Available in UI or Input File")
else:
    results.append("Tenure of Property Data not Available in UI or Input File or Tenure of Property may be Freehold")        

if (Excel_Dictionary1['Property Status'] != None):
    if (Excel_Dictionary1['Property Status'] == 'Completed'):
        if (outputdict['Self_Property_Status_Completed'] == 'Yes'):
            results.append("Property Status 'Completed' tick Passed")
        else:
            results.append("Property Status 'Completed' tick Failed")    
        if (Excel_Dictionary1['Year Built'] != None) and (str(Excel_Dictionary1['Year Built']) == outputdict['Self_Property_Years_BuiltText']):
            results.append("Property Status 'Completed' with 'Year Built' Data Passed") 
        elif (Excel_Dictionary1['Year Built'] == None) and (outputdict['Self_Property_Years_BuiltText'] == ''):
            results.append("Property Status 'Completed' without 'Year Built' Data Passed")
        else:
            results.append("Property Status 'Completed' Failed")            

    elif (Excel_Dictionary1['Property Status'] == 'Under Construction'):
        if (outputdict['Self_Property_Status_Under_Construction'] == 'Yes'):
            results.append("Property Status 'Under Construction' tick Passed")
        else:
            results.append("Property Status 'Under Construction' tick Failed")    
        if (Excel_Dictionary1['Expected TOP (Date)'] != None) and (str(Excel_Dictionary1['Expected TOP (Date)']) == outputdict['Self_Property_Expected_TOPText']):
            results.append("Property Status 'Under Construction' with 'Expected TOP (Date)' Data Passed") 
        elif (Excel_Dictionary1['Expected TOP (Date)'] == None) and (outputdict['Self_Property_Expected_TOPText'] == ''):
            results.append("Property Status 'Under Construction' without 'Expected TOP (Date)' Data Passed")     
        else:
            results.append("Property Status 'Under Construction' Failed")      
    else:
        results.append("Property Status Data is Incorrect")            
else:
    results.append("Property Status Data not Available in UI or Input File")

if (Excel_Dictionary1['Property Size - Built In (sqft)'] != None):
    if (str(Excel_Dictionary1['Property Size - Built In (sqft)']) == outputdict['Self_BuiltIn_Area']):
        results.append("Property Size - Built In (sqft) Passed")
        if '___' in outputdict['Self_BuiltIn_Area_Sq_m_Cut']:
            results.append("Property Size - Built In (sqft) strike Passed")
        else:
            results.append("Property Size - Built In (sqft) strike Failed")    
    else:
        results.append("Property Size - Built In (sqft) Failed")    
else:
    results.append("Property Size - Built In (sqft) Data not Available in UI or Input File")

if (Excel_Dictionary1['Property Size - Land(sqft)'] != None):
    if(str(Excel_Dictionary1['Property Size - Land(sqft)']) == outputdict['Self_Land_Area']):
        results.append("Property Size - Land(sqft) Passed")
        if '___' in outputdict['Self_Land_Area_Sq_m_Cut']:
            results.append("Property Size - Land(sqft) strike Passed")
        else:
            results.append("Property Size - Land(sqft) strike Failed")    
    else:
        results.append("Property Size - Land(sqft) Failed")
else:
    results.append("Property Size - Land(sqft) Data not Available in UI or Input File")        

if (Excel_Dictionary1['Number of Storeys'] != None):
    if (str(Excel_Dictionary1['Number of Storeys']) == outputdict['Number_of_Stories']):
        results.append("Number of Storeys Passed")
    else:
        results.append("Number of Storeys Failed")    
else:
    results.append("Number of Storeys Data not Available in UI or Input File")






########################################################## Validation Ending #############################################################

#Preparing Results File 
writesheet = openpyxl.Workbook()
Wsheet = writesheet.active
Wsheet.column_dimensions['A'].width = float(100)
if len(results) > 0:
    j = 1
    for i in range(len(results)):
        Wsheet.cell(j,1).value = results[i]
        j+=1

writesheet.save("D:\\Automation Testing\\MayBank\\Property\\Results\\private_executive_results.xlsx")   
