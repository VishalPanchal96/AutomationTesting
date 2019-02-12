from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import PyPDF2
import openpyxl
import os
import time

driver = webdriver.Chrome('D:\\Automation Testing\\chromedriver.exe')
driver.maximize_window()
driver.get("http://localhost:8000/")
driver.implicitly_wait(5000)

path = "D:\\Automation Testing\\MayBank\\Property\\property_hdb_INPUT.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
Excel_Dictionary1 = {}
for i in range(1,26):
    Excel_Dictionary1[sheet.cell(i,1).value] = sheet.cell(i,2).value

driver.find_element_by_id("userName").send_keys("abhinay.k@sankeysolutions.com")
driver.find_element_by_id("password").send_keys("123")
driver.find_element_by_class_name("login-button-label").click()

driver.find_element_by_id("AddLoanApplication").click()
driver.find_element_by_id("homeBuyerSearchInputId").send_keys("A K S")
driver.find_element_by_id("homeBuyerSearchButtonId").click()
driver.find_element_by_id("selectedClientId").click()
driver.find_element_by_xpath('//*[@id="homeBuyerDetailsModelDiv"]/div/div/div[2]/div[3]/div[2]/button').click()
driver.find_element_by_id("singleApplicantID").click()
driver.find_element_by_id('hdbFlatID').click()

if Excel_Dictionary1['application type 2'] != None:
    if Excel_Dictionary1['application type 2'] == 'New':
        driver.find_element_by_id("newHomeLoanImage").click()
    elif Excel_Dictionary1['application type 2'] == 'Refinance':
        driver.find_element_by_id("refinanceID").click()    
    else:
        driver.find_element_by_id("newHomeLoanImage").click()
else:
    driver.find_element_by_id("newHomeLoanImage").click()

driver.find_element_by_xpath("(.//*[normalize-space(text()) and normalize-space(.)='Cancel'])[2]/following::button[2]").click()

if Excel_Dictionary1['Salutation'] != None:
    driver.find_element_by_xpath('//*[@id="salutation"]').send_keys(Excel_Dictionary1['Salutation'])
if Excel_Dictionary1['Full Name as in NRIC/PASSPORT'] != None:     
    driver.find_element_by_xpath('//*[@id="fullName"]').clear()
    driver.find_element_by_xpath('//*[@id="fullName"]').send_keys(Excel_Dictionary1['Full Name as in NRIC/PASSPORT'])     
if Excel_Dictionary1['Email Address'] != None:   
    driver.find_element_by_name('M0031').send_keys(Excel_Dictionary1['Email Address']) 

driver.find_element_by_xpath('//*[@id="propertyDetailsLabel"]').click()
time.sleep(4)

if Excel_Dictionary1['Property Name'] != None:
    driver.find_element_by_name('M0183').send_keys(Excel_Dictionary1['Property Name'])    
if Excel_Dictionary1['Country'] != None:
    driver.find_element_by_name('M0507').send_keys(Excel_Dictionary1['Country'])
if Excel_Dictionary1['Zip / Postal Code'] != None:
    driver.find_element_by_name('M0186').send_keys(Excel_Dictionary1['Zip / Postal Code'])
if Excel_Dictionary1['Property Address'] != None:
    driver.find_element_by_name('M0182').send_keys(Excel_Dictionary1['Property Address'])  
if Excel_Dictionary1['Unit No'] != None:
    driver.find_element_by_name('M0185').send_keys(Excel_Dictionary1['Unit No'])    

driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
time.sleep(2)        
if Excel_Dictionary1['application type 2'] == 'New':
    if Excel_Dictionary1['Have you purchase the property? '] == 'Yes':
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="isAlreadyPurchased"]').click()
        time.sleep(1)
        if Excel_Dictionary1['Date of Purchase'] != None:
            driver.find_element_by_name('M0173').send_keys(Excel_Dictionary1['Date of Purchase'])
    elif Excel_Dictionary1['Have you purchase the property? '] == 'No':
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="isAlreadyPurchasedNo"]').click()   

if Excel_Dictionary1['Type of Property HDB'] != None:
    driver.find_element_by_name('M0177').send_keys(Excel_Dictionary1['Type of Property HDB'])

if Excel_Dictionary1['Usage of Property'] != None:
    driver.find_element_by_name('M0176').send_keys(Excel_Dictionary1['Usage of Property'])
    if Excel_Dictionary1['Usage of Property'] == 'Investment':
        if Excel_Dictionary1['Is Vacant or Occupied'] != None:
            if Excel_Dictionary1['Is Vacant or Occupied'] == 'Vacant':
                time.sleep(1)
                driver.find_element_by_id('vacant').click()
                if Excel_Dictionary1['Expected Rental Amount'] != None:
                    driver.find_element_by_name('M0189').send_keys(Excel_Dictionary1['Expected Rental Amount'])
                    if Excel_Dictionary1['Expected Rental Monthly/Yearly'] != None:
                        driver.find_element_by_name('periodDropdown').send_keys(Excel_Dictionary1['Expected Rental Monthly/Yearly'])
                        time.sleep(1)
            elif Excel_Dictionary1['Is Vacant or Occupied'] == 'Occupied':
                time.sleep(1)
                driver.find_element_by_id('occupied').click()
                if Excel_Dictionary1['Actual Rental Amount'] != None:
                    driver.find_element_by_name('M0190').send_keys(Excel_Dictionary1['Actual Rental Amount'])
                    if Excel_Dictionary1['Actual Rental Monthly/Yearly'] != None:
                        driver.find_element_by_name('expectedPeriodDropdown').send_keys(Excel_Dictionary1['Actual Rental Monthly/Yearly']) 
                        time.sleep(1)
                if Excel_Dictionary1['Rental Expires On'] != None:
                    driver.find_element_by_name('M0191').send_keys(Excel_Dictionary1['Rental Expires On'])              

if Excel_Dictionary1['Tenure w.e.f'] != None:                           
    driver.find_element_by_name('M0187').send_keys(Excel_Dictionary1['Tenure w.e.f'])

if Excel_Dictionary1['Property Status'] != None:
    driver.find_element_by_name('M0180').send_keys(Excel_Dictionary1['Property Status'])
    time.sleep(1)
    if Excel_Dictionary1['Property Status'] == 'Completed':
        if Excel_Dictionary1['Year Built'] != None:
            driver.find_element_by_id('propertyStatusYearBuilt').send_keys(Excel_Dictionary1['Year Built'])
    elif Excel_Dictionary1['Property Status'] == 'Under Construction':
        if Excel_Dictionary1['Expected TOP (Date)'] != None:
            driver.find_element_by_id('propertyStatusExpectedTOP').send_keys(Excel_Dictionary1['Expected TOP (Date)'])

if Excel_Dictionary1['Property Size - Built In (sqft)'] != None:
    driver.find_element_by_name('M0163').send_keys(Excel_Dictionary1['Property Size - Built In (sqft)'])           

time.sleep(3)
driver.find_element_by_xpath('//*[@id="saveAndNext"]/button').click()
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
time.sleep(4)
driver.find_element_by_class_name('loan-submit-button').click()
time.sleep(3)
driver.find_element_by_xpath('//*[@id="checkbox-select"]').click()
driver.find_element_by_xpath('//*[@id="maybank-deselected"]').click()
driver.find_element_by_xpath('//*[@id="downloadForms"]').click()
time.sleep(70)
driver.quit()
print("Form Filling Done")